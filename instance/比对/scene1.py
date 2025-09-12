# -*- coding: utf-8 -*-
# *********************************************************************
# Author        : John
# Date          : 2025-9-12
# Description   : 场景一：CSV / Excel 结构化数据字段级比对
# 解决思路先把两张表按主键对齐，再逐列标记差异，最后导出差异报告。
# 　　· 对齐：使用商品 ID 作为唯一键，避免错位。
# 　　· 标记：为每列生成布尔型“变更”标志，True 即差异。
# 　　· 导出：将含 True 的行写入新 Excel，直接送交业务方。
# *********************************************************************

import pandas as pd

def excel_diff_only_red(file1, file2, key_column, compare_columns=None):
    """
    Excel文件差异比对核心实现, 差异数据用红色背景，不显示无差异的列
    :param file1: 基准文件路径
    :param file2: 比对文件路径
    :param key_column: 关键索引列（如商品ID）
    :param compare_columns: 需比对的字段列表，默认为None表示比对所有列
    """
    # 文件加载
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # 如果未指定比对列，则获取除关键列外的所有列
    if compare_columns is None:
        all_columns = set(df1.columns) & set(df2.columns)
        compare_columns = list(all_columns - {key_column})
    # print(compare_columns)  # ['name', 'stock', 'price', 'supplier']

    # 按关键字段智能对齐
    merged = pd.merge(df1, df2, on=key_column, suffixes=('_基准', '_比对'))

    # 差异标记与筛选
    diff_mask = False
    change_columns = []  # 记录创建的变更列名
    for col in compare_columns:
        # 新增 列名_变更列
        change_col_name = f'{col}_变更'
        merged[change_col_name] = merged[f'{col}_基准'] != merged[f'{col}_比对']
        change_columns.append(change_col_name)  # 记录变更列名
        diff_mask = diff_mask | merged[change_col_name]

    diff_df = merged[diff_mask]

    # 优化：只保留有差异的列，删除没有变化的列
    if not diff_df.empty:
        columns_to_keep = [key_column]  # 始终保留关键列
        for col in compare_columns:
            change_col_name = f'{col}_变更'
            # 如果该字段存在差异，则保留基准列、比对列和变更列
            if diff_df[change_col_name].any():
                columns_to_keep.extend([f'{col}_基准', f'{col}_比对', change_col_name])
            # 如果该字段无差异，则不保留任何列

        # 只保留需要的列
        diff_df = diff_df[columns_to_keep]

    # 保存差异报告，为有差异的单元格添加红色背景
    with pd.ExcelWriter('diff_report.xlsx', engine='openpyxl') as f:
        diff_df.to_excel(f, index=False, sheet_name='差异报告')

        # 获取工作表对象
        worksheet = f.sheets['差异报告']

        # 创建红色背景样式
        from openpyxl.styles import PatternFill
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

        # 遍历所有列，为有差异的基准值和比对值添加红色背景
        for col in compare_columns:
            change_col_name = f'{col}_变更'
            if change_col_name in diff_df.columns:
                # 找到基准列和比对列的列索引
                base_col_idx = diff_df.columns.get_loc(f'{col}_基准') + 1  # +1因为Excel列从1开始
                compare_col_idx = diff_df.columns.get_loc(f'{col}_比对') + 1

                # 遍历所有行，为有差异的单元格添加红色背景
                for row_idx, is_diff in enumerate(diff_df[change_col_name], start=2):  # start=2因为数据从第2行开始
                    if is_diff:  # 如果该行该列有差异
                        worksheet.cell(row=row_idx, column=base_col_idx).fill = red_fill
                        worksheet.cell(row=row_idx, column=compare_col_idx).fill = red_fill

    return diff_df

def excel_diff_only(file1, file2, key_column, compare_columns=None):
    """
    Excel文件差异比对核心实现，不显示无差异的列
    :param file1: 基准文件路径
    :param file2: 比对文件路径
    :param key_column: 关键索引列（如商品ID）
    :param compare_columns: 需比对的字段列表，默认为None表示比对所有列
    """
    # 文件加载
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # 如果未指定比对列，则获取除关键列外的所有列
    if compare_columns is None:
        all_columns = set(df1.columns) & set(df2.columns)
        compare_columns = list(all_columns - {key_column})

    # 按关键字段智能对齐
    merged = pd.merge(df1, df2, on=key_column, suffixes=('_基准', '_比对'))

    # 差异标记与筛选
    diff_mask = False
    change_columns = []  # 记录创建的变更列名
    for col in compare_columns:
        # 新增 列名_变更列
        change_col_name = f'{col}_变更'
        merged[change_col_name] = merged[f'{col}_基准'] != merged[f'{col}_比对']
        change_columns.append(change_col_name)  # 记录变更列名
        diff_mask = diff_mask | merged[change_col_name]

    diff_df = merged[diff_mask]

    # 优化：只保留有差异的列，删除没有变化的列
    if not diff_df.empty:
        columns_to_keep = [key_column]  # 始终保留关键列
        for col in compare_columns:
            change_col_name = f'{col}_变更'
            # 如果该字段存在差异，则保留基准列、比对列和变更列
            if diff_df[change_col_name].any():
                columns_to_keep.extend([f'{col}_基准', f'{col}_比对', change_col_name])
            # 如果该字段无差异，则不保留任何列

        # 只保留需要的列
        diff_df = diff_df[columns_to_keep]

    # 保存差异报告
    with pd.ExcelWriter('diff_report.xlsx') as f:
        diff_df.to_excel(f, index=False)

    return diff_df

def excel_diff_showBase_noChange(file1, file2, key_column, compare_columns=None):
    """
    Excel文件差异比对核心实现, 不显示无差异的比较列（显示基准列）
    :param file1: 基准文件路径
    :param file2: 比对文件路径
    :param key_column: 关键索引列（如商品ID）
    :param compare_columns: 需比对的字段列表，默认为None表示比对所有列
    """
    # 文件加载
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # 如果未指定比对列，则获取除关键列外的所有列
    if compare_columns is None:
        all_columns = set(df1.columns) & set(df2.columns)
        compare_columns = list(all_columns - {key_column})

    # 按关键字段智能对齐
    merged = pd.merge(df1, df2, on=key_column, suffixes=('_基准', '_比对'))

    # 差异标记与筛选
    diff_mask = False
    change_columns = []  # 记录创建的变更列名
    for col in compare_columns:
        # 新增 列名_变更列
        change_col_name = f'{col}_变更'
        merged[change_col_name] = merged[f'{col}_基准'] != merged[f'{col}_比对']
        change_columns.append(change_col_name)  # 记录变更列名
        diff_mask = diff_mask | merged[change_col_name]

    diff_df = merged[diff_mask]

    # 优化：删除没有变化的基准列、比对列和值全为False的变更列
    if not diff_df.empty:
        columns_to_keep = [key_column]  # 始终保留关键列
        for col in compare_columns:
            change_col_name = f'{col}_变更'
            # 如果该字段存在差异，则保留基准列、比对列和变更列
            if diff_df[change_col_name].any():
                columns_to_keep.extend([f'{col}_基准', f'{col}_比对', change_col_name])
            # 如果该字段无差异，则只保留其中一个列（避免重复信息）
            else:
                columns_to_keep.append(f'{col}_基准')  # 可以选择保留基准列或比对列

        # 只保留需要的列
        diff_df = diff_df[columns_to_keep]

    # 保存差异报告
    with pd.ExcelWriter('diff_report.xlsx') as f:
        diff_df.to_excel(f, index=False)

    return diff_df

def excel_diff_showAll_noChange(file1, file2, key_column, compare_columns=None):
    """
    Excel文件差异比对核心实现, 显示所有基准列和比较列，不显示无差异变更列
    :param file1: 基准文件路径
    :param file2: 比对文件路径
    :param key_column: 关键索引列（如商品ID）
    :param compare_columns: 需比对的字段列表，默认为None表示比对所有列
    """
    # 文件加载
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # 如果未指定比对列，则获取除关键列外的所有列
    if compare_columns is None:
        all_columns = set(df1.columns) & set(df2.columns)
        compare_columns = list(all_columns - {key_column})

    # 按关键字段智能对齐
    merged = pd.merge(df1, df2, on=key_column, suffixes=('_基准', '_比对'))

    # 差异标记与筛选
    diff_mask = False
    change_columns = []  # 记录创建的变更列名
    for col in compare_columns:
        # 新增 列名_变更列
        change_col_name = f'{col}_变更'
        merged[change_col_name] = merged[f'{col}_基准'] != merged[f'{col}_比对']
        change_columns.append(change_col_name)  # 记录变更列名
        diff_mask = diff_mask | merged[change_col_name]

    diff_df = merged[diff_mask]

    # 优化：删除值全为False的变更列
    if not diff_df.empty:
        for col in change_columns:
            # 如果该变更列的所有值都是False，则删除该列
            if not diff_df[col].any():
                diff_df = diff_df.drop(columns=[col])

    # 保存差异报告
    with pd.ExcelWriter('diff_report.xlsx') as f:
        diff_df.to_excel(f, index=False)

    return diff_df


def excel_diff_showAll(file1, file2, key_column, compare_columns=None):
    """
    Excel文件差异比对核心实现, 显示所有基准列、比较列、变更列
    :param file1: 基准文件路径
    :param file2: 比对文件路径
    :param key_column: 关键索引列（如商品ID）
    :param compare_columns: 需比对的字段列表，默认为None表示比对所有列
    """
    # 文件加载
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # 如果未指定比对列，则获取除关键列外的所有列
    if compare_columns is None:
        all_columns = set(df1.columns) & set(df2.columns)
        compare_columns = list(all_columns - {key_column})
    print(compare_columns)  # ['name', 'stock', 'price', 'supplier']

    # 按关键字段智能对齐
    merged = pd.merge(df1, df2, on=key_column, suffixes=('_基准', '_比对'))

    # 差异标记与筛选
    diff_mask = False
    for col in compare_columns:
        # 新增 列名_变更列

        # 生成差异标记列
        # 为每个比对列创建一个新的布尔列，列名为 {原列名}_变更
        # 通过比较 _基准 和 _比对 两列的值来判断是否存在差异
        # 如果两列值不同，该位置为 True，否则为 False
        merged[f'{col}_变更'] = merged[f'{col}_基准'] != merged[f'{col}_比对']

        # 累积差异标记
        # 使用逻辑或运算符 | 将当前列的差异标记合并到总标记中
        # 只要有任何一列存在差异，对应行的标记就为
        # True
        diff_mask = diff_mask | merged[f'{col}_变更']

    # 筛选差异数据
    # 使用布尔索引筛选出存在差异的行
    # 只保留至少在一列中有差异的记录
    diff_df = merged[diff_mask]

    # 保存差异报告
    with pd.ExcelWriter('diff_report.xlsx') as f:
        diff_df.to_excel(f, index=False)

    return diff_df


if __name__ == '__main__':

    excel_diff_only_red('old.xlsx', 'new.xlsx',key_column='product_id', compare_columns=['price', 'stock'])  # 指定比较2列，只显示差异列，背景色红色。
    # excel_diff_only_red('old.xlsx', 'new.xlsx', key_column='product_id')  # 只显示差异列，背景色红色。

    # excel_diff_only('old.xlsx', 'new.xlsx', key_column='product_id')  # 只显示差异列
    # excel_diff_showBase_noChange('old.xlsx', 'new.xlsx', key_column='product_id')  # 显示差异列和无差异基准列
    # excel_diff_showAll_noChange('old.xlsx', 'new.xlsx', key_column='product_id')  # 显示差异列和无差异列
    # excel_diff_showAll('old.xlsx', 'new.xlsx', key_column='product_id')  # 显示所有列（基准列、比较列、变更列）