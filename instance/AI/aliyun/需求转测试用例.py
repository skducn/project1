import docx
import openpyxl
import requests
import json, os
import re
from datetime import datetime

# ===================== 配置项 =====================
# 替换为你的通义千问 API-KEY（阿里云获取）
DASHSCOPE_API_KEY = "sk-f3e3d8f64cab416fb028d582533c1e01"
# 通义千问 API 地址（固定）
DASHSCOPE_API_URL = "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation"

# 获取当前文件所在目录作为基础目录
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


# # 需求文档路径（Word 文件）
# WORD_FILE_PATH = "【专病库】需求规格说明书.docx"
# # 测试用例 Excel 保存路径
# EXCEL_SAVE_PATH = "测试用例_【专病库】需求规格说明书.xlsx"

# 需求文档路径（Word 文件）
WORD_FILE_PATH = os.path.join(BASE_DIR, "【专病库】需求规格说明书.docx")
# 测试用例 Excel 保存路径
EXCEL_SAVE_PATH = os.path.join(BASE_DIR, "测试用例_【专病库】需求规格说明书.xlsx")



# ===================== 核心函数 =====================
def read_word_document(file_path):
    """读取 Word 文档内容"""
    try:
        doc = docx.Document(file_path)
        full_text = []
        # 读取文档中所有段落
        for para in doc.paragraphs:
            if para.text.strip():  # 跳过空行
                full_text.append(para.text.strip())
        # 拼接所有内容
        requirement_text = "\n".join(full_text)
        if not requirement_text:
            print("Word 文档内容为空")
            return None
        return requirement_text
    except FileNotFoundError:
        print(f"未找到 Word 文件：{file_path}")
        return None
    except Exception as e:
        print(f"读取 Word 文档失败：{str(e)}")
        return None


    # ... existing code ...

def parse_markdown_table_to_json(markdown_text):
    """将 Markdown 表格转换为测试用例 JSON 列表"""
    test_cases = []

    try:
        # 按行分割
        lines = markdown_text.strip().split('\n')
        
        if len(lines) < 3:
            print(f"Markdown 表格行数不足：{len(lines)}")
            return None

        # 提取表头（第一行）
        header_line = lines[0].strip()
        
        # 移除首尾的 |
        if header_line.startswith('|'):
            header_line = header_line[1:]
        if header_line.endswith('|'):
            header_line = header_line[:-1]
        
        headers = [h.strip() for h in header_line.split('|')]
        
        if not headers or len(headers) < 5:
            print(f"表头解析失败，列数：{len(headers)}")
            return None

        # 跳过第二行（分隔线 --- | --- | ---）
        # 从第三行开始提取数据
        success_count = 0
        for line in lines[2:]:
            line = line.strip()
            if not line or line.startswith('---'):
                continue

            # 移除首尾的 |
            if line.startswith('|'):
                line = line[1:]
            if line.endswith('|'):
                line = line[:-1]
            
            # 分割单元格
            parts = line.split('|')
            cells = [part.strip() for part in parts]
            
            # 动态适配列数：如果单元格数少于表头，补充空值
            while len(cells) < len(headers):
                cells.append("")
            
            # 如果单元格数多于表头，合并多余部分到最后一个字段
            if len(cells) > len(headers):
                merged_cells = cells[:len(headers)-1]
                last_cell = '|'.join(cells[len(headers)-1:])
                merged_cells.append(last_cell)
                cells = merged_cells

            # 构建测试用例字典
            case = {}
            for i, header in enumerate(headers):
                case[header] = cells[i] if i < len(cells) else ""

            # 验证必要字段
            if "用例 ID" in case and case["用例 ID"]:
                test_cases.append(case)
                success_count += 1

        print(f"DEBUG: 成功解析 {success_count} 行数据")
        return test_cases if test_cases else None

    except Exception as e:
        print(f"解析 Markdown 表格失败：{str(e)}")
        import traceback
        traceback.print_exc()
        return None

    # ... existing code ...

    
def clean_markdown_blocks(text):
    """清理所有 markdown 代码块标记"""
    # 移除 ```json ... ``` 块
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text)
    
    # 移除 ``` 开头和结尾
    if text.startswith('```'):
        text = text[3:]
    if text.endswith('```'):
        text = text[:-3]
    
    return text.strip()


def call_qianfan_api(requirement_text):
    """调用通义千问 API，将需求文档转换为测试用例"""
    # 构造精准的提示词，确保返回结构化 JSON
    prompt = f"""
    你是资深软件测试工程师，现在请解析这份需求文档自动生成标准测试用例。
请严格按照下面要求执行，不要额外解释：
1. 先从需求文档中提取：
- 功能模块
- 字段/参数名称
- 数据类型
- 约束规则（必填/非必填、长度、范围、枚举、格式、正则等）
2. 为每个字段生成三类测试用例，必须全覆盖：
【1】正向用例：符合所有规则的正常合法数据
【2】反向用例：
- 空值（必填字段为空）
- 非法字符
- 格式错误
- 越界（长度过长/过短）
- 数据类型错误
【3】边界用例：
- 最小长度/最小值
- 最大长度/最大值
- 枚举边界（第一个、最后一个）
- 临界值
3. 输出格式要求：
- 直接输出标准 Markdown 表格，方便我一键复制到 Excel
- 列固定为：["用例 ID", "优先级", "模块", "子模块", "前置条件", "用例类型", "用例检查点", "操作步骤", "预期结果", "备注"]
- 用例ID从 UC001 开始自动递增
- 语言：中文
- 不要多余内容，只输出表格
需求文档内容：
{requirement_text}
"""

    # 构造通义千问 API 请求参数
    headers = {
        "Authorization": f"Bearer {DASHSCOPE_API_KEY}",
        "Content-Type": "application/json"
    }
    data = {
        "model": "qwen-turbo",  # 通义千问轻量版，也可换 qwen-plus/qwen-max
        "input": {
            "messages": [
                {"role": "user", "content": prompt}
            ]
        },
        "parameters": {
            "result_format": "text",  # 返回文本格式
            "temperature": 0.1,  # 低随机性，保证结果稳定
            "max_tokens": 5000  # 最大返回字符数
        }
    }

    try:
        # 发送请求
        response = requests.post(DASHSCOPE_API_URL, headers=headers, json=data, timeout=120)
        response.raise_for_status()  # 抛出 HTTP 错误
        result = response.json()

        # 提取 AI 返回的测试用例内容
        ai_content = result["output"]["text"].strip()

        # 清理所有可能的 markdown 代码块标记
        ai_content = clean_markdown_blocks(ai_content)

        # 优先尝试解析为 JSON
        try:
            test_cases = json.loads(ai_content)
            print(f"✅ 成功解析 JSON 格式的测试用例，共 {len(test_cases)} 条")
            return test_cases
        except json.JSONDecodeError as e:
            print(f"JSON 解析失败：{e}")
            print(f"AI 返回内容前 200 字符：{ai_content[:200]}")
            
            # 尝试提取 JSON 部分（如果 AI 添加了额外说明文字）
            json_match = re.search(r'\[\s*\{.*\}\s*\]', ai_content, re.DOTALL)
            if json_match:
                try:
                    extracted_json = json_match.group()
                    test_cases = json.loads(extracted_json)
                    print(f"✅ 从混合内容中提取 JSON 成功，共 {len(test_cases)} 条")
                    return test_cases
                except:
                    pass
            
            pass

        # 如果不是 JSON，尝试解析 Markdown 表格
        test_cases = parse_markdown_table_to_json(ai_content)
        if test_cases:
            print(f"✅ 成功从 Markdown 表格解析出 {len(test_cases)} 个测试用例")
            return test_cases

        print(f"AI 返回的内容无法解析为 JSON 或 Markdown 表格：{ai_content[:200]}")
        return None

    except requests.exceptions.HTTPError as e:
        print(f"❌ HTTP 错误：{e}")
        print(f"状态码：{response.status_code}")
        if response.status_code == 401:
            print("💡 提示：401 未授权，可能原因：")
            print("   1. API-KEY 无效或缺少 sk-前缀")
            print("   2. API-KEY 已过期")
            print("   3. 账号欠费或未开通服务")
            print("   ✅ 解决：访问 https://bailian.console.aliyun.com/api-key 重新获取")
        return None
    except Exception as e:
        print(f"调用通义千问 API 失败：{str(e)}")
        return None


def save_test_cases_to_excel(test_cases, save_path):
    """将测试用例写入 Excel，带格式优化"""
    if not test_cases:
        print("无测试用例数据可保存")
        return

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "功能测试用例"

    # 定义表头
    # headers = ["用例 ID", "模块", "用例名称", "前置条件", "操作步骤", "预期结果", "优先级", "创建时间"]
    headers = ["用例 ID", "优先级", "模块", "子模块", "前置条件", "用例类型", "用例检查点", "操作步骤", "预期结果", "备注"]
    ws.append(headers)

    # 设置表头样式（加粗、居中、浅灰色背景）
    header_style = openpyxl.styles.Font(bold=True)
    header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    header_fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_style
        cell.alignment = header_alignment
        cell.fill = header_fill

    # 写入测试用例数据
    create_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for case in test_cases:
        # 处理可能为列表的字段，转换为字符串
        steps = case.get("操作步骤", "")
        if isinstance(steps, list):
            steps = "\n".join(steps)
        
        result = case.get("预期结果", "")
        if isinstance(result, list):
            result = "\n".join(result)
        
        row_data = [
            case.get("用例 ID", ""),
            case.get("优先级", ""),
            case.get("模块", ""),
            case.get("子模块", ""),
            case.get("前置条件", ""),
            case.get("用例类型", ""),
            case.get("用例检查点", ""),
            steps,
            result,
            case.get("备注", ""),
            create_time
        ]
        ws.append(row_data)

    # 调整列宽（适配内容长度）
    column_widths = [18, 15, 15, 15, 35, 25, 45, 35, 8, 20]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # 自动换行（操作步骤/预期结果字段）
    wrap_alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # 操作步骤（第 5 列）、预期结果（第 6 列）开启自动换行
        row[4].alignment = wrap_alignment
        row[5].alignment = wrap_alignment

    # 保存 Excel
    try:
        wb.save(save_path)
        print(f"✅ 测试用例已成功保存至：{save_path}")
    except PermissionError:
        print(f"❌ 保存失败：{save_path} 文件已被打开，请关闭后重试")
    except Exception as e:
        print(f"❌ 保存 Excel 失败：{str(e)}")


# ===================== 主程序 =====================
if __name__ == "__main__":
    # 1. 读取 Word 需求文档
    print("📄 正在读取 Word 需求文档...")
    requirement_text = read_word_document(WORD_FILE_PATH)
    if not requirement_text:
        exit(1)

    # 2. 调用通义千问 API 生成测试用例
    print("🤖 正在调用通义千问 API 生成测试用例...")
    test_cases = call_qianfan_api(requirement_text)
    if not test_cases:
        exit(1)

    # 3. 保存到 Excel
    print("💾 正在将测试用例写入 Excel...")
    save_test_cases_to_excel(test_cases, EXCEL_SAVE_PATH)