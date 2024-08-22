# -*- coding: utf-8 -*-
# *********************************************************************
# Author        : John
# Date          : 2020-12-8
# Description   : openpyxl 对象层
# openpyxl 官网 （http://openpyxl.readthedocs.org/en/latest/），支持Excel 2010 xlsx/xlsm/xltx/xltm 的读写，最新版本3.0.6（Oct 23, 2020）
# 1，请安装 openpyxl 3.0.0，其他版本如3.0.2使用中会报错。 pip3 install openpyxl == 3.0.0
# 报错：File "src\lxml\serializer.pxi", line 1652, in lxml.etree._IncrementalFileWriter.write TypeError: got invalid input value of type <类与实例 'xml.etree.ElementTree.Element'>, expected string or Element
# 解决方法: pip uninstall lxml   或更新最新openpyxl 3.0.7以上版本
# 2，如果文字编码是“gb2312” 读取后就会显示乱码，请先转成Unicode
# 3，openpyxl 的首行、首列 是 （1,1）而不是（0,0）
# 4，openpyxl 的NULL空值对应于python中的None，表示这个cell里面没有数据。
# 5，openpyxl 的numberic数字型，统一按照浮点数来进行处理，对应于python中的float
# 6，openpyxl 的string字符串型，对应于python中的unicode
# 7，默认情况下，openpyxl会将整个xlsx读入到内存中，方便处理。
# 8，openpyxl 操作大文件时可使用 Optimized reader 和 Optimized writer 两种模式，它们提供了流式的接口，速度更快，使我们可以用常量级的内存消耗来读取和写入无限量的数据。
# Optimized reader，打开文件使用use_iterators=True参数，如：wb = load_workbook(filename = 'haggle.xlsx',use_iterators=True)
# 9，openpyxl 读取大数据的效率没有 xlrd 高
# 10，openpyxl 与 xlsxwriter xlrd xlwt xlutils 的比较，这些库都不支持 excel 写操作，一般只能将原excel中的内容读出、做完处理后，再写入一个新的excel文件。
# 11，openpyxl常用模块用法：https://www.debug8.com/python/t_41519.html

# 参考：https://blog.csdn.net/m0_47590417/article/details/119082064
# https://blog.csdn.net/four91/article/details/106141274
# *********************************************************************

from openpyxl import load_workbook
import openpyxl, sys, platform, os
import openpyxl.styles
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块
from datetime import date
from time import sleep
import psutil

# from PO.ColorPO import *
# Color_PO = ColorPO()
#
# from PO.CharPO import *
# Char_PO = CharPO()

'''
1 新建excel newExcel()
2 获取总行数和总列数 l_getTotalRowCol()
3 设置单元格的值 setCellValue()
4 批量设置单元格的值  setMoreCellValue()
5 获取每行数据 l_getRowData()
6 获取每列数据 l_getColData()
7 获取N列的行数据 l_getRowDataByPartCol()
8 获取N列的列数据，可忽略多行 l_getColDataByPartCol()
9 设置单元格背景色 setCellColor()
10 删除行 delRowData()
11 删除列 delColData()
12 清空行 clsRowData()
13 清空列 clsColData()
14 两表比较，输出差异 cmpExcel()
15 获取单元格的值 getCellValue()
16 设置工作表标题选项卡背景颜色 setSheetColor()
17 关闭excel进程
18 添加工作表（工作表名重复时，保留原工作表）
19 添加工作表（工作表名重复时，覆盖原工作表）
20 删除工作表
21 初始化数据
22 get_column_letter 得到表格列的字母编号
'''

class OpenpyxlPO():

    def __init__(self, file):
        self.file = file
        self.wb = openpyxl.load_workbook(self.file)
        self.wb.sheetnames  # 获取所有的工作表名称
        # # self.nameSizeColor = openpyxl.styles.Font(name="宋体", size=33, color="00CCFF")

    def save(self):
        self.wb.save(self.file)

    def open(self):
        if platform.system() == 'Darwin':
            os.system("open " + self.file)
        if platform.system() == 'Windows':
            os.system("start " + self.file)

    def sh(self, varSheet):
        if isinstance(varSheet, int):
            sh = self.wb[self.wb.sheetnames[varSheet]]
            return sh
        elif isinstance(varSheet, str):
            sh = self.wb[varSheet]
            return sh
        else:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")
            exit(0)

    def ws(self, active_row):
        worksheet = self.wb.get_sheet_by_name(self.sh)
        coords = "A" + str(active_row)
        print(coords)
        # worksheet.cell(row=active_row, column=1)
        worksheet.sheet_view.selection[0].activeCell = coords
        worksheet.sheet_view.selection[0].sqref = coords

    # 1 新建excel（ok）
    def newExcel(self, varFileName, *varSheetName):
        # 新建excel，生成N个工作表（默认一个Sheet1）
        # 注意：如果文件已存在则会先删除后再新建！
        # Openpyxl_PO.newExcel("d:\\444.xlsx")  # 新建excel，默认生成一个工作表Sheet1
        # Openpyxl_PO.newExcel("d:\\444.xlsx", "mySheet1", "mySheet2","mySheet3")  # 新建excel，生成三个工作表（mySheet1,mySheet2,mySheet3），默认定位在第一个mySheet1表。

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            if len(varSheetName) == 0:
                ws.title = "Sheet1"
            else:
                ws.title = varSheetName[0]
            for i in range(1, len(varSheetName)):
                wb.create_sheet(varSheetName[i])
            wb.save(varFileName)
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 2 获取总行数和总列数(ok)
    def l_getTotalRowCol(self, varSheet=0):
        # print(Openpyxl_PO.l_getTotalRowCol())  # [4,3] //返回第1个工作表的总行数和总列数
        # print(Openpyxl_PO.l_getTotalRowCol(1))  # [4,3] //返回第2个工作表的总行数和总列数
        # print(Openpyxl_PO.l_getTotalRowCol("python"))  # [4,3] //返回python工作表的总行数和总列数
        sh = self.sh(varSheet)
        rows = sh.max_row
        columns = sh.max_column
        return [rows, columns]


    # 3 设置单元格的值(ok)
    def setCellValue(self, varRow, varCol, varContent, l_color, varSheet=0):
        # ['c6efce', '006100']  背景色淡绿底，字体深绿色
        # ['ffffff', '000000'] 白色，黑色
        # ['ffeb9c', '000000'] 橙色，黑色
        # Openpyxl_PO.setCellValue(5, 3, "777777",['ffffff', '000000'])  # 对第一个sheet表的第5行第3列写入数据
        # Openpyxl_PO.setCellValue(5, 3, "12345678", ['ffeb9c', '000000'],"python")  # 对python工作表的的第5行第3列写入数据，
        # Openpyxl_PO.setCellValue(5, 3, "12345678", "","python")  # 对python工作表的的第5行第3列写入数据
        try:
            sh = self.sh(varSheet)
            if l_color == "":
                sh.cell(row=varRow, column=varCol, value=varContent)
                sh.cell(row=varRow, column=varCol, value=varContent)
            else:
                # fill(填充类，可设置单元格填充颜色等)
                fille = PatternFill('solid', fgColor=l_color[0])  # 背景色
                # font(字体类，可设置字号、字体颜色、下划线等)
                font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=l_color[1])  # 字体色
                sh.cell(row=varRow, column=varCol, value=varContent).fill = fille
                sh.cell(row=varRow, column=varCol, value=varContent).font = font
                # alignment(位置类、可以设置单元格内数据各种对齐方式)
                # number_format(格式类，可以设置单元格内各种类型的数据格式)
                # protection(保护类，可以设置单元格写保护等)。

        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

            # Color_PO.consoleColor("31", "31", "[ERROR] ", "call " + sys._getframe(1).f_code.co_name + " (line " + str(
            #     sys._getframe(1).f_lineno) + ", call " + sys._getframe(
            #     0).f_code.co_name + " from '" + sys._getframe().f_code.co_filename + "')")

    # 4 批量设置单元格的值(ok)
    def setMoreCellValue(self, varList_Row_Col_Content, varSheet=0):
        # Openpyxl_PO.setMoreCellValue([[7, "你好", "测试", "报告"], [9, "再见", "", "好了"]])  # 对第7行第9行分别写入内容
        # Openpyxl_PO.setMoreCellValue([[2, "a", "b", "c"], [3, "d", "", "f"]], -1)  # 对最后一个sheet表第2行第3行分别写入内容

        try:
            sh = self.sh(varSheet)
            for i in range(len(varList_Row_Col_Content)):
                for j in range(1, len(varList_Row_Col_Content[i])):
                    sh.cell(row=varList_Row_Col_Content[i][0], column=j, value=varList_Row_Col_Content[i][j])
            # self.wb.save(self.file)
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 5 获取每行数据(ok)
    def l_getRowData(self, varSheet=0):
        # print(Openpyxl_PO.l_getRowData())
        # print(Openpyxl_PO.l_getRowData("python"))

        l_rowData = []  # 每行数据
        l_allData = []  # 所有行数据
        sh = self.sh(varSheet)
        for cases in list(sh.rows):
            for i in range(len(cases)):
                l_rowData.append(cases[i].value)
            l_allData.append(l_rowData)
            l_rowData = []
        return (l_allData)

    # 6 获取每列数据(ok)
    def l_getColData(self, varSheet=0):
        # print(Openpyxl_PO.l_getColData())
        # print(Openpyxl_PO.l_getColData("python"))

        l_colData = []  # 每列数据
        l_allData = []  # 所有行数据
        sh = self.sh(varSheet)
        for cases in list(sh.columns):
            for i in range(len(cases)):
                l_colData.append(cases[i].value)
            l_allData.append(l_colData)
            l_colData = []
        return (l_allData)

    # 7 获取N列的行数据（ok）
    def l_getRowDataByPartCol(self, l_varCol, varSheet=0):
        # print(Openpyxl_PO.l_getRowDataByPartCol([1, 2, 4]))  # 获取第1，2，4列的行数据
        # print(Openpyxl_PO.l_getRowDataByPartCol([1, 2, 4], -1))  # 获取最后一个工作表的第1，2，4列的行数据

        l_rowData = []  # 每行的数据
        l_allData = []  # 所有的数据
        sh = self.sh(varSheet)
        for row in range(1, sh.max_row + 1):
            try:
                for column in l_varCol:
                    l_rowData.append(sh.cell(row, column).value)
                l_allData.append(l_rowData)
                l_rowData = []
            except:
                print("errorrrrrrrrrr, line " + str(
                    sys._getframe(1).f_lineno) + ", in " + sys._getframe().f_code.co_name + "() ")
                print("建议：参数列表元素不能是0或负数")
                exit(0)

        return l_allData

    # 8 获取N列的列数据，可忽略多行(ok)
    def l_getColDataByPartCol(self, l_varCol, l_varIgnoreRowNum, varSheet=0):
        # print(Openpyxl_PO.l_getColDataByPartCol([1, 3], [1, 2]))  # 获取第二列和第四列的列值，并忽略第1，2行的行值。
        # print(Openpyxl_PO.l_getColDataByPartCol([2], [], "python"))  # 获取第2列所有值。

        l_colData = []  # 每列的数据
        l_allData = []  # 所有的数据
        sh = self.sh(varSheet)
        for col in l_varCol:
            try:
                for row in range(1, sh.max_row+1):
                    if row not in l_varIgnoreRowNum:
                        l_colData.append(sh.cell(row, col).value)
                l_allData.append(l_colData)
                l_colData = []
            except:
                print("errorrrrrrrrrr, line " + str(
                    sys._getframe(1).f_lineno) + ", in " + sys._getframe().f_code.co_name + "() ")
                print("建议：参数列表元素不能是0或负数")
                exit(0)

        return l_allData

    # 9 设置单元格背景色(ok)
    def setCellColor(self, row, col, varColor, varSheet=0):
        # Openpyxl_PO.setCellColor(6, 7, "FF0000")   将单元格第6行第7列的背景色设置为红色（FF0000）
        # 绿色 = 00E400，黄色 = FFFF00，橙色 = FF7E00，红色 = FF0000，粉色 = 99004C，褐色 =7E0023

        sh = self.sh(varSheet)
        style = PatternFill("solid", fgColor=varColor)
        sh.cell(row, col).fill = style

    # 10 删除行(ok)
    def delRowData(self, varFrom, varSeries=1, varSheet=0):
        # Openpyxl_PO.delRowData(2, 3)  # 删除第2行之连续3行（删除2，3，4行）

        try:
            sh = self.sh(varSheet)
            sh.delete_rows(varFrom, varSeries)  # 删除从某行开始连续varSeries行
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")
            exit(0)

    # 11 删除列(ok)
    def delColData(self, varFrom, varSeries=1, varSheet=0):
        # Openpyxl_PO.delColData(1, 2)  # 删除第1列之连续2列（删除1，2列）
        # Openpyxl_PO.delColData(2, 1, "python")  # 删除第2列之连续1列（删除2列）

        try:
            sh = self.sh(varSheet)
            sh.delete_cols(varFrom, varSeries)  # 删除从某列开始连续varSeries行
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 12 清空行(ok)
    def clsRowData(self, varNums, varSheet=0):
        # Openpyxl_PO.clsRowData(2)  # 清空第2行

        try:
            sh = self.sh(varSheet)
            for i in range(sh.max_row):
                sh.cell(row=varNums, column=i + 1, value="")
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(
                sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 13 清空列(ok)
    def clsColData(self, varCol, varSheet=0):
        # 注意：保留标题，从第二行开始清空
        # Openpyxl_PO.clsColData(2)  # 清空第2列
        # Openpyxl_PO.clsColData(1, "python")  # 清空第1列
        try:
            sh = self.sh(varSheet)
            for i in range(sh.max_row-1):
                sh.cell(row=i + 2, column=varCol).value = None
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(
                sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")




    # 14 两表比较，输出差异
    def cmpExcel(self, file1, file1Sheet, file2, file2Sheet):
        # 输出的列表中，第一个元素是序号。
        try:
            list1 = self.l_getRowColNums(file1, file1Sheet)
            list2 = self.l_getRowColNums(file2, file2Sheet)
            tmpList1 = []
            tmpList2 = []
            mainList1 = []
            mainList2 = []
            wb = load_workbook(file1)
            wk_sheet = wb[file1Sheet]
            for i in range(1, list1[0] + 1):
                tmpList1.append(i)
                for j in range(1, list1[1] + 1):
                    tmpList1.append(wk_sheet.cell(row=i, column=j).value)
                mainList1.append(tmpList1)
                tmpList1 = []
            # print(mainList1)

            wb = load_workbook(file2)
            wk_sheet = wb[file2Sheet]
            for i in range(1, list2[0] + 1):
                tmpList2.append(i)
                for j in range(1, list1[1] + 1):
                    tmpList2.append(wk_sheet.cell(row=i, column=j).value)
                mainList2.append(tmpList2)
                tmpList2 = []
            # print(mainList2)

            a = [x for x in mainList1 if x in mainList2]  # 两个列表中都存在
            b = [y for y in (mainList1) if y not in a]  # 两个列表中的不同元素，输出 mainList1 中差异部分
            c = [y for y in (mainList2) if y not in a]  # 两个列表中的不同元素，输出 mainList2 中差异部分

            if b == []:
                print("ok，两表一致")
            else:
                return file1, b, file2, c
                # print(file1 + " => " + str(b))
                # print(file2 + " => " + str(c))

        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(
                sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 15 获取单元格值
    def getCellValue(self, varRow, varCol, varSheet=0):
        sh = self.sh(varSheet)
        cell_value = sh.cell(row=varRow, column=varCol).value
        return cell_value

    # 16 设置工作表标题选项卡背景颜色
    def setSheetColor(self, varColor, varSheet=0):
        # Openpyxl_PO.setSheetColor("FF0000")
        # 绿色 = 00E400，黄色 = FFFF00，橙色 = FF7E00，红色 = FF0000，粉色 = 99004C，褐色 =7E0023
        sh = self.sh(varSheet)
        sh.sheet_properties.tabColor = varColor

    # 17 关闭excel进程
    def closeExcelPid(self, varApplication):

        '''关闭进程
        os.system 输出如果出现乱码，需将 File->Settings->Editor->File Encodings 中 Global Encoding 设置成 GBK'''
        pids = psutil.pids()
        for pid in pids:
            try:
                p = psutil.Process(pid)
                # print('pid=%s,pname=%s' % (pid, p.name()))
                # 关闭excel进程
                if p.name() == varApplication:
                    cmd = 'taskkill /F /IM ' + varApplication + ">> null"
                    os.system(cmd)
                    sleep(2)
            except Exception as e:
                pass

    # 18 添加工作表（工作表名重复时，保留原工作表）
    def addSheet(self, varSheetName, varIndex=0):
        # 默认在左侧第一个位置上添加工作表，如果工作表名已存在，第一次重名时自动更名为原有名字后加1，第二次重名时自动更名为第一次更名后数字运算加1。如 test1工作表添加了2次后，生成test11,test12两个工作表。
        # Openpyxl_PO.addSheet("test1")  # 默认在左侧第一个位置上添加工作表 test1
        # Openpyxl_PO.addSheet("haha12", 99)   # 当index足够大时，则在最后一个位置添加工作表 haha12
        # Openpyxl_PO.addSheet("7894", -1)   # 则倒数第二个位置添加工作表 7894

        try:
            self.wb.create_sheet(title=varSheetName, index=varIndex)
            self.save()
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")


    # 19 添加工作表（工作表名重复时，删除旧的工作表，谨慎！）
    def addSheetByCover(self, varSheetName, varIndex=0):
        # 已作了处理，如果添加的工作表已存在，则先删除旧的
        # Openpyxl_PO.addSheet("test1")  # 默认在左侧第一个位置上添加工作表，如果工作表名已存在，第一次重名时自动更名为原有名字后加1，第二次重名时自动更名为第一次更名后数字运算加1。如 test1工作表添加了2次后，生成test11,test12两个工作表。
        # Openpyxl_PO.addSheet("haha12", 99)   # 当index足够大时，则在最后一个位置添加工作表
        # Openpyxl_PO.addSheet("7894", -1)   # 则倒数第二个位置添加工作表。

        for i in self.wb.sheetnames:
            if i == varSheetName:
                del self.wb[i]
                break
        self.wb.create_sheet(title=varSheetName, index=varIndex)
        self.save()


    # 20 删除工作表
    def delSheet(self, varSheetName):
        # Openpyxl_PO.delSheet("test1")  # 删除test1工作表
        # 注意确保工作表存在
        try:
            del self.wb[varSheetName]
            self.save()
        except:
            print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 21 初始化数据
    def initData(self, data, varSheet=0):
        sh = self.sh(varSheet)
        for r in range(len(data)):
            sh.append(data[r])
        self.save()


    # 22 get_column_letter得到表格列的字母编号
    def get_column_letter(self, varRowFrom, varRowTo, varColFrom, varColTo, varSheet = 0):
        # 用get_column_letter得到表格列的字母编号 https://www.pynote.net/archives/2269
        sh = self.sh(varSheet)
        for row in range(varRowFrom, varRowTo):
            for col in range(varColFrom, varColTo):
                _ = sh.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
        self.save()

    def freeze(self, varCell, varSheet=0):
        # 冻结单元格 sheet.freeze_panes
        # 如:Openpyxl_PO.freeze("h2") 表示将a1-h2区间固定（不移动）
        sh = self.sh(varSheet)
        sh.freeze_panes = varCell

        #
        # coords = "W48"
        # # sh.sheet_view.selection[0].activeCell = coords
        # # sh.sheet_view.selection[0].sqref = coords
        # sh.sheet_view.topLeftCell = coords
        self.save()

    def filter(self, varCell="all", varSheet=0):
        # .auto_filter.ref = sheet.dimensions 给所有字段添加筛选器；
        #  .auto_filter.ref = "A1" 给 A1 这个格子添加“筛选器”，就是给第一列添加“筛选器
        sh = self.sh(varSheet)
        if varCell == "all":
            sh.auto_filter.ref = sh.dimensions
        else:
            sh.auto_filter.ref = varCell
        self.save()


if __name__ == "__main__":

    Openpyxl_PO = OpenpyxlPO("d:\\1.xlsx")
    Openpyxl_PO.save()

    Openpyxl_PO.closeExcelPid('EXCEL.EXE')
    # print(Openpyxl_PO.wb.sheetnames)
    #
    # print("1 新建excel".center(100, "-"))
    # # Openpyxl_PO.newExcel("d:\\444.xlsx")  # 新建excel，默认生成一个工作表Sheet1
    # # Openpyxl_PO.newExcel("d:\\444.xlsx", "mySheet1", "mySheet2", "mySheet3")  # 新建excel，生成三个工作表（mySheet1,mySheet2,mySheet3），默认定位在第一个mySheet1表。
    #
    # print("2 获取总行数和总列数".center(100, "-"))
    # # print(Openpyxl_PO.l_getTotalRowCol())  # [4,3] //返回第1个工作表的总行数和总列数
    # # print(Openpyxl_PO.l_getTotalRowCol(1))  # [4,3] //返回第2个工作表的总行数和总列数
    # # print(Openpyxl_PO.l_getTotalRowCol("python"))  # [4,3] //返回python工作表的总行数和总列数
    #
    # print("3 设置单元格的值".center(100, "-"))
    # Openpyxl_PO.setCellValue(5, 6, "jinhao",['ffffff', '000000'])  # 对第一个sheet表的第5行第3列写入数据
    # Openpyxl_PO.setCellValue(5, 7, "12345678", ['ffffff', '000000'], "case")  # 对python工作表的的第5行第3列写入数据
    # Openpyxl_PO.save()
    #
    # print("4 批量设置单元格的值".center(100, "-"))
    # # Openpyxl_PO.setMoreCellValue([[7, "你好", "测试", "报告"], [9, "再见", "", "好了"]])  # 对第7行第9行分别写入内容
    # # Openpyxl_PO.setMoreCellValue([[2, "a", "b", "c"], [3, "d", "", "f"]], -1)  # 对最后一个sheet表第2行第3行分别写入内容
    # # Openpyxl_PO.save()
    #
    # print("5 获取每行数据".center(100, "-"))
    # # print(Openpyxl_PO.l_getRowData())
    # # print(Openpyxl_PO.l_getRowData("python"))
    #
    # print("6 获取每列数据".center(100, "-"))
    # # print(Openpyxl_PO.l_getColData())
    # # print(Openpyxl_PO.l_getColData("python"))
    #
    # print("7 获取指定列的行数据".center(100, "-"))
    # # print(Openpyxl_PO.l_getRowDataByPartCol([1, 2, 4]))   # 获取第1，2，4列的行数据
    # # print(Openpyxl_PO.l_getRowDataByPartCol([1, 2, 4], -1))   # 获取最后一个工作表的第1，2，4列的行数据
    #
    # print("8 获取某些列的列数据，可忽略多行".center(100, "-"))
    # # print(Openpyxl_PO.l_getColDataByPartCol([1, 3], [1, 2]))   # 获取第二列和第四列的列值，并忽略第1，2行的行值。
    # # print(Openpyxl_PO.l_getColDataByPartCol([2], [], "python"))  # 获取第2列所有值。
    #
    # print("9 设置单元格背景色".center(100, "-"))
    # Openpyxl_PO.setCellColor(11, 1, "00E400")
    # Openpyxl_PO.save()
    #
    # print("10 删除行".center(100, "-"))
    # # Openpyxl_PO.delRowData(2, 3)  # 删除第2行之连续3行（删除2，3，4行）
    # # Openpyxl_PO.save()
    #
    # print("11 删除列".center(100, "-"))
    # # Openpyxl_PO.delColData(1, 2)  # 删除第1列之连续2列（删除1，2列）
    # # Openpyxl_PO.delColData(2, 1, "python")  # 删除第2列之连续1列（删除2列）
    # # Openpyxl_PO.save()
    #
    # print("12 清空行".center(100, "-"))
    # # Openpyxl_PO.clsRowData(2)  # 清空第2行
    # # Openpyxl_PO.save()
    #
    # print("13 清空列".center(100, "-"))
    # # Openpyxl_PO.clsColData(2)  # 清空第2列
    # # Openpyxl_PO.clsColData(1, "python")  # 清空第2列
    # # Openpyxl_PO.save()
    #
    # print("14 两表比较，输出差异".center(100, "-"))
    # # file1,list1,file2,list2 = Openpyxl_PO.cmpExcel("d:\\test1.xlsx", "mySheet1", "d:\\test2.xlsx", "mySheet1")
    # # print(file1 + ">"*50)
    # # for l in list1:
    # #     print(l)
    # # print("\n" + file2 + ">"*50)
    # # for l in list2:
    # #     print(l)
    #
    # print("15 获取单元格值".center(100, "-"))
    # # print(Openpyxl_PO.getCellValue(3, 2))  # 获取第3行第2列的值
    #
    # print("16 设置工作表标题选项卡背景颜色".center(100, "-"))
    # # Openpyxl_PO.setSheetColor("FF0000")
    # # Openpyxl_PO.save()
    #
    # print("17 关闭excel进程".center(100, "-"))
    # # Openpyxl_PO.closeExcelPid('EXCEL.EXE')
    #
    # print("18 添加工作表（工作表名重复时，保留原工作表）".center(100, "-"))
    # # Openpyxl_PO.addSheet("test1")  # 默认在左侧第一个位置上添加工作表，如果工作表名已存在，第一次重名时自动更名为原有名字后加1，第二次重名时自动更名为第一次更名后数字运算加1。如 test1工作表添加了2次后，生成test11,test12两个工作表。
    # # Openpyxl_PO.addSheet("haha12", 99)   # 当index足够大时，则在最后一个位置添加工作表
    # # Openpyxl_PO.addSheet("7894", -1)   # 则倒数第二个位置添加工作表。
    #
    # print("19 添加工作表（工作表名重复时，删除旧的工作表，谨慎！）".center(100, "-"))
    # # Openpyxl_PO.addSheetByCover("haha12", 99)   # 当index足够大时，则在最后一个位置添加工作表
    # # Openpyxl_PO.addSheetByCover("7894", -1)   # 则倒数第二个位置添加工作表。
    #
    # print("20 删除工作表".center(100, "-"))
    # # Openpyxl_PO.delSheet("test1")  # 删除test1工作表
    #
    #
    # print("21 初始化数据".center(100, "-"))
    # # Openpyxl_PO.initData([['姓名', '电话', '成绩', '学科'], ['毛泽东', 15266606298, 14, '化学'], ['周恩来', 15201077791, 78, '美术']])   # 初始化数据，一般用于空工作表，如果工作表里有内容则在原有内容下面生成。
    # # Openpyxl_PO.initData([['姓名', '电话', '成绩', '学科'], ['金浩', 13816109050, 119, '语文'], ['Marry', 15201062191, 28, '数学']], "haha")  # 在haha工作表中初始化数据
    #
    #
    # print("22 get_column_letter得到表格列的字母编号".center(100, "-"))
    # # Openpyxl_PO.get_column_letter(10, 20, 3, 10)
    #
    # Openpyxl_PO.open()