# coding: utf-8
# ********************************************************************************************************************
# Author     : John
# Date       : 2020-11-13
# Description: openpyxl 学习
# openpyxl模块是一个读写Excel 2010文档的Python库，如果要处理更早格式的Excel文档，需要用到额外的库，openpyxl是一个比较综合的工具，能够同时读取和修改Excel文档。
# https://www.debug8.com/python/t_40455.html  openpyxl3.0.3中文手册--教程
# ********************************************************************************************************************

import openpyxl
excel = openpyxl.Workbook()  # 创建本地工作簿
excel = openpyxl.load_workbook("abc.xlsx")  # 加载本地已存在的工作簿
# 操作工作簿完毕后需要保存工作簿
excel.save("hello.xlsx")
# PS：Workbook和load_workbook相同，返回的都是一个Workbook对象

'''Workbook对象'''
# 属性 attribute
excel.active # 获取当前活跃的Worksheet对象
excel.worksheets # 以列表的形式返回所有的Worksheet对象
excel.sheetnames # 获取工作簿中的表名[name1, name2, name3]
excel.encoding # 获取文档的字符集编码
excel.properties # 获取文档的元数据，如标题，创建者，创建日期等
# 通过索引值设置当前活跃的worksheet
excel.active = 0

ws = excel.active
# 注：此方法默认获取工作簿中索引为 0 的工作表（工作簿中的第一个工作表），除非手动改为其他的值。

# 可以通过工作表名获取到worksheet对象
excel_sheet = excel["Sheet1"]
# 工作表的创建和删除
excel.remove(excel_sheet) # 删除一个表格，参数为worksheet对象
# excel.remove_sheet()已被废弃sh
excel.create_sheet() # 创建一个空的表格，参数为表名和index

# 使用 excel.create_sheet（）方法创建新的工作表：
# 在末尾创建，工作表名自动按（Sheet, Sheet1, Sheet2, ...）的顺序命名
ws0 = excel.create_sheet()
# 在末尾插入（默认）名字为 Mysheet 的工作表
ws1 = excel.create_sheet("Mysheet")
# 在前端插入（即插入的表索引号为 0）名字为 Mysheet 的工作表
ws2 = excel.create_sheet("Mysheet", 0)
# 在倒数第二个位置插入名字为 Mysheet 的工作表
ws3 = excel.create_sheet("Mysheet", -1)

创建工作表示时，若不指定创建的工作表名，则自动安照（Sheet, Sheet1, Sheet2, ...）的顺序为工作表命名。之后可以使用 Worksheet.title 属性为工作表重命名：

ws.title = "New Title"
工作表标题选项卡的背景颜色默认为白色，可以通过Worksheet.sheet_properties.tabColor 属性更改其背景颜色，赋值为 RRGGBB 颜色代码：

ws.sheet_properties.tabColor = "1072BA"
给工作表命名后，可以通过表名获取工作表：

ws3 = wb["New Title"]
使用 Workbook.sheetname 属性查看工作簿所有工作表的名称：

print(wb.sheetnames)

输出结果：
['Sheet2', 'New Title', 'Sheet1']
可以遍历工作表：

for sheet in wb:
    print(sheet.title)
使用 Workbook.copy_worksheet() 方法在单个工作簿中创建工作表的副本:

source = wb.active
target = wb.copy_worksheet(source)

注：此方法仅复制单元格（包括值、样式、超链接和注释）和工作表一些特性（包括维度、格式和属性），工作簿或工作表其他属性不会复制，例如图像、图表。
此方法只能在工作簿内复制工作表，不能在工作簿之间复制工作表，如果工作薄是以以只读或仅写模式打开，则无法复制工作表。



'''Worksheet对象'''
# 属性 attribute
excel_sheet.title # 表格的标题(可读可写)
excel_sheet.dimensions # 表格的大小，这里的大小是指含有数据的表格的大小，即：左上角的坐标:右下角的坐标

excel_sheet.max_row # 表格的最大行数
excel_sheet.min_row # 表格的最小行数
excel_sheet.max_column # 表格的最大列数
excel_sheet.min_column # 表格的最小列数
excel_sheet.rows # 按行获取单元格(Cell对象) - 生成器
excel_sheet.columns # 按列获取单元格(Cell对象) - 生成器
excel_sheet.values # 按行获取表格的内容(数据)  - 生成器
# 方法
excel_sheet.iter_rows() # 按行获取所有单元格，内置属性有(min_row,max_row,min_col,max_col)
excel_sheet.iter_cols() # 按列获取所有的单元格

# 通过key值获取
excel_sheet["A"]  # 返回A列中所有的单元格cell对象
excel_sheet["1"]  # 返回第一行中所有的单元格cell对象
excel_sheet["A1"] # 返回该单元格对象
for row in excel2[‘abc‘].iter_rows(min_row=2,max_row=4,min_col=2,max_col=4):
    print(row)

# (<Cell ‘abc‘.B2>, <Cell ‘abc‘.C2>, <Cell "abc".D2>)
# (<Cell ‘abc‘.B3>, <Cell ‘abc‘.C3>, <Cell "abc".D3>)
# (<Cell ‘abc‘.B4>, <Cell ‘abc‘.C4>, <Cell "abc".D4>)

'''cell对象'''
# 属性
row # 单元格所在的行
column # 单元格坐在的列
value # 单元格的值
coordinate # 单元格的坐标
>>> excel2[‘abc‘].cell(row=1,column=2).coordinate
‘B1
>>> excel2[‘abc‘].cell(row=1,column=2).value
‘test‘
>>> excel2[‘abc‘].cell(row=1,column=2).row
1
>>> excel2[‘abc‘].cell(row=1,column=2).column
‘B‘