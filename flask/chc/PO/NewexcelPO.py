# -*- coding: utf-8 -*-
# *********************************************************************
# Author        : John
# Date          : 2020-12-8
# Description   : openpyxl 对象层
# openpyxl 官网 （http://openpyxl.readthedocs.org/en/latest/），支持Excel 2010 xlsx/xlsm/xltx/xltm 的读写，最新版本3.0.6（Oct 23, 2020）
# 1，请安装 openpyxl 3.0.0，其他版本如3.0.2使用中会报错。 pip3 install openpyxl == 3.0.0
# 报错：File "src\lxml\serializer.pxi", line 1652, in lxml.etree._IncrementalFileWriter.write TypeError: got invalid input value of type <类与实例 'xml.etree.ElementTree.Element'>, expected string or Element
# 解决方法: pip uninstall lxml   或更新最新openpyxl 3.0.7以上版本
# 2，如果文字编码是“gb2312” 读取后就会显示乱码，请先转成Unicode
# 3，openpyxl 的首行、首列 是 （1,1）而不是（0,0）
# 4，openpyxl 的NULL空值对应于python中的None，表示这个cell里面没有数据。
# 5，openpyxl 的numberic数字型，统一按照浮点数来进行处理，对应于python中的float
# 6，openpyxl 的string字符串型，对应于python中的unicode
# 7，默认情况下，openpyxl会将整个xlsx读入到内存中，方便处理。
# 8，openpyxl 操作大文件时可使用 Optimized reader 和 Optimized writer 两种模式，它们提供了流式的接口，速度更快，使我们可以用常量级的内存消耗来读取和写入无限量的数据。
# Optimized reader，打开文件使用use_iterators=True参数，如：wb = load_workbook(filename = 'haggle.xlsx',use_iterators=True)
# 9，openpyxl 读取大数据的效率没有 xlrd 高
# 10，openpyxl 与 xlsxwriter xlrd xlwt xlutils 的比较，这些库都不支持 excel 写操作，一般只能将原excel中的内容读出、做完处理后，再写入一个新的excel文件。
# 11，openpyxl常用模块用法：https://www.debug8.com/python/t_41519.html

# 参考：https://blog.csdn.net/m0_47590417/article/details/119082064
# https://blog.csdn.net/four91/article/details/106141274

# 绿色 = 00E400，黄色 = FFFF00，橙色 = FF7E00，红色 = FF0000，粉色 = 99004C，褐色 =7E0023
# 'c6efce = 淡绿', '006100 = 深绿'，'ffffff=白色', '000000=黑色'，'ffeb9c'= 橙色

# 用get_column_letter得到表格列的字母编号 https://www.pynote.net/archives/2269
# *********************************************************************

from openpyxl import load_workbook
import openpyxl, sys, platform, os
import openpyxl.styles
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Alignment,
)
from openpyxl.utils import get_column_letter
from datetime import date
from time import sleep
import psutil


from PO.SysPO import *
Sys_PO = SysPO()

"""
1.1 新建excel  newExcel()
"""


class NewexcelPO:
    def __init__(self, varFileName, *varSheetName):
        """
        1.1 新建excel(覆盖)
        :param varFileName: 文件名
        :param varSheetName: N个工作表
        # NewexcelPO.newExcel("d:\\444.xlsx")  # 新建excel默认一个Sheet1工作表
        # NewexcelPO.newExcel("d:\\444.xlsx", "mySheet1", "mySheet2","mySheet3")  # 新建excel生成三个工作表，默认在第一个mySheet1表。
        # 注意：如果文件已存在则会先删除后再新建！
        """

        # try:
        wb = openpyxl.Workbook()
        ws = wb.active
        if len(varSheetName) == 0:
            ws.title = "Sheet1"
        else:
            ws.title = varSheetName[0]
        for i in range(1, len(varSheetName)):
            wb.create_sheet(varSheetName[i])
        wb.save(varFileName)



if __name__ == "__main__":


    # 新建excel，默认生成一个工作表Sheet1
    # Newexcel_PO = NewexcelPO("./ExcelPO/KILL.xlsx", "mySheet1", "mySheet2", "mySheet3")
    Newexcel_PO = NewexcelPO("./ExcelPO/KILL.xlsx")

