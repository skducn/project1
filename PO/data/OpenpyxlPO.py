# -*- coding: utf-8 -*-
# *****************************************************************
# Author        : John
# Date          : 2019-1-26
# Description   : openpyxl1.py 表格处理（insertImage，writeExcel）
# 官网学习文档中文版3.0.0 http://yumos.gitee.io/openpyxl3.0/
# openpyxl 是一个 Python 库 用来读/写 Excel 2010 xlsx/xlsm/xltx/xltm 类型文件.
# 默认情况下openpyxl不防范quadratic blowup或 billion laughs的xml 攻击。为防范这些攻击，请安装defusedxml.
# 默认创建sheet的位置是在最后，但可选择插入的位置，如 ws2 = wb.create_sheet("Mysheet", 0) # 在最前端插入
# title的默认背景色是白色. 通过对Worksheet.sheet_properties.tabColor 属性赋值不同的 RRGGBB 颜色代码来改变title的背景色，如：ws.sheet_properties.tabColor = "1072BA"
# http://openpyxl.readthedocs.org/en/latest/
# https://openpyxl.readthedocs.io/en/stable/usage.html#write-a-workbook
# openpyxl（可读写excel表）专门处理Excel2007及以上版本产生的xlsx文件，xls和xlsx之间转换容易
# 注意：如果文字编码是“gb2312” 读取后就会显示乱码，请先转成Unicode
# openpyxl 的首行、首列 是 （1,1）而不是（0,0）
# NULL空值：对应于python中的None，表示这个cell里面没有数据。
# numberic： 数字型，统一按照浮点数来进行处理。对应于python中的float。
# string： 字符串型，对应于python中的unicode。
# 在默认情况下，openpyxl会将整个xlsx都读入到内存中，方便处理。
# 这使得操作大文件的时候，速度较慢，可以使用Optimized reader和Optimized writer。它们提供了流式的接口，速度更快。
# *****************************************************************

import os, time, json
from openpyxl import Workbook

# from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import openpyxl, os, datetime
from PO.TimePO import *
from PO.ExcelPO import *

Time_PO = TimePO()


from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
    AreaChart3D,
)


# 实例，新建第3个工作表 Data，遍历10行（从10到20），每行从 C-I 分别填入C-I
# ws3 = wb.create_sheet(title="Data")
# for row in range(10, 20):
#     for col in range(3, 10):
#         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))


class OpenpyxlPO:
    def __init__(self, file):
        self.file = file
        if os.path.exists(self.file):
            self.wb = openpyxl.load_workbook(self.file)
            self.sheets = self.wb.sheetnames
        else:
            self.wb = Workbook()
            self.wb.save(self.file)
            self.wb = openpyxl.load_workbook(self.file)
            self.sheets = self.wb.sheetnames

    # 所有工作表名
    def getSheetName(self):
        # 获取所有工作表名
        # print(Openpyxl_PO.getSheetName())  # ['test2', 'john', 'john1999999213', 'Sheet']
        return self.sheets

    # 总行数和总列数
    def getRowsColsNum(self, varSheet=0):
        # print(Openpyxl_PO.getRowsColsNum())  # 默认定位第1个Sheet，获取行和列总数
        # print(Openpyxl_PO.getRowsColsNum("john"))  # 定位john工作表，获取行和列总数
        # print(Openpyxl_PO.getRowsColsNum(1))  # 定位第2个Sheet，获取行和列总数
        # print(Openpyxl_PO.getRowsColsNum("john1212"))  # 定位一个不存在的工作表，返回None
        try:
            if isinstance(varSheet, int):
                ws = self.wb[self.sheets[varSheet]]
            elif isinstance(varSheet, str):
                ws = self.wb[varSheet]
            else:
                exit()
            rows = ws.max_row
            columns = ws.max_column
            return rows, columns
        except:
            return None

    # 获取某行某列所有值
    def getValues(self, varMode, varNum, varSheet=0):

        try:
            if isinstance(varSheet, int):
                ws = self.wb[self.sheets[varSheet]]
            elif isinstance(varSheet, str):
                ws = self.wb[varSheet]
            else:
                exit()
            if varMode == "row":
                columns = ws.max_column
                rowdata = []
                for i in range(1, columns + 1):
                    cellvalue = ws.cell(row=varNum, column=i).value
                    rowdata.append(cellvalue)
                return rowdata
            elif varMode == "col":
                rows = ws.max_row
                columndata = []
                for i in range(1, rows + 1):
                    cellvalue = ws.cell(row=i, column=varNum).value
                    columndata.append(cellvalue)
                return columndata
            else:
                return None
        except:
            return None

    # 设置某行某列所有值
    def setValues(self, varMode, varRowColNum, varValue, varSheet=0):
        # 设置某行或某列数据
        # print(Openpyxl_PO.setValues('row', 6, ll))  # 对第六行写入
        # print(Openpyxl_PO.setValues('col', 6, l2))  # 对第六列写入
        # print(Openpyxl_PO.setValues('', 6, l2))  # 如果varMode为空,返回None
        try:
            if isinstance(varSheet, int):
                ws = self.wb[self.sheets[varSheet]]
            elif isinstance(varSheet, str):
                ws = self.wb[varSheet]
            else:
                exit()
            try:
                for i in range(len(varValue)):
                    if varMode == "row":
                        ws.cell(row=varRowColNum, column=i + 1).value = varValue[i]
                    elif varMode == "col":
                        ws.cell(row=i + 1, column=varRowColNum).value = varValue[i]
                    else:
                        return None
                self.wb.save(self.file)
                return 0
            except:
                exit()
        except:
            return None

    # 获取某个单元格的值
    def getCellValue(self, row, column, varSheet=0):
        # print(Openpyxl_PO.getCellValue(1, 2))  # 默认定位第1个Sheet，获取第一行第二列的值
        # print(Openpyxl_PO.getCellValue(1, 2, 1))  # 定位第2个Sheet，获取第一行第二列的值
        # print(Openpyxl_PO.getCellValue(1, 2, "john"))  # 定位john工作表获取第一行第二列的值
        # print(Openpyxl_PO.getCellValue(1, -2))  # 错误的单元格号，返回None
        # print(Openpyxl_PO.getCellValue(1, 2,"john1212"))  # 错误不存在的工作表，返回None
        try:
            if isinstance(varSheet, int):
                ws = self.wb[self.sheets[varSheet]]
            elif isinstance(varSheet, str):
                ws = self.wb[varSheet]
            else:
                exit()
            cellvalue = ws.cell(row=row, column=column).value
            return cellvalue
        except:
            return None

    # 设置某个单元格的值
    def setCellValue(self, row, colunm, cellvalue, varSheet=0):
        # print(Openpyxl_PO.setCellValue(1, 1, "测试内容"))  # 默认定位第1个Sheet，对第1行第1列的值赋值。
        # print(Openpyxl_PO.setCellValue(1, 2, "我的家", 1))  # 定位第2个Sheet，对第1行第2列的值赋值
        # print(Openpyxl_PO.setCellValue(1, 3, "我的家", "john"))  # 定位john工作表，对第1行第3列的值赋值
        # print(Openpyxl_PO.setCellValue(1, -2, "我的家"))  # 错误的单元格号，返回None
        # print(Openpyxl_PO.setCellValue(1, 2, "我的家", "john1212"))  # 错误不存在的工作表，返回None
        try:
            if isinstance(varSheet, int):
                ws = self.wb[self.sheets[varSheet]]
            elif isinstance(varSheet, str):
                ws = self.wb[varSheet]
            else:
                exit()
            try:
                ws.cell(row=row, column=colunm).value = cellvalue
                self.wb.save(self.file)
                return 0
            except:
                exit()
        except:
            return None

    # 插入图片
    def setImage(self, varLocCell, varImage, varSheet=0):
        # 打开表格工作表, 在指定位置插入图片, 如果工作表名不存在,则新增工作表
        # print(Openpyxl_PO.setImage('d5', 'logo.png'))  # 默认定位第1个Sheet，在D5单元格处插入logo.png
        # print(Openpyxl_PO.setImage('k5', 'logo.png', "Sheet"))  # 默认定位第1个Sheet，在D5单元格处插入logo.png
        # print(Openpyxl_PO.setImage('d11', 'logo.png', 1))  # 默认定位第1个Sheet，在D5单元格处插入logo.png
        # print(Openpyxl_PO.setImage('11', 'logo.png', "Sheet"))  # 错误的单元格，返回None
        # print(Openpyxl_PO.setImage('d11', 'logo16666623.png', "Sheet"))  # 错误不存在的文件，返回None
        # print(Openpyxl_PO.setImage('d11', 'logo.png', "yoyo"))  # 如果工作表不存在,则新增工作表

        try:
            if isinstance(varSheet, int):
                ws = self.wb[self.sheets[varSheet]]
            elif isinstance(varSheet, str):
                if varSheet not in self.wb.sheetnames:  # 如果工作表名不存在,则新增工作表
                    self.wb.create_sheet(
                        index=0, title=varSheet
                    )  # index=0表示工作表排在第一个,默认新建后排在最后.
                    ws = self.wb[varSheet]
            else:
                exit()
            try:
                img = Image(varImage)
                try:
                    ws[varLocCell] = ""
                except:
                    print("error, varLoc单元格不存在!")
                    return None
                ws.add_image(img, varLocCell)  # 插入图片的起始位置在A3 ，缺省值是A1
            except:
                print("error, " + varImage + "不存在!")
                return None
            self.wb.save(self.file)
            return 0
        except:
            return None

        # if varMode == 'a':
        #     if os.path.exists(self.file):
        #         wb = openpyxl.load_workbook(self.file)
        #         ws = wb.worksheets[0]
        #     else:
        #         wb = Workbook()
        #         wb.save(varFileName)
        #         wb = openpyxl.load_workbook(varFileName)
        #         ws = wb.worksheets[0]
        # else:
        #     wb = Workbook()
        #     ws = wb.active
        # ws.title = varSheet

        return True

    # 设置行列值
    def setRowsColsValue(self, varDictData):
        # 字典方式批量设置行列值(追加),如果工作表不存在,则自动新建
        # print(Openpyxl_PO.setRowsColsValue(d))
        try:
            for k in varDictData:
                if k not in self.wb.sheetnames:  # 比较字典中工作表名是否在原有表格中，如果没有则新建工作表名。
                    ws2 = self.wb.create_sheet(
                        index=0, title=k
                    )  # index=0表示在最前端插入， ws3 = wb.create_sheet("Mysheet", -1) 表示插入倒数第二个位置
                    ws2 = self.wb.active
                    for row in range(len(varDictData[k])):
                        ws2.append(varDictData[k][row])
                else:
                    ws1 = self.wb[k]
                    for row in range(len(varDictData[k])):
                        ws1.append(varDictData[k][row])
        except:
            print("error, varDictData 数据格式有误，字典value应该是[[]],如{'Sheet1',[[1,2,3]]}")
            return None

        self.wb.save(self.file)
        return 0

        #
        # else:
        #     wb = Workbook()
        #     if len(varDictData) > 1:
        #         for k in varDictData:
        #             if tmp == 0:
        #                 # 新建第1个工作表，初始化数据
        #                 ws1 = wb.active
        #                 ws1.title = k
        #                 for row in range(len(varDictData[k])):
        #                     ws1.append(varDictData[k][row])
        #             else:
        #                 # 新建第2个工作表，初始化数据
        #                 ws2 = wb.create_sheet(title=k)
        #                 for row in range(len(varDictData[k])):
        #                     ws2.append(varDictData[k][row])
        #             tmp = tmp + 1
        #     else:
        #         # 新建第1个工作表，初始化数据
        #         for k in varDictData:
        #             ws1 = wb.active
        #             ws1.title = k
        #             for row in range(len(varDictData[k])):
        #                 ws1.append(varDictData[k][row])
        # else:
        #     wb = Workbook()
        #     if len(varDictData) > 1 :
        #         for k in varDictData:
        #             if tmp == 0 :
        #                 # 新建第1个工作表，初始化数据
        #                 ws1 = wb.active
        #                 ws1.title = k
        #                 ws1.sheet_properties.tabColor = "1072BA"
        #                 for row in range(len(varDictData[k])):
        #                     ws1.append(varDictData[k][row])
        #             else:
        #                 # 新建第2个工作表，初始化数据
        #                 ws2 = wb.create_sheet(title=k)
        #                 for row in range(len(varDictData[k])):
        #                     ws2.append(varDictData[k][row])
        #             tmp = tmp + 1
        #     else:
        #         # 新建第1个工作表，初始化数据
        #         for k in varDictData:
        #             ws1 = wb.active
        #             ws1.title = k
        #             for row in range(len(varDictData[k])):
        #                 ws1.append(varDictData[k][row])

    def areachart(self):
        # 2DArea Charts
        wb = Workbook()
        ws = wb.active
        rows = [
            ["Number具体数", "高地", "山丘"],
            [2, 40, 30],
            [3, 40, 25],
            [4, 50, 30],
            [5, 30, 10],
            [6, 25, 5],
            [7, 150, 10],
        ]
        for row in rows:
            ws.append(row)

        chart = AreaChart()
        chart.title = "我的内容上门"
        chart.style = 13
        chart.x_axis.title = "标题测试组"
        chart.y_axis.title = "比率"
        cats = Reference(ws, min_col=1, min_row=1, max_row=7)
        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "G10")
        wb.save("area.xlsx")

    def areachart3(self):
        wb = Workbook()
        ws = wb.active

        rows = [
            ["Number", "Batch 1", "Batch 2"],
            [2, 30, 40],
            [3, 25, 40],
            [4, 30, 50],
            [5, 10, 30],
            [6, 5, 25],
            [7, 10, 50],
        ]

        for row in rows:
            ws.append(row)

        chart = AreaChart3D()
        chart.title = "Area Chart"
        chart.style = 13
        chart.x_axis.title = "Test"
        chart.y_axis.title = "Percentage"
        chart.legend = None

        cats = Reference(ws, min_col=1, min_row=1, max_row=7)
        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "A10")

        wb.save("area3D.xlsx")

    # 折叠
    def outline(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet()
        ws.column_dimensions.group("A", "D", hidden=True)
        ws.row_dimensions.group(1, 10, hidden=True)
        wb.save(self.file)


def readExcel(varFileName):

    """
    编辑一个excel文件（xls,xlsx）
    :param varFileName: empty_book.xlsx
    :return:
    """

    from openpyxl import load_workbook

    wb = load_workbook(filename=varFileName)
    ws = wb["test2"]
    # // 遍历第2-3行，第2-4列数据
    for row in ws.iter_rows(
        min_row=2, max_row=3, min_col=2, max_col=4, values_only=True
    ):
        print(row)

    # // 获取某2列数据
    for row in ws.iter_rows(min_col=2, max_col=4, values_only=True):
        print(row)

    # # 4、遍历工作表所有内容，返回列表 ， 注意 表格中空单元输出None
    # rows = ws2.rows
    # # cols = ws2.columns
    # for row in rows:
    #     line = [col.value for col in row]
    #     print(line)
    #
    # # 4, 遍历工作表所有行和列， 获取指定的列数据
    # for i in range(1, ws2.max_row + 1):
    #     print(ws2.cell(row=i, column=1).value)

    # 5, 写单元格的值，三种方式
    # ws2['E1'] = 66
    # ws2.cell(row=2, column=5).value = 99  # (E2)
    # ws2.cell(row=3, column=5, value=100)  # (E3)
    # 支持用公式设定单元格的值(不建议)
    # ws2["A11"] = "=SUM(" + str(ws2['A1'].value) + ", " + str(ws2['A2'].value) + ")"
    # ws2["A12"] = "=SUM(B1:C1)"

    # 6, 逐行写 ws.append(iterable)
    # 功能：在当前sheet的最底部逐行追加一行
    # iterable 必须是 list,tuple,dict,range,generator 类型。
    # list类型，将list从头到尾顺序添加。
    # dict类型，按照相应的键添加相应的键值。
    # ws2.append(['This is A1', 'This is B1', 'This is C1'])
    # ws2.append({'A' : 'This is A1', 'C' : 'This is C1'})
    # ws2.append({1 : 'This is A1', 3 : 'This is C1'})
    # ws2.append((55,66,77))

    # 7, 新建工作表
    # ws0 = wb.create_sheet('SheetName', 0) # 插入在第1个前面
    # ws7 = wb.create_sheet('SheetName') # 插入到最后

    # 8, 读取单元格值：
    # print(ws2.cell(row=2,column=1).value)
    # print(ws2['A1'].value)

    # 9, 默认sheet的tab是白色的，通过RGB颜色来修改tab的颜色，如下是蓝色
    # ws2.sheet_properties.tabColor = '1072BA'

    # 10, 将excel数据存储为字典
    # dict1 = {}
    # list1 = []
    # for i in range(1, ws2.max_row + 1):
    #     for j in range(1, ws2.max_column + 1):
    #         w = ws2.cell(row=i, column=j).value
    #         list1.append(w)
    #     dict1[i] = list1
    #     list1 =[]
    # print('Total:%d' % len(dict1))
    # print(json.dumps(dict1,  ensure_ascii=False))
    # # Total:15
    # # {"1": [42, null, null, null, 66, null],
    # # "2": [8, null, null, null, 99, null],
    # # "3": ["=SUM(1, 1)", null, null, null, 100, null],
    # # "4": ["=SUM(4, 9)", null, null, null, null, null],
    # # "5": [null, null, null, null, null, 3.14],
    # # "6": [null, null, null, null, null, null],
    # # "7": [null, null, null, null, null, null],
    # # "8": [null, null, null, null, null, null],
    # # "9": [null, null, null, null, null, null],
    # # "10": [666, null, null, null, null, null],
    # # "11": ["=SUM(42, 8)", null, null, null, null, null],
    # # "12": ["This is A1", "This is B1", "This is C1", null, null, null],
    # # "13": ["This is A1", null, "This is C1", null, null, null],
    # # "14": ["This is A1", null, "This is C1", null, null, null],
    # # "15": [55, 66, 77, null, null, null]}

    # 11, 插入图片到表格，图片左上角定位在D10
    # from openpyxl.drawing.image import Image
    # img = Image('logo.png')
    # ws3.add_image(img, 'F20')

    # 12, 以时间格式命名excel名，如 Excel2019-01-27_17-15-16.xlsx
    # # excelPath = os.path.join(os.getcwd(), 'ExcelData','test','peter') 获取路径
    # # 获取日期时间
    # nameTime = time.strftime('%Y-%m-%d_%H-%M-%S')
    # excelName = 'Excel' + nameTime + '.xlsx'
    # # ExcelFullName= os.path.join(excelPath,excelName)
    # wb.save(excelName)

    wb.save(varFileName)


def mergeCell(varFileName):
    """
    Merge / Unmerge cells
    When you merge cells all cells but the top-left one are removed from the worksheet. See Styling Merged Cells for information on formatting merged cells.
    :param varFileName: empty_book.xlsx
    :return:
    """

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.unmerge_cells("A2:D2")
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
    ws.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
    wb.save(filename=varFileName)


def fold(varFileName):

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet()
    ws.column_dimensions.group("A", "D", hidden=True)
    ws.row_dimensions.group(1, 10, hidden=True)
    wb.save(varFileName)


def style(varFileName):

    from openpyxl.styles import Font, Color, Fill
    from openpyxl.styles.colors import RED
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    from openpyxl import load_workbook

    from openpyxl.styles import NamedStyle, Font, Border, Side

    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True, size=20)
    bd = Side(style="thick", color="FFBB00")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    wb = load_workbook(filename=varFileName)
    ws = wb.active
    c = ws["A1"]
    c.font = Font(size=15, color=RED)  # 文字颜色
    d = ws["B1"]
    d.font = Font(
        size=15,
        color="FFBB00",
        name="Arial",
        italic=True,
        bold=False,
        vertAlign=None,
        underline="none",
        strike=False,
    )

    # ws['D8'].style = highlight
    # a1 = ws['A1']
    # d4 = ws['D4']
    # ft = Font(color=colors.RED)
    # a1.font = ft
    # d4.font = ft
    # a1.font.italic = True  # is not allowed
    # # If you want to change the color of a Font, you need to reassign it::
    # a1.font = Font(color=colors.RED, italic=True)  # the change only affects A1
    wb.save(varFileName)


# *****************************************************************
if __name__ == "__main__":

    Openpyxl_PO = OpenpyxlPO("test2222.xlsx")  # 打开excle，如果找不到则自动创建，默认一个Sheet。

    # print(Openpyxl_PO.getSheetName())  # ['test2', 'john', 'john1999999213', 'Sheet']

    print(Openpyxl_PO.getRowsColsNum())  # 默认定位第1个Sheet，获取行和列总数
    print(Openpyxl_PO.getRowsColsNum("john"))  # 定位john工作表，获取行和列总数
    print(Openpyxl_PO.getRowsColsNum(1))  # 定位第2个Sheet，获取行和列总数
    print(Openpyxl_PO.getRowsColsNum("john1212"))  # 定位一个不存在的工作表，返回None
    #
    # print(Openpyxl_PO.getValues('row', 3))  # 默认定位第1个Sheet，获取第3行数据
    # print(Openpyxl_PO.getValues('row', 3, "john"))  # 定位john工作表，获取第3行数据
    # print(Openpyxl_PO.getValues('row', 3, 1))  # 定位第2个Sheet，获取第3行数据
    # print(Openpyxl_PO.getValues('row', -3))  # 错误的行数，返回None
    # print(Openpyxl_PO.getValues('row', 3, "john1212"))  # 错误不存在的工作表，返回None
    # print(Openpyxl_PO.getValues('col', 3))  # 默认定位第1个Sheet，获取第3行数据
    # print(Openpyxl_PO.getValues('col', 3, "john"))  # 定位john工作表，获取第3行数据
    # print(Openpyxl_PO.getValues('col', 3, 1))  # 定位第2个Sheet，获取第3行数据
    # print(Openpyxl_PO.getValues('col', -3))  # 错误的行数，返回None
    # print(Openpyxl_PO.getValues('col', 3, "john1212"))  # 错误不存在的工作表，返回None
    # print(Openpyxl_PO.getValues('', 3, "john1212"))  # 错误不存在的工作表，返回None

    # ll = ['', '姓名', '电话', '成绩', '学科']
    # l2 = ['john', 'yoyo', 'meter', 'haha']
    # print(Openpyxl_PO.setValues('row', 6, ll))  # 对第六行写入
    # print(Openpyxl_PO.setValues('col', 6, l2))  # 对第六列写入
    # print(Openpyxl_PO.setValues('', 6, l2))  # 如果varMode为空,返回None

    # print(Openpyxl_PO.getCellValue(1, 2))  # 默认定位第1个Sheet，获取第一行第二列的值
    # print(Openpyxl_PO.getCellValue(1, 2, 1))  # 定位第2个Sheet，获取第一行第二列的值
    # print(Openpyxl_PO.getCellValue(1, 2, "john"))  # 定位john工作表获取第一行第二列的值
    # print(Openpyxl_PO.getCellValue(1, -2))  # 错误的单元格号，返回None
    # print(Openpyxl_PO.getCellValue(1, 2,"john1212"))  # 错误不存在的工作表，返回None

    # print(Openpyxl_PO.setCellValue(1, 1, "测试内容"))  # 默认定位第1个Sheet，对第1行第1列的值赋值。
    # print(Openpyxl_PO.setCellValue(1, 2, "我的家", 1))  # 定位第2个Sheet，对第1行第2列的值赋值
    # print(Openpyxl_PO.setCellValue(1, 3, "我的家", "john"))  # 定位john工作表，对第1行第3列的值赋值
    # print(Openpyxl_PO.setCellValue(1, -2, "我的家"))  # 错误的单元格号，返回None
    # print(Openpyxl_PO.setCellValue(1, 2, "我的家", "john1212"))  # 错误不存在的工作表，返回None

    # print(Openpyxl_PO.setImage('d5', 'logo.png'))  # 默认定位第1个Sheet，在D5单元格处插入logo.png
    # print(Openpyxl_PO.setImage('k5', 'logo.png', "Sheet"))  # 默认定位第1个Sheet，在D5单元格处插入logo.png
    # print(Openpyxl_PO.setImage('d11', 'logo.png', 1))  # 默认定位第1个Sheet，在D5单元格处插入logo.png
    # print(Openpyxl_PO.setImage('11', 'logo.png', "Sheet"))  # 错误的单元格，返回None
    # print(Openpyxl_PO.setImage('d11', 'logo16666623.png', "Sheet"))  # 错误不存在的文件，返回None
    # print(Openpyxl_PO.setImage('d11', 'logo.png', "yoyo"))  # 如果工作表不存在,则新增工作表

    tableValues = [
        ["", "姓名", "电话", "成绩", "学科"],
        ["", "李雷", "15201062598", 19, datetime.datetime.now()],
        ["", "Marry", "15201062191", 28, Time_PO.getDate()],
        ["", "董承瑞", "13451062100", 38, "物理"],
        ["", "毛泽东", "15266606298", 14, "化学"],
        ["", "周恩来", "15201077791", 78, "美术"],
    ]
    # d = {"test": tableValues}
    d = {"john": [["测试组", "15201077791", 278, "美术3"]], "test2": tableValues}
    # print(Openpyxl_PO.setRowsColsValue(d))

    # Openpyxl_PO.areachart()
    Openpyxl_PO.areachart3()

    # readExcel('test2222.xlsx')

    # # mergeCell('empty_book.xlsx')
    # # fold('fold.xlsx')
    # # style('empty_book.xlsx')
    #
    # # insertImage('TEST111111.xlsx', 'Sheet1', 'd5', 'logo.png', 'a')
    # # insertImage('TEST111111.xlsx', 'Sheet1', 'd5', 'lgo.png'))
    # # insertImage('TEST111111.xlsx', 'Sheet1', 'd5', 'logo.png', 'w')
    #
    # getColValues("test2222.xlsx",2)

    # Openpyxl_PO.zhedie()
