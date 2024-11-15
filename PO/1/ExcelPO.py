# -*- coding: utf-8 -*-
# *********************************************************************
# Author        : John
# Date          : 2019-9-26
# Description   : excel 对象层

# python 处理 Excel 表格  http://www.cnblogs.com/snow-backup/p/4021554.html
# xlrd（读excel）表、xlwt（写excel）表、openpyxl（可读写excel表）
# xlrd读数据较大的excel表时效率高于 openpyxl
# Working with Excel Files in Python   下载地址：http://www.python-excel.org/

# python处理Excel，pythonexcel https://www.cnblogs.com/wanglle/p/11455758.html
# Python利用pandas处理Excel数据的应用 # https://www.cnblogs.com/liulinghua90/p/9935642.html

# python中处理excel表格，常用的库有xlrd（读excel）表、xlwt（写excel）表、openpyxl（可读写excel表）等。xlrd读数据较大的excel表时效率高于openpyxl，建议使用 xlrd 和 xlwt 两个库。
# openpyxl , xlsxwriter xlrd xlwt xlutils 下载地址：http://www.python-excel.org/
# 以上这些库都没有提供修改 excel 表格内容功能，一般只能将原excel中的内容读出、做完处理后，再写入一个新的excel文件。
#
# 常见问题1：处理 excel 表格时遇到的 unicode编码。
# python默认字符编码为unicode，所以从excel中读取的中文sheet或中文名时，会提示“UnicodeEncodeError: 'ascii' codec can't encode characters in position 0-2: ordinal not in range(128)。”
# 这是由于 windows里中文使用了gb2312编码方式，python误将其当作 unicode 和 ascii 来解码所造成的错误，使用“.encode('gb2312')”即可解决打印中文的问题。
# 可在文件名前加‘u’表示将该中文文件名采用unicode编码。

# 问题2：处理excel表格时遇到 excel中 时间显示错误。
# 有excel中，时间和日期都使用浮点数表示。如单元格中时间是 ‘2013年3月20日’，python输出为‘41353.0’；当其单元格格式改变为日期后，内容又变为了‘2013年3月20日’。
# 而使用xlrd读出excel中的日期和时间后，得到是的一个浮点数。所以当向excel中写入的日期和时间为一个浮点数也不要紧，只需将表格的表示方式改为日期和时间，即可得到正常的表示方式。
# excel中，用浮点数1表示1899年12月31日。

# python读取excel中单元格的内容返回的有5种类型
# ctype： 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
# 如 sh.cell(2, 1).ctype == 3  判断某一单元格内容是否是日期

# pip install pypiwin32   安装win32com
# Mac上是无法安装pypiwin32的
# *********************************************************************

from datetime import date
import xlrd, xlwt
from xlutils.copy import copy
from openpyxl import load_workbook
import pandas as pd
import openpyxl
import openpyxl.styles
from openpyxl.styles import PatternFill

# import win32com.client as win32
import openpyxl, sys, platform, os, psutil
from time import sleep

from PO.ColorPO import *

Color_PO = ColorPO()
from PO.CharPO import *

Char_PO = CharPO()
from PO.FilePO import *

File_PO = FilePO()

"""
1 新建excel（for xlsx） newExcel()

2.1 写操作（for xlsx） wrtCellValue()
2.2 批量写操作（for xlsx） wrtMoreCellValue()
2.3 写操作（by xlrd） 忽略未测试
2.4 删除行列 delRowColValues()
2.5 清空行列（for xlsx）clrRowColValues()

3.1 获取所有工作表名 l_getSheet()
3.2 获取工作表名 getSheet()
3.3 获取总行数和总列数 t_getRowColNums()
3.4 获取一行及单个值 l_getRowValues()
3.5 获取一列及单个值 l_getColValues()
3.6 获取单元格的值 getCellValue()
3.7 获取每行数据 l_getRowData()
3.8 获取指定列的行数据 l_getRowDataByPartCol()
3.9 获取每列数据 l_getColData()
3.10 获取某些列的列数据，可忽略多行 l_getColDataByPartCol()
3.11 获取任意两列的计算值（+ - * /） l_getCalcData(2,3,"求和",1,"4.电子健康档案规范建档率")

4 设置表格样式 set_style()

5 两表比较，输出差异 cmpExcel()
6 检查列值，返回行记录 l_getRecords()
"""


class ExcelPO:
    def __init__(self, file):
        self.file = file
        # 判断是不是xlsx文件，如果是xls则转换成xlsx,同时删除xls文件
        # if str(self.file).split(".")[1] == "xls":
        #     self.xls2xlsx(self.file)
        #     self.file = str(self.file).replace(".xls", ".xlsx")
        #     # 删除xls文件
        #     File_PO.delFile(file)
        # self.wb = xlrd.open_workbook(self.file)

        # self.wb1 = load_workbook(self.file)
        # self.sheetNames = self.wb1.sheetnames

        # # self.nameSizeColor = openpyxl.styles.Font(name="宋体", size=33, color="00CCFF")
        self.wb2 = openpyxl.load_workbook(self.file)
        self.sheetNames = self.wb2.sheetnames

    def save(self):
        self.wb2.save(self.file)

    # def __del__(self):
    #     self.wb1.save(self.file)

    # 1 新建excel（for xlsx）
    def newExcel(self, varFileName, *varSheetName):
        # 新建excel，生成N个工作表（默认一个Sheet1）
        # 注意：如果文件已存在则先删除再新建
        # Excel_PO.newExcel("d:\\test1.xlsx")  # 新建excel，默认生成一个工作表Sheet1
        # Excel_PO.newExcel("d:\\test1.xlsx", "mySheet1")  # 新建excel，生成一个工作表mySheet1
        # Excel_PO.newExcel("d:\\test1.xlsx", "mySheet1", "mySheet2", "mySheet3")  # 新建excel，生成三个工作表（mySheet1,mySheet2,mySheet3），默认定位在第一个mySheet1表。
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            if len(varSheetName) == 0:
                ws.title = "Sheet1"
            else:
                ws.title = varSheetName[0]
            for i in range(1, len(varSheetName)):
                wb.create_sheet(varSheetName[i])
            wb.save(varFileName)
        except:
            print(
                "errorrrrrrrrrr, call "
                + sys._getframe().f_code.co_name
                + "() from "
                + str(sys._getframe(1).f_lineno)
                + " row, error from "
                + str(sys._getframe(0).f_lineno)
                + " row"
            )

    # 2.1 写操作（for xlsx）
    def wrtCellValue(self, varRow, varCol, varContent, varSheet=0):
        # Excel_PO.writeXlsx(5, 3, "测1212试一下")  # 对第5行第3列写入内容
        # Excel_PO.writeXlsx(5, 3, "测1212试一下", "测试")  # 打开测试工作表，对第5行第3列写入内容
        try:
            if isinstance(varSheet, int):
                # 定位到sheet页,[0]为sheet页索引
                sh = self.wb1[self.sheetNames[varSheet]]
            elif isinstance(varSheet, str):
                sh = self.wb1[varSheet]
            else:
                print(
                    "errorrrrrrrrrr, call "
                    + sys._getframe().f_code.co_name
                    + "() from "
                    + str(sys._getframe(1).f_lineno)
                    + " row, error from "
                    + str(sys._getframe(0).f_lineno)
                    + " row"
                )
            sh.cell(row=varRow, column=varCol, value=varContent)
            self.wb1.save(self.file)
        except:
            Color_PO.consoleColor(
                "31",
                "31",
                "[ERROR] ",
                "call "
                + sys._getframe(1).f_code.co_name
                + " (line "
                + str(sys._getframe(1).f_lineno)
                + ", call "
                + sys._getframe(0).f_code.co_name
                + " from '"
                + sys._getframe().f_code.co_filename
                + "')",
            )

    # 2.2 批量写操作（for xlsx）
    def wrtMoreCellValue(self, varList_Row_Col_Content, varSheet=0):
        # 对多个元素写操作（只支持.xlsx）
        # Excel_PO.writeXlsxByMore([[1, "你好", "测试", "报告"], [2, "再见", "", "好了"]])
        # 第一行写入 "你好", "测试", "报告"
        # 第二行写入 "再见", "", "好了"

        try:
            if isinstance(varSheet, int):
                sh = self.wb1[self.sheetNames[varSheet]]
            elif isinstance(varSheet, str):
                sh = self.wb1[varSheet]
            else:
                print(
                    "errorrrrrrrrrr, call "
                    + sys._getframe().f_code.co_name
                    + "() from "
                    + str(sys._getframe(1).f_lineno)
                    + " row, error from "
                    + str(sys._getframe(0).f_lineno)
                    + " row"
                )
            for i in range(len(varList_Row_Col_Content)):
                for j in range(1, len(varList_Row_Col_Content[i])):
                    sh.cell(
                        row=varList_Row_Col_Content[i][0],
                        column=j,
                        value=varList_Row_Col_Content[i][j],
                    )
            self.wb1.save(self.file)
        except:
            print(
                "errorrrrrrrrrr, call "
                + sys._getframe().f_code.co_name
                + "() from "
                + str(sys._getframe(1).f_lineno)
                + " row, error from "
                + str(sys._getframe(0).f_lineno)
                + " row"
            )

    # 2.3 写操作（for xls） 忽略未测试
    def wrtXls(self, varFileName, varRow, varCol, varContent, varIgnore, varSheet=0):
        # 功能：写表操作，对单元格进行单个及批量写入（只支持.xls）
        # Excel_PO.writeXls("excel1.xls", "Sheet2", 2, 5, "令狐冲")  # 对Sheet2，第2行第5列写入令狐冲
        # Excel_PO.writeXls("excel1.xls", "Sheet2", "*", 5, "")  # 对Sheet2，清除第6列 (保留第一行，一般第一行是标题)
        # Excel_PO.writeXls("excel1.xls", 1, 2, '*', 'aaa')  # 对Sheet2，第3行写入aaa (保留第一列，一般第一列是标题)
        try:
            wb = xlrd.open_workbook(filename=varFileName)
            if isinstance(varSheet, int):  # 判断是int类型
                sh = self.wb.sheet_by_index(varSheet)
            elif isinstance(varSheet, str):
                sh = self.wb.sheet_by_name(varSheet)
            else:
                print(
                    "errorrrrrrrrrr, call "
                    + sys._getframe().f_code.co_name
                    + "() from "
                    + str(sys._getframe(1).f_lineno)
                    + " row, error from "
                    + str(sys._getframe(0).f_lineno)
                    + " row"
                )
            wbk = copy(self.wb)
            sheet = wbk.get_sheet(varSheet)
            if varRow == "*":
                for i in range(sh.nrows - varIgnore):
                    sheet.write(i + varIgnore, varCol - 1, varContent)
            elif varCol == "*":
                for i in range(sh.ncols - varIgnore):
                    sheet.write(varRow - 1, i + varIgnore, varContent)
            else:
                sheet.write(varRow - 1, varCol - 1, varContent)
            wbk.save(self.file)
        except:
            print(
                "errorrrrrrrrrr, call "
                + sys._getframe().f_code.co_name
                + "() from "
                + str(sys._getframe(1).f_lineno)
                + " row, error from "
                + str(sys._getframe(0).f_lineno)
                + " row"
            )

    # 2.4 删除行列（for xlsx）
    def delRowColValues(self, varType, varFrom, varSeries=1, varSheet=0):
        # 功能删除多行或多列。（for xlsx）
        # Excel_PO.delRowColForXlsx("excel3.xlsx", "col", 1, 2)  # 从第1列开始，连续删除2列
        # Excel_PO.delRowColForXlsx("excel3.xlsx", "row", 2, 3)  # 从第2行开始，连续删除3行
        try:
            if isinstance(varSheet, int):
                sh = self.wb1[self.sheetNames[varSheet]]
            elif isinstance(varSheet, str):
                sh = self.wb1[varSheet]
            if varType == "row":
                sh.delete_rows(varFrom + 1, varSeries)  # 删除从某行开始连续varSeries行
            elif varType == "col":
                sh.delete_cols(varFrom + 1, varSeries)  # 删除从某列开始连续varSeries行
            self.wb1.save(self.file)
        except:
            print(
                "errorrrrrrrrrr, call "
                + sys._getframe().f_code.co_name
                + "() from "
                + str(sys._getframe(1).f_lineno)
                + " row, error from "
                + str(sys._getframe(0).f_lineno)
                + " row"
            )

    # 2.5 清空行列（for xlsx）
    def clrRowColValues(self, varType, varNums, varSheet=0):
        # Excel_PO.clrRowColValues("row", 2)  # 清空第3行
        # Excel_PO.clrRowColValues("col", 3)  # 清空第4列
        # Excel_PO.clrRowColValues("col", 3, "mySheet3")  # 清空mySheet3的第4列
        try:

            # 判断 varSheet 是数字还是字符
            if isinstance(varSheet, int):  # 判断是int类型
                # 定位到相应sheet页,[0]为sheet页索引
                sh = self.wb2[self.sheetNames[varSheet]]
            elif isinstance(varSheet, str):
                sh = self.wb2[varSheet]
            else:
                print(
                    "errorrrrrrrrrr, call "
                    + sys._getframe().f_code.co_name
                    + "() from "
                    + str(sys._getframe(1).f_lineno)
                    + " row, error from "
                    + str(sys._getframe(0).f_lineno)
                    + " row"
                )
            if varType == "col":
                for i in range(sh.max_row):
                    sh.cell(row=i + 1, column=varNums + 1, value="")  # 清除第row行的第col列
            elif varType == "row":
                for i in range(sh.max_row):
                    sh.cell(row=varNums + 1, column=i + 1, value="")  # 清除第row行的第col列
            self.wb2.save(self.file)
        except:
            print(
                "errorrrrrrrrrr, call "
                + sys._getframe().f_code.co_name
                + "() from "
                + str(sys._getframe(1).f_lineno)
                + " row, error from "
                + str(sys._getframe(0).f_lineno)
                + " row"
            )

    # 2.6 清空所有内容（for xlsx）
    def clrAllValues(self, varSheet=0):

        # try:
        # 判断 varSheet 是数字还是字符
        if isinstance(varSheet, int):  # 判断是int类型
            # 定位到相应sheet页,[0]为sheet页索引
            sh = self.wb2[self.sheetNames[varSheet]]
        elif isinstance(varSheet, str):
            sh = self.wb2[varSheet]
        else:
            print(
                "errorrrrrrrrrr, call "
                + sys._getframe().f_code.co_name
                + "() from "
                + str(sys._getframe(1).f_lineno)
                + " row, error from "
                + str(sys._getframe(0).f_lineno)
                + " row"
            )

        print(sh.max_row)
        for i in range(sh.max_row):
            sh.cell(row=i + 1, column=i + 1, value="")

        # sh.cell(row=3, column=3, value="")

        self.wb2.save(self.file)
        # except:
        #     print("errorrrrrrrrrr, call " + sys._getframe().f_code.co_name + "() from " + str(
        #         sys._getframe(1).f_lineno) + " row, error from " + str(sys._getframe(0).f_lineno) + " row")

    # 3.1 获取所有工作表名
    def l_getSheet(self):
        try:
            return self.wb.sheet_names()
        except:
            print(
                "errorrrrrrrrrr, call "
                + sys._getframe().f_code.co_name
                + "() from "
                + str(sys._getframe(1).f_lineno)
                + " row, error from "
                + str(sys._getframe(0).f_lineno)
                + " row"
            )

    # 3.2 获取单个工作表名
    def getSheet(self, varIndex):
        # 通过 index 获取工作表名，0=第一个Sheet，1=第二个Sheet，-1表示从后往前数，以此类推
        # print(Excel_PO.getSheet(0))  # mySheet1
        # print(Excel_PO.getSheet(1))  # mySheet2
        # print(Excel_PO.getSheet(-1))  # mySheet3  //-1表示从后往前数
        try:
            return self.wb.sheet_by_index(varIndex).name
        except:
            print(
                "errorrrrrrrrrr, call "
                + sys._getframe().f_code.co_name
                + "() from "
                + str(sys._getframe(1).f_lineno)
                + " row, error from "
                + str(sys._getframe(0).f_lineno)
                + " row"
            )

    # 3.3 获取总行数和总列数
    def t_getRowColNums(self, varSheet=0):
        """返回元组(sh,行,列)"""
        # print(Excel_PO.t_getRowColNums())  # （4,3） //默认返回第一个工作表的总行数和总列数
        # print(Excel_PO.t_getRowColNums()[0])  # 4 //返回总行数
        # print(Excel_PO.t_getRowColNums()[1])  # 3 //返回总列数
        # print(Excel_PO.t_getRowColNums("mySheet3")[0])  # 8 //返回工作表名是测试的总行数
        # print(Excel_PO.t_getRowColNums("mySheet3")[1])  # 6 //返回工作表名是测试的总列数
        try:
            if isinstance(varSheet, int):
                sh = self.wb.sheet_by_index(varSheet)
            else:
                sh = self.wb.sheet_by_name(varSheet)
            # 检查 sheet是否导入完毕
            if self.wb.sheet_loaded(varSheet) == True:
                return sh, sh.nrows, sh.ncols
        except:
            print(
                "errorrrrrrrrrr, line "
                + str(sys._getframe(1).f_lineno)
                + "（ctrl + G）, 执行 "
                + sys._getframe().f_code.co_name
                + "(), 可能工作表名不存在！"
            )
            exit(0)

    def l_getRowColNums(self, filename, varSheet=0):
        """返回列表[行，列]"""
        list1 = []
        try:
            wb = xlrd.open_workbook(filename)
            if isinstance(varSheet, int):
                sh = self.wb.sheet_by_index(varSheet)
            else:
                sh = self.wb.sheet_by_name(varSheet)
            # 检查 sheet是否导入完毕
            if self.wb.sheet_loaded(varSheet) == True:
                rows = sh.nrows
                cols = sh.ncols
                list1.append(rows)
                list1.append(cols)
            return list1
        except:
            print(
                "errorrrrrrrrrr, line "
                + str(sys._getframe(1).f_lineno)
                + "（ctrl + G）, 执行 "
                + sys._getframe().f_code.co_name
                + "(), 参数错误！"
            )
            exit(0)

    # 3.4 获取一行及单个值
    def l_getRowValues(self, row, varSheet=0):
        # 获取行值,可使用切片获取某行某个单元值
        # print(Excel_PO.l_getRowValues(2))  # //默认第一个工作表，获取第3行所有值
        # print(Excel_PO.l_getRowValues(2, "测试"))  # //定位“测试”工作表
        # print(Excel_PO.l_getRowValues(2)[3])  # 获取列表中第3行第4列值

        list1 = []

        try:
            sh, rows, cols = self.t_getRowColNums(varSheet)
        except:
            print(
                "errorrrrrrrrrr, line "
                + str(sys._getframe(1).f_lineno)
                + "（ctrl + G）, 执行 "
                + sys._getframe().f_code.co_name
                + "() "
            )
            exit(0)
        try:
            for c in range(cols):
                # ctype： 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
                if sh.cell(row, c).ctype == 0:
                    cellvalue = ""
                elif sh.cell(row, c).ctype == 2:
                    cellvalue = sh.cell_value(rowx=row, colx=c)
                    # cellvalue = Char_PO.zeroByDel(cellvalue)
                elif sh.cell(row, c).ctype == 3:
                    cellvalue = xlrd.xldate_as_tuple(
                        sh.cell_value(rowx=row, colx=c), self.wb.datemode
                    )
                    cellvalue = date(*cellvalue[:3])
                elif sh.cell(row, c).ctype == 4:
                    if sh.cell_value(rowx=row, colx=c) == True:
                        cellvalue = 1
                    else:
                        cellvalue = 0
                else:
                    cellvalue = sh.cell_value(rowx=row, colx=c)

                list1.append(cellvalue)
            return list1
        except:
            print(
                "errorrrrrrrrrr, line "
                + str(sys._getframe(1).f_lineno)
                + "（ctrl + G）, 执行 "
                + sys._getframe().f_code.co_name
                + "(), 参数列表元素不能是0或负数！"
            )
            exit(0)

    # 3.5 获取一列及单个值
    def l_getColValues(self, column, varSheet=0):
        # 获取某列的值，可使用切片获取某行某个单元值
        # print(Excel_PO.l_getColValues(1))  # //默认第一个工作表，获取第2列所有值
        # print(Excel_PO.l_getColValues(3, "测试"))  # //定位“测试”工作表
        # print(Excel_PO.l_getColValues(1)[2])  # 获取列表中第2行第3列值
        list1 = []

        try:
            sh, rows, cols = self.t_getRowColNums(varSheet)
        except:
            print(
                "errorrrrrrrrrr, line "
                + str(sys._getframe(1).f_lineno)
                + "（ctrl + G）, 执行 "
                + sys._getframe().f_code.co_name
                + "() "
            )
            exit(0)
        try:
            for row in range(rows):
                # ctype： 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
                if sh.cell(row, column).ctype == 0:
                    cellvalue = ""
                elif sh.cell(row, column).ctype == 2:
                    cellvalue = sh.cell_value(rowx=row, colx=column)
                    # cellvalue = Char_PO.zeroByDel(cellvalue)
                elif sh.cell(row, column).ctype == 3:
                    cellvalue = xlrd.xldate_as_tuple(
                        sh.cell_value(rowx=row, colx=column), self.wb.datemode
                    )
                    cellvalue = date(*cellvalue[:3])
                elif sh.cell(row, column).ctype == 4:
                    if sh.cell_value(rowx=row, colx=column) == True:
                        cellvalue = 1
                    else:
                        cellvalue = 0
                else:
                    cellvalue = sh.cell_value(rowx=row, colx=column)

                list1.append(cellvalue)
            return list1
        except:
            print(
                "errorrrrrrrrrr, line "
                + str(sys._getframe(1).f_lineno)
                + "（ctrl + G）, 执行 "
                + sys._getframe().f_code.co_name
                + "(), 参数列表元素不能是0或负数！"
            )
            exit(0)

    # 3.6 获取单元格的值
    def getCellValue(self, row, column, varSheet=0):
        # print(Excel_PO.getCellValue(2, 2))

        sh, rows, cols = self.t_getRowColNums(varSheet)
        # 检查 sheet是否导入完毕
        if self.wb.sheet_loaded(varSheet) == True:
            # ctype： 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
            if sh.cell(row, column).ctype == 0:
                return ""
            if sh.cell(row, column).ctype == 2:
                cellvalue = sh.cell_value(rowx=row, colx=column)
                # cellvalue =Char_PO.zeroByDel(cellvalue)
                return cellvalue
            if sh.cell(row, column).ctype == 3:
                cellvalue = xlrd.xldate_as_tuple(
                    sh.cell_value(rowx=row, colx=column), self.wb.datemode
                )
                # cellvalue = date(*cellvalue[:3]).strftime('%Y/%d/%m')  #输出特定分隔符的日期
                cellvalue = date(*cellvalue[:3])
                return cellvalue
            elif sh.cell(row, column).ctype == 4:
                if sh.cell_value(rowx=row, colx=column) == True:
                    return 1
                else:
                    return 0
            else:
                cellvalue = sh.cell_value(rowx=row, colx=column)
                return cellvalue

    # 3.7 获取每行数据
    def l_getRowData(self, varSheet=0):

        """获取每行数据"""

        l_rowData = []  # 每行数据
        l_allData = []  # 所有行数据

        if isinstance(varSheet, int):
            self.sh = self.wb2[self.sheetNames[varSheet]]
        elif isinstance(varSheet, str):
            self.sh = self.wb2[varSheet]

        for cases in list(self.sh.rows):
            for i in range(len(cases)):
                l_rowData.append(cases[i].value)
            l_allData.append(l_rowData)
            l_rowData = []

        return l_allData

    # 3.8 获取某些列的行数据
    def l_getRowDataByPartCol(self, l_varCol, varSheet=0):
        # print(Excel_PO.l_getRowDataByPartCol([1, 2, 4]))  # 获取第2，3，5列的行数据

        """
        获取某些列的行数据
        """

        l_rowData = []  # 每行的数据
        l_allData = []  # 所有的数据

        if isinstance(varSheet, int):
            self.sh = self.wb1[self.sheetNames[varSheet]]
        elif isinstance(varSheet, str):
            self.sh = self.wb1[varSheet]

        max_r = self.sh.max_row
        for row in range(1, max_r + 1):
            try:
                for column in l_varCol:
                    l_rowData.append(self.sh.cell(row, column + 1).value)
                l_allData.append(l_rowData)
                l_rowData = []
            except:
                print(
                    "errorrrrrrrrrr, line "
                    + str(sys._getframe(1).f_lineno)
                    + ", in "
                    + sys._getframe().f_code.co_name
                    + "() "
                )
                print("建议：参数列表元素不能是0或负数")
                exit(0)

        return l_allData

    # 3.9 获取每列数据
    def l_getColData(self, varSheet=0):

        """获取每列数据"""

        l_colData = []  # 每列数据
        l_allData = []  # 所有行数据

        if isinstance(varSheet, int):
            self.sh = self.wb1[self.sheetNames[varSheet]]
        elif isinstance(varSheet, str):
            self.sh = self.wb1[varSheet]

        for cases in list(self.sh.columns):
            for i in range(len(cases)):
                l_colData.append(cases[i].value)
            l_allData.append(l_colData)
            l_colData = []

        return l_allData

    # 3.10 获取某些列的列数据，可忽略多行
    def l_getColDataByPartCol(self, l_varCol, l_varIgnoreRowNum, varSheet=0):
        # print(Excel_PO.l_getColDataByPartCol([1, 3], [0, 2]))  # 获取第二列和第四列的列值，并忽略第一行和第三行的行值。
        # l_data = Excel_PO.l_getColDataByPartCol([2], [], "7.家庭医生电子健康建档率")，获取第三列所有值。

        """
        获取某些列的列数据，可忽略某行
        l_varCol = 获取的列
        l_varIgnoreRowNum = 忽略的行
        """

        l_colData = []  # 每列的数据
        l_allData = []  # 所有的数据

        if isinstance(varSheet, int):
            self.sh = self.wb1[self.sheetNames[varSheet]]
        elif isinstance(varSheet, str):
            self.sh = self.wb1[varSheet]

        max_r = self.sh.max_row
        for col in l_varCol:
            try:
                for row in range(1, max_r + 1):
                    if row - 1 not in l_varIgnoreRowNum:
                        l_colData.append(self.sh.cell(row, col + 1).value)
                l_allData.append(l_colData)
                l_colData = []
            except:
                print(
                    "errorrrrrrrrrr, line "
                    + str(sys._getframe(1).f_lineno)
                    + ", in "
                    + sys._getframe().f_code.co_name
                    + "() "
                )
                print("建议：参数列表元素不能是0或负数")
                exit(0)

        return l_allData

    # 4 设置表格样式
    def set_style(self, name, height, bold=False):
        style = xlwt.XFStyle()
        font = xlwt.Font()
        font.name = name
        font.bold = bold
        font.color_index = 4
        font.height = height
        style.font = font
        return style

    # 5 两表比较，输出差异
    def cmpExcel(self, file1, file1Sheet, file2, file2Sheet):
        # 输出的列表中，第一个元素是序号。
        try:
            list1 = self.l_getRowColNums(file1, file1Sheet)
            list2 = self.l_getRowColNums(file2, file2Sheet)
            tmpList1 = []
            tmpList2 = []
            mainList1 = []
            mainList2 = []
            wb = load_workbook(file1)
            wk_sheet = wb[file1Sheet]
            for i in range(1, list1[0] + 1):
                tmpList1.append(i)
                for j in range(1, list1[1] + 1):
                    tmpList1.append(wk_sheet.cell(row=i, column=j).value)
                mainList1.append(tmpList1)
                tmpList1 = []
            # print(mainList1)

            wb = load_workbook(file2)
            wk_sheet = wb[file2Sheet]
            for i in range(1, list2[0] + 1):
                tmpList2.append(i)
                for j in range(1, list1[1] + 1):
                    tmpList2.append(wk_sheet.cell(row=i, column=j).value)
                mainList2.append(tmpList2)
                tmpList2 = []
            # print(mainList2)

            a = [x for x in mainList1 if x in mainList2]  # 两个列表中都存在
            b = [y for y in (mainList1) if y not in a]  # 两个列表中的不同元素，输出 mainList1 中差异部分
            c = [y for y in (mainList2) if y not in a]  # 两个列表中的不同元素，输出 mainList2 中差异部分

            if b == []:
                print("ok，两表一致")
            else:
                return file1, b, file2, c
                # print(file1 + " => " + str(b))
                # print(file2 + " => " + str(c))

        except:
            print(
                "errorrrrrrrrrr, call "
                + sys._getframe().f_code.co_name
                + "() from "
                + str(sys._getframe(1).f_lineno)
                + " row, error from "
                + str(sys._getframe(0).f_lineno)
                + " row"
            )

    # 6 检查列值，返回行记录
    def l_getRecords(self, varFileName, varSheet, varCol, varValue):

        list1 = []
        list2 = []
        wb = xlrd.open_workbook(varFileName)
        if isinstance(varSheet, int):
            sh = wb.sheet_by_index(varSheet)
        else:
            sh = wb.sheet_by_name(varSheet)
        if wb.sheet_loaded(varSheet) == True:  # 检查 sheet是否导入完毕，返回True 或 False
            rows = sh.nrows
            cols = sh.ncols
            for row in range(rows):
                if sh.cell_value(row, varCol) == varValue:
                    list1.append(row + 1)
                    for column in range(cols):
                        # ctype： 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
                        if sh.cell(row, column).ctype == 0:
                            cellvalue = ""
                        elif sh.cell(row, column).ctype == 2:
                            cellvalue = sh.cell_value(rowx=row, colx=column)
                            # cellvalue = Char_PO.zeroByDel(cellvalue)
                        elif sh.cell(row, column).ctype == 3:
                            cellvalue = xlrd.xldate_as_tuple(
                                sh.cell_value(rowx=row, colx=column), self.wb.datemode
                            )
                            cellvalue = date(*cellvalue[:3])
                        elif sh.cell(row, column).ctype == 4:
                            if sh.cell_value(rowx=row, colx=column) == True:
                                cellvalue = 1
                            else:
                                cellvalue = 0
                        else:
                            cellvalue = sh.cell_value(rowx=row, colx=column)

                        list1.append(cellvalue)

                    list2.append(list1)
                    list1 = []
        return list2

    # 7 设置单元格背景色
    def setCellColor(self, row, col, varColor, varSheet=0):
        # Excel_PO.setCellColor(6, 7, "FF0000")   设置单元格第六行第七列背景色为红色
        # 绿色 = 00E400，黄色 = FFFF00，橙色 = FF7E00，红色 = FF0000，粉色 = 99004C，褐色 =7E0023
        if isinstance(varSheet, int):
            sh = self.wb2[self.sheetNames[varSheet]]
        else:
            sh = self.wb2[varSheet]
        style = PatternFill("solid", fgColor=varColor)
        sh.cell(row, col).fill = style

    # 8 xls转xlsx
    def xls2xlsx(self, varFile):
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        wb = excel.Workbooks.Open(varFile)
        wb.SaveAs(
            varFile + "x", FileFormat=51
        )  # FileFormat = 51 is for .xlsx extension
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()

    # 17 关闭excel进程
    def closeExcelPid(self):
        pids = psutil.pids()
        for pid in pids:
            try:
                p = psutil.Process(pid)
                # print('pid=%s,pname=%s' % (pid, p.name()))
                # 关闭excel进程
                if p.name() == "EXCEL.EXE":
                    cmd = "taskkill /F /IM EXCEL.EXE"
                    os.system(cmd)
                    sleep(2)
            except Exception as e:
                pass


if __name__ == "__main__":

    Excel_PO = ExcelPO("ExcelPO/fold.xlsx")

    # Excel_PO = ExcelPO("d:\\1.xlsx")
    # Excel_PO.setCellColor(11, 7, "00E400")
    # Excel_PO.setCellColor(12, 7, "FFFF00")
    # Excel_PO.setCellColor(13, 7, "FF7E00")
    # Excel_PO.setCellColor(14, 7, "FF0000")
    # Excel_PO.setCellColor(15, 7, "99004C")
    # Excel_PO.save()

    # print("1，新建excel（by openpyxl）".center(100, "-"))
    # Excel_PO.newExcel("d:\\123.xlsx")  # 新建excel，默认生成一个工作表Sheet1
    # Excel_PO.newExcel("d:\\test1.xlsx", "mySheet1", "mySheet2", "mySheet3")  # 新建excel，生成三个工作表（mySheet1,mySheet2,mySheet3），默认定位在第一个mySheet1表。
    #
    # print("2.1 写操作（只支持.xlsx）".center(100, "-"))
    # Excel_PO.wrtCellValue(5, 3, "测1212试一下")  # 对第5行第3列写入内容
    # Excel_PO.wrtCellValue(5, 3, "测1212试一下", "mySheet2")  # 打开测试工作表，对第5行第3列写入内容
    #
    # print("2.2 批量写操作（只支持.xlsx）".center(100, "-"))
    # Excel_PO.wrtMoreCellValue([[7,"你好","测试","报告"], [9,"再见", "", "好了"]])
    #
    # print("2.3 单个元素及批量行列写操作（by xlrd）".center(100, "-"))
    # Excel_PO1 = ExcelPO("d:\\123.xls")
    # Excel_PO1.wrtXls(2, 5, "令狐冲","")  # 对 Sheet2 的第2行第5列写入令狐冲
    # # Excel_PO.wrtXls("d:\\test.xls", "Sheet1", "*", 5, "c", 2)  # 对 Sheet2 的第5列写入c，忽略前2行
    # # Excel_PO.wrtXls("d:\\test.xls", "Sheet1", 4, '*', 'b', 3)  # 对 Sheet2 的第4行写入b，忽略前3列
    #
    # print("2.4 删除行列".center(100, "-"))
    # Excel_PO.delRowColValues("col", 2, 1)  # 默认第一个工作表，删除第3列
    # Excel_PO.delRowColValues("col", 2, 3)  # 默认第一个工作表，删除第3，4，5列 （从第三列开始连续3列）
    # Excel_PO.delRowColValues("col", 0, 1, "mySheet3")  # 对mySheet3工作表，删除第1列
    # Excel_PO.delRowColValues("row", 3, 2)  # 删除第4、5行

    # print("2.5 清空行列".center(100, "-"))
    # Excel_PO.clrRowColValues("row", 2)  # 清空第3行
    # Excel_PO.clrRowColValues("col", 3)  # 清空第4列
    # Excel_PO.clrRowColValues("col", 3, "mySheet3")  # 清空mySheet3的第4列

    # print("2.6 清空所有内容".center(100, "-"))
    Excel_PO.clrAllValues()

    # print("3.1 获取所有工作表名".center(100, "-"))
    # print(Excel_PO.l_getSheet())  # ['mySheet1', 'mySheet2', 'mySheet3']
    # print(Excel_PO.l_getSheet()[1])  # ['mySheet1', 'mySheet2', 'mySheet3']
    #
    # print("3.2 获取单个工作表名".center(100, "-"))
    # print(Excel_PO.getSheet(0))  # mySheet1
    # print(Excel_PO.getSheet(1))  # mySheet2
    # print(Excel_PO.getSheet(-1))  # mySheet3  //-1表示从后往前数

    # print("3.3 获取总行数和总列数".center(100, "-"))
    # print(Excel_PO.t_getRowColNums())  # （4,3） //默认返回第一个工作表的总行数和总列数
    # print(Excel_PO.t_getRowColNums()[0])  # 4 //返回总行数
    # print(Excel_PO.t_getRowColNums()[1])  # 3 //返回总列数
    # print(Excel_PO.t_getRowColNums("测试")[0])  # 8 //返回工作表名是测试的总行数
    # print(Excel_PO.t_getRowColNums("测试")[1])  # 6 //返回工作表名是测试的总列数

    # print("3.4 获取一行及单个值".center(100, "-"))
    # print(Excel_PO.l_getRowValues(2))  # 获取第3行值
    # print(Excel_PO.l_getRowValues(2)[3])  # 获取列表中第3行第4列值
    #
    # print("3.5 获取一列及单个值".center(100, "-"))
    # print(Excel_PO.l_getColValues(1))   # 获取第2列所有值
    # print(Excel_PO.l_getColValues(1)[2])  # 获取列表中第2行第3列值
    #
    # print("3.6 获取单元格的值".center(100, "-"))
    # print(Excel_PO.getCellValue(2, 2))

    # print("3.7 获取每行数据".center(100, "-"))
    # print(Excel_PO.l_getRowData())
    #
    # print("3.8 获取指定列的行数据".center(100, "-"))
    # print(Excel_PO.l_getRowDataByPartCol([1,2,4]))   # 获取第2，3，5列的行数据
    #
    # print("3.9 获取每列数据".center(100, "-"))
    # print(Excel_PO.l_getColData("test_case"))

    # print("3.10 获取某些列的列数据，可忽略多行".center(100, "-"))
    # print(Excel_PO.l_getColDataByPartCol([1, 3], [0, 2]))   # 获取第二列和第四列的列值，并忽略第一行和第三行的行值。

    # # print("5 两表比较，输出差异".center(100, "-"))
    # file1,list1,file2,list2 = Excel_PO.cmpExcel("d:\\test1.xlsx", "mySheet1", "d:\\test2.xlsx", "mySheet1")
    # print(file1 + ">"*50)
    # for l in list1:
    #     print(l)
    # print("\n" + file2 + ">"*50)
    # for l in list2:
    #     print(l)

    # print("6 检查列值，返回行记录）".center(100, "-"))
    # list1 = (Excel_PO.l_getRecords("D:\\test1.xlsx", "mySheet1", 2, "a"))  # 检查第三列，如果有a值，则输出哪一行所有数据。
    # for l in list1:
    #     print(l)
