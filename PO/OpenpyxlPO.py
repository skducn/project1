# -*- coding: utf-8 -*-
# *********************************************************************
# Author        : John
# Date          : 2020-12-8
# Description   : openpyxl 对象层
# *********************************************************************
# 官网：http://openpyxl.readthedocs.org/en/latest/
# openpyxl 工作原理：将整个xlsx读入内存，方便处理。
# openpyxl 操作大文件时可使用 Optimized reader 和 Optimized writer 两种模式，它们提供了流式的接口，速度更快，使我们可以用常量级的内存消耗来读取和写入无限量的数据。
# Optimized reader 可用 use_iterators=True参数打开，如：wb = load_workbook(filename = 'haggle.xlsx',use_iterators=True)
# openpyxl 读取大数据的效率没有 xlrd 高。
# xlsxwriter xlrd xlwt xlutils 这些库都不支持 excel 写操作，一般只能将原excel中的内容读出、做完处理后，再写入一个新的excel文件。
# 扩展学习 xlwings https://blog.csdn.net/qfxietian/article/details/123822358

# 支持excel2010 .xlsx / .xlsm / .xltx / .xltm格式的文件
# 首行、首列 是（1,1）而不是（0,0）
# NULL表示空值，即cell里面没有数据 ，等同于 python中的None
# numberic 数字型 = python中的float
# string 字符串型 = python中的unicode

# 使用方法
# 参考：openpyxl常用模块用法 https://www.debug8.com/python/t_41519.html
# 基础使用方法：https://blog.csdn.net/four91/article/details/106141274
# 高级使用方法：https://blog.csdn.net/m0_47590417/article/details/119082064

# 安装包
# pip3 install openpyxl == 3.0.0
# 注意！：其他版本（如3.0.2使用中会报错）如有报错，请安装3.0.0

# 报错
# 如：File "src\lxml\serializer.pxi", line 1652, in lxml.etree._IncrementalFileWriter.write TypeError: got invalid input value of type <类与实例 'xml.etree.ElementTree.Element'>, expected string or Element
# 解决方法: pip uninstall lxml   及更新 openpyxl 版本，3.0.7以上

# 乱码
# gb2312 文字编码，在读取后会显示乱码，需转换成 Unicode

# 颜色
# 颜色码对照表（RGB与十六进制颜色码互转） https://www.sioe.cn/yingyong/yanse-rgb-16/
# 绿色 = 00E400，黄色 = FFFF00，橙色 = FF7E00，红色 = FF0000，粉色 = 99004C，褐色 =7E0023,'c6efce = 淡绿', '006100 = 深绿'，'ffffff=白色', '000000=黑色'，'ffeb9c'= 橙色

# 表格列 A，B，C 与 1，2，3 互转
from openpyxl.utils import get_column_letter, column_index_from_string
# get_column_letter(2)  # 'B'
# a = column_index_from_string('B')  # 2
# *********************************************************************

# todo [菜单]
"""
【工作表】
1.1 新建 Openpyxl_PO = OpenpyxlPO("1212.xlsx")
		Openpyxl_PO = OpenpyxlPO("1212.xlsx",l_sheet=['Sheet1','Sheet2','Sheet3'])
1.2 打开 open()
1.3 保存 save()
1.4 工作表 sh()
1.5 获取工作表 getL_sheet()
1.6 切换工作表 switchSheet("Sheet2") 
1.7 添加工作表 addSheet("Sheet1", overwrite=True)
1.8 删除工作表 delSheet("Sheet1")
1.9 重命名工作表 renameSheet("sheet1", "sheet2")

【操作数据】
2.0 在第N行前插入多行空白（优化版）
	insertNullRow(3)  在第3行前插入1行空白
	insertNullRow(3，5)  在第3行前插入5行空白
2.1 在第N列前插入多列空白（优化版）
	insertNullCol(3) 在第3列前插入1列空白
	insertNullCol(3,5)  在第3列前插入5列空白
2.2 设置单元格
	setCell(1, 2, "hello") # 将第一行B列写入hello
	setCell(1, 'B', "john") # 将第一行第三列写入john
2.3 插入行数据 insertRow({2: ["金浩", "101", "102"]})
2.4 更新行数据 setRow({2: ["金浩", "101", "102"], 5: ["yoyo", "123", "666"]})
2.5 追加行数据 appendRow([['姓名', '电话'], ['毛泽东', 15266606298]])
2.6 插入列数据 insertCol({"a": ["姓名", "张三", "李四"], "c": ["年龄", "55", "34"]})
2.7 更新列数据 setCol({"A": ["公司", "百度"], "F": ["学校", "清华大学"]})
2.8 追加列数据 appendCol([["姓名", "张三", "李四"], ["年龄", "55", "34"]])
2.9 清空行 clsRow(2)  # 清空第2行
2.10 清空列 
    clsCol(2, clear_header=True)  # 清空第2列
    clsCol(2, clear_header=False)  # 清空第2列,保留列标签
2.11 删除连续行 delRow(2, 3)  # 删除从第二行开始连续三行数据 （即删除2，3，4行）
2.12 删除连续列 delCol(2, 3)  # 删除从第二列开始连续三列数据 （即删除2，3，4列）
2.13 移动区域 moveBlock(rows, cols, 'C1:D2')
2.14 将excel中标题（第一行字段）排序（从小打大）sortColHeader()

【样式】
3.0 设置列宽 
	setColWidth(1, 20) # 设置第1列宽度20
	setColWidth('C', 30) # 设置第f列宽度30
3.1 设置行高与列宽
	setRowColSize([(3, 30), (5, 40)], None)  # 设置第三行行高30，第五行行高40
	setRowColSize(None, [('a', 22), ('C', 33)])  # 设置第a列宽22, 第c列宽33
	setRowColSize([(3, 30), (5, 40)], [('a', 22), ('C', 33)])  # 设置第三行行高30，第五行行高40, 设置第a列宽22, 第c列宽33
3.2 设置所有单元格的行高与列宽 setAllSize(30, 20) //设置所有单元格高30，宽20
3.3 设置自动换行 setWordWrap()
3.4 冻结窗口 setFreezePanes(）
	setFreezePanes('A2') # 冻结首行    A表示不冻结列，2表示冻结首行
	setFreezePanes('B1')  # 冻结第A列。  B表示冻结A列，1表示不冻结行
	setFreezePanes('B2')  # 冻结第1行和第A列。
	setFreezePanes('C3')  # 冻结第1、2行和第B列。  C表示冻结A、B列，3表示冻结第1、2行
3.5 设置单元格对齐样式  setAlignment(5, 4, 'center', 'center')
	setAlignment(5, 4, horizontal="center", vertical="center") # 设置第5行第4列居中对齐
	setAlignment(2, "E", vertical="top", wrap_text=True) # 设置第5行F列顶部对齐并自动换行
3.6 设置单行多列对齐样式 
	setRowColAlignment(5, [1,4], 'center', 'center') # 第5行第1,4列居中
	setRowColAlignment(1, ["c", "e"], 'center', 'top')  # 第一行第c,e列居中
	setRowColAlignment(1, "all", 'center', 'center')  # 第1行全部居中
3.7 设置所有单元格对齐样式 setAllAlignment('center', 'center')
3.8 设置筛选列  
	setFilterCol("all") # 全部筛选, 
	setFilterCol("") # 取消筛选 
	setFilterCol("A2") # 对A2筛选 
3.9 设置单元格字体（字体、字号、粗斜体、下划线、颜色） 
setCellFont(1, 1, name=u'微软雅黑', size=16, bold=True, italic=True, color="000000")
3.10 设置单行多列字体
	setRowColFont(2, ["b", "D"],name="微软雅黑",size=16, bold=False, italic=False, color="000000")  # 第1行第b-h列
	setRowColFont(3, "all")  # 第3行
3.11 设置所有单元格字体  setAllCellFont(color="000000")
3.12 设置单元格边框 setBorder(1, 2, left = ['thin','ff0000'], right = ['thick','ff0000'], top = ['thin','ff0000'],bottom = ['thick','ff0000'])
3.13 设置单元格填充渐变色 setGradientFill(3, 3, stop=["FFFFFF", "99ccff", "000000"])
3.14 设置单元格背景色".center(100, "-"))
	setBackgroundColor(2, 1, "ff0000")  # 设置第五行第1列设置红色
	setBackgroundColor(2, "c", "0000FF")  # 设置第五行e列设置红色
	setBackgroundColor(2, 1, clear_color=True)  # 清除第5行第1列的背景色
	setBackgroundColor(2, "c", clear_color=True)  # 清除第5行d列的背景色
3.15 设置单行多列背景色
	setRowColBackgroundColor(2, ['b', 'd'], "ff0000") # 设置第2行第b，c，d列背景色
	etRowColBackgroundColor(1, "all", "ff0000")  # 设置第7行所有列背景色
	setRowColBackgroundColor(2, ['B', 'D'], clear_color=True)  # 清除第2行第b，c，d列背景色
	setRowColBackgroundColor(1, "all", clear_color=True)  #  清除第1行所有列背景色
3.16 设置所有单元格背景色
	setAllBackgroundColor("ff0000")  # 设置所有单元格背景色
	setAllBackgroundColor(clear_color=True)  # 清除所有单元格背景色
3.17 设置整行(可间隔)背景色  
	setBandRowsColor(5, 0, "ff0000")  # 从第5行开始每行颜色标红色
	setBandRowsColor(3, 1, "ff0000")  # 从第3行开始每隔1行颜色标红色
	setBandRowsColor(3, 1, clear_color=True) # 清除从第3行开始每隔1行的背景色
	setBandRowsColor(5, 0, clear_color=True) # 清除从第5行开始每行颜色标红色
3.18 设置整列(可间隔)背景色 
	setBandColsColor(2, 0, "ff0000")  # 从第2列开始每列颜色为红色
	setBandColsColor(2, 1, "ff0000")  # 从第2列开始每隔1列设置颜色为红色
	setBandColsColor(2, 1, clear_color=True) # 清除从第2列开始每隔1列的背景色
	setBandColsColor(2, 0, clear_color=True)  # 从第2列开始每列颜色为红色
3.19 设置工作表背景颜色 
	setSheetColor("FF0000")
	setSheetColor(clear_color=True)	

【获取】
4.1 获取总行列数 getL_shape()  # [5,10]
4.2 获取单元格的值 getCell(3,2)  //获取第3行第2列的值
4.3 获取一行数据 getL_row(2) # 获取第2行值
4.4 获取一列数据 
	getL_col(2))  # ['高地', 40, 44, 50, 30, 25, 150]
	getL_col('B'))  # ['高地', 40, 44, 50, 30, 25, 150]
4.5 获取行数据 getLL_row()  # [['状态', '名字'],['ok', 'jinhao']...]

4.6 获取带行号的行数据 getD_rowNumber_row()  # { 2 : ['状态', '名字'], 3 : ['ok', 'jinhao']...}
4.7 获取部分列的行数据 
	getLL_rowOfPartialCol([1, 3])   # [['Number具体数', 'jinhaoyoyo'], [2, 30], [3, 10]] //获取1和3列的行数据
	getLL_rowOfPartialCol(['a', 'C']))  # 同上
	getLL_rowOfPartialCol(["A", 3]))   # 同上
4.8 获取带行号的部分列的行数据 
	getD_rowNumber_rowOfpartialCol([1, 3]))   # {1: ['Number具体数', 'jinhaoyoyo'], 2: [2, 30], 3: [3, 25], 4: [4, 30], 5: [5, 10], 6: [6, 5], 7: [7, 10]}
	getD_rowNumber_rowOfpartialCol([1, 'C']) 
	getD_rowNumber_rowOfpartialCol(['a', 'C'])) 
4.9 获取每列数据 getLL_col() 
4.10 获取带列序号的每列数据 getD_colNumber_col()  # { 2 : ['状态', '名字'], 3 : ['ok', 'jinhao']...}
4.11 获取带列字母的每列数据 getD_colLetter_col()  # { 'a' : ['状态', '名字'], 'b' : ['ok', 'jinhao']...}
4.12 获取列标签的序号 getL_columnHeaderNumber(['测试'，‘开发’])  # [2，4]
4.13 获取列标签的字母 getL_columnHeaderLetter(['测试'，‘开发’])  # ['A', 'C']
4.14 将标题转列字典序列 getD_colNumber_columnTitle(['测试'，‘开发’]）# {2: '姓名', 5: '性别'}
4.15 将标题转列字典字母 getD_colLetter_columnTitle(['测试'，‘开发’]）# {'B': '姓名', 'E': '性别'}
4.16 获取部分列的列值(可忽略多行) getLL_partialColOfPartialCol([1, 3], [1, 2]))   # 获取第二列和第四列的列值，并忽略第1，2行的行值。
4.17 获取单元格的坐标 getCoordinate(2, 5))   # E2
4.18 获取所有数据的坐标 getDimensions())  # A1:E17

【两表比较】
5.1 两个excel的sheet进行比较，输出有差异值。（两表标题与行数必须一致）
Openpyxl_PO = OpenpyxlPO("./data/11.xlsx")
Openpyxl_PO2 = OpenpyxlPO("./data/22.xlsx")
getD_excel_cell_By_Diff(Openpyxl_PO.getLL_row("hello_标题升序"), Openpyxl_PO2.getLL_row("hello_标题升序")))
5.2 对一张表的两个sheet进行数据比对，对第一张表差异数据标注颜色 
    setColorByDiff("Sheet1", "Sheet2")
5.3 对一张表的两个sheet进行数据比对，将结果写入第一个sheet ".center(100, "-"))
    genSheetByDiff("hello1", "hello2")  //结果写入 hello1_hello2_diff



"""

from openpyxl import load_workbook
from datetime import date
from time import sleep
import psutil
# import xlwings as xw

import openpyxl, platform, os, pdfplumber
import numpy as np
# import openpyxl.styles
from openpyxl.styles import (Font, PatternFill, GradientFill, Border, Side, Protection, Alignment)
from openpyxl.utils import get_column_letter, column_index_from_string

from PO.ListPO import *
List_PO = ListPO()

from PO.ColorPO import *
Color_PO = ColorPO()

# print(openpyxl.__version__)
import subprocess
import pandas as pd


class OpenpyxlPO:

    def __init__(self, pathFile, l_sheet=[]):

        self.file = pathFile
        self._wb = None  # 延迟加载标志
        self.l_sheet = l_sheet

        # 检查文件是否存在，若不存在则根据 l_sheet 创建新文件
        if not os.path.exists(self.file):
            wb = openpyxl.Workbook()
            ws = wb.active

            # 根据 l_sheet 参数设置工作表名称
            if not self.l_sheet:
                ws.title = "Sheet1"
            else:
                ws.title = self.l_sheet[0]
                for sheet_name in self.l_sheet[1:]:
                    wb.create_sheet(sheet_name)

            # 保存文件并输出提示信息
            wb.save(self.file)
            print(f"已创建文件: {self.file}")

        # 延迟加载工作簿属性
        @property
        def wb(self):
            if self._wb is None:
                try:
                    # 如果文件仍然不存在（理论上不会发生），再次尝试创建
                    if not os.path.exists(self.file):
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        if self.l_sheet:
                            ws.title = self.l_sheet[0]
                            for sheet_name in self.l_sheet[1:]:
                                wb.create_sheet(sheet_name)
                        else:
                            ws.title = "Sheet1"
                        wb.save(self.file)
                    self._wb = openpyxl.load_workbook(self.file)
                except PermissionError:
                    raise IOError(f"权限不足，无法访问或创建文件: {self.file}")
                except FileNotFoundError:
                    raise IOError(f"文件路径不存在: {self.file}")
                except Exception as e:
                    raise IOError(f"初始化文件失败: {e}") from e
            return self._wb

        # 加载工作簿
        self.wb = openpyxl.load_workbook(self.file)

        # self.wb = openpyxl.load_workbook(self.file, keep_vba=True)
        # self.wb.protect_worksheet(1)
        # self.wb.active = 1
        # self.wb.save(self.file)

        # 获取文档的字符集编码
        # print(self.wb.encoding)  # utf-8

        # 获取文档的元数据，如标题，创建者，创建日期等
        # print(self.wb.properties)
        # <openpyxl.packaging.core.DocumentProperties object>
        # Parameters:
        # creator='openpyxl', title=None, description=None, subject=None, identifier=None, language=None, created=datetime.datetime(2019, 1, 27, 18, 43, 19), modified=datetime.datetime(2024, 4, 11, 3, 55, 45), lastModifiedBy='jh', category=None, contentStatus=None, version=None, revision=None, keywords=None, lastPrinted=None
        # print(self.wb.properties.lastModifiedBy)  # jh

        # 通过索引值设置当前活跃的sheet
        # self.wb.active= 2


    # todo [工作表]

    def open(self):
        # 1.1 打开
        try:
            # 检查文件是否存在
            if not os.path.exists(self.file):
                raise FileNotFoundError(f"文件不存在: {self.file}")

            # 缓存平台信息
            system = platform.system()

            # 根据不同操作系统执行对应命令
            if system == "Darwin":  # macOS
                subprocess.run(["open", self.file], check=True)
            elif system == "Windows":  # Windows
                subprocess.run(["start", self.file], shell=True, check=True)
            elif system == "Linux":  # Linux
                subprocess.run(["xdg-open", self.file], check=True)
            else:
                raise OSError(f"不支持的操作系统: {system}")

            print(f"[INFO] 已成功打开文件: {self.file}")

        except FileNotFoundError as e:
            print(f"[ERROR] 文件未找到: {e}")
        except subprocess.CalledProcessError as e:
            print(f"[ERROR] 打开文件失败: {e}")
        except Exception as e:
            print(f"[ERROR] 发生未知错误: {e}")

    def save(self):
        """
        1.2 保存（优化版）
        :raises IOError: 如果保存过程中发生错误
        """
        try:
            # 校验文件路径是否存在
            if not os.path.exists(os.path.dirname(self.file)):
                raise FileNotFoundError(f"文件路径不存在: {os.path.dirname(self.file)}")

            # 保存文件
            self.wb.save(self.file)

        except PermissionError as e:
            raise IOError(f"权限不足，无法保存文件: {self.file}") from e
        except FileNotFoundError as e:
            raise IOError(f"文件路径不存在: {self.file}") from e
        except Exception as e:
            raise IOError(f"保存文件失败: {self.file}") from e

    def sh(self, varSheet):
        """
        1.3 工作表（优化版）
        :param varSheet: 工作表索引（int）或名称（str）
        :return: 工作表对象
        :raises ValueError: 如果参数不合法或工作表不存在
        """
        # 参数校验
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            if isinstance(varSheet, int):
                # 检查索引是否越界
                if varSheet < 0 or varSheet >= len(self.wb.sheetnames):
                    raise ValueError(f"工作表索引 {varSheet} 超出范围，可用索引为 0 到 {len(self.wb.sheetnames) - 1}")
                sh = self.wb[self.wb.sheetnames[varSheet]]
            elif isinstance(varSheet, str):
                # 检查工作表名称是否存在
                if varSheet not in self.wb.sheetnames:
                    raise ValueError(f"工作表 '{varSheet}' 不存在于当前文件中。可用的工作表包括：{self.wb.sheetnames}")
                sh = self.wb[varSheet]
            return sh
        except Exception as e:
            raise ValueError(f"获取工作表失败: {e}") from e

    def getL_sheet(self):

        ''' 1.4 获取工作表
        如 ['mySheet1', 'mySheet2', 'mySheet3']
        '''

        return self.wb.sheetnames

    # def switchSheet(self, varSheet=0):
    #
    #     '''
    #     1.5 切换工作表
    #     # switchSheet("Sheet2")
    #     '''
    #
    #     sh = self.sh(varSheet)
    #     self.wb.active = 2
    #     for sheet in self.wb:
    #         # print(sheet,sh)
    #         if sheet.title == sh.title:
    #             sheet.sheet_view.tabSelected = True
    #         else:
    #             sheet.sheet_view.tabSelected = False
    #     self.save()
    def switchSheet(self, varSheet=0):
        """
        1.5 切换工作表（优化版）

        :param varSheet: 工作表索引（int）或名称（str），默认为0（第一个工作表）
        :raises ValueError: 如果参数不合法或工作表不存在
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取目标工作表对象
            sh = self.sh(varSheet)

            # 检查工作表是否存在于工作簿中
            if sh.title not in self.wb.sheetnames:
                raise ValueError(f"工作表 '{varSheet}' 不存在于当前文件中。可用的工作表包括：{self.wb.sheetnames}")

            # 设置活动工作表
            self.wb.active = self.wb.sheetnames.index(sh.title)

            # 更新 tabSelected 属性（仅激活目标工作表）
            for sheet in self.wb:
                sheet.sheet_view.tabSelected = (sheet.title == sh.title)

            # 保存更改
            self.save()

            # 日志输出
            print(f"[INFO] 成功切换到工作表: {sh.title}")

        except ValueError as ve:
            raise ValueError(f"参数错误: {ve}") from ve
        except IOError as ioe:
            raise IOError(f"文件操作失败: {ioe}") from ioe
        except Exception as e:
            raise IOError(f"切换工作表失败: {e}") from e

    def addSheet(self, varSheetName, varIndex=0, overwrite=False):
        """
        1.6 添加工作表（优化版）

        :param varSheetName: 工作表名称，必须是字符串
        :param varIndex: 索引位置，默认为0
        :param overwrite: 是否覆盖已存在的工作表，默认False（不覆盖）
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSheetName, str):
            raise ValueError("varSheetName 必须是字符串")
        if not isinstance(varIndex, int):
            raise ValueError("varIndex 必须是整数")
        if not isinstance(overwrite, bool):
            raise ValueError("overwrite 必须是布尔值")

        try:
            # 检查工作表是否已存在
            sheet_exists = varSheetName in self.wb.sheetnames

            # 根据 overwrite 参数决定是否删除已有工作表
            if overwrite and sheet_exists:
                del self.wb[varSheetName]
                print(f"[INFO] 已删除已存在的工作表: {varSheetName}")

            # 如果工作表不存在或允许覆盖，则创建新工作表
            if not sheet_exists or overwrite:
                self.wb.create_sheet(title=varSheetName, index=varIndex)
                print(f"[INFO] 已成功添加工作表: {varSheetName}，位置索引: {varIndex}")
            else:
                print(f"[WARNING] 工作表 '{varSheetName}' 已存在且未启用覆盖模式，跳过创建。")

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"添加工作表失败: {e}") from e

    def delSheet(self, varSheetName):
        """
        1.7 删除工作表（优化版）

        :param varSheetName: 要删除的工作表名称，必须是字符串
        :raises ValueError: 如果参数不合法或工作表不存在
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSheetName, str):
            raise ValueError("varSheetName 必须是字符串")

        try:
            # 检查工作表是否存在
            if varSheetName not in self.wb.sheetnames:
                raise ValueError(f"工作表 '{varSheetName}' 不存在于当前文件中。可用的工作表包括：{self.wb.sheetnames}")

            # 检查是否只剩一个工作表
            if len(self.wb.sheetnames) <= 1:
                raise ValueError("[WARNING] Excel 文件必须保留至少一个工作表，无法删除最后一个工作表。")

            # 删除工作表
            del self.wb[varSheetName]
            self.save()

            # 日志输出
            print(f"[INFO] 已成功删除工作表: {varSheetName}")

        except ValueError as ve:
            raise ValueError(f"参数错误: {ve}") from ve
        except Exception as e:
            raise IOError(f"删除工作表失败: {e}") from e

    def renameSheet(self, varOldSheet, varNewSheet):
        """
        1.8 重命名工作表（优化版）

        :param varOldSheet: 原工作表名称或索引，必须是字符串或整数
        :param varNewSheet: 新工作表名称，必须是字符串
        :raises ValueError: 如果参数不合法或工作表不存在
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varOldSheet, (int, str)):
            raise ValueError("varOldSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(varNewSheet, str):
            raise ValueError("varNewSheet 必须是字符串")

        try:
            # 获取目标工作表对象
            ws = self.sh(varOldSheet)

            # 检查新名称是否与现有工作表冲突
            if varNewSheet in self.wb.sheetnames:
                raise ValueError(f"工作表 '{varNewSheet}' 已存在，无法重命名为相同名称。")

            # 执行重命名操作
            ws.title = varNewSheet

            # 保存更改
            self.save()

            # 日志输出
            print(f"[INFO] 已成功将工作表 '{varOldSheet}' 重命名为 '{varNewSheet}'")

        except KeyError:
            raise ValueError(f"工作表 '{varOldSheet}' 不存在于当前文件中。可用的工作表包括：{self.wb.sheetnames}")
        except ValueError as ve:
            raise ValueError(f"参数错误: {ve}") from ve
        except Exception as e:
            raise IOError(f"重命名工作表失败: {e}") from e



    # todo [操作数据]

    def insertNullRow(self, varRow, varStep=1, varSheet=0):
        """
        2.0 在第N行前插入多行空白（优化版）

        insertNullRow(3)  在第3行前插入1行空白
        insertNullRow(3，5)  在第3行前插入5行空白

        :param varRow: 插入位置的行号，必须是大于0的整数
        :param varStep: 插入的行数，默认为1，必须是大于0的整数
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varRow, int) or varRow <= 0:
            raise ValueError("seq 必须是大于0的整数")
        if not isinstance(varStep, int) or varStep <= 0:
            raise ValueError("moreRow 必须是大于0的整数")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 执行插入操作
            sh.insert_rows(idx=varRow, amount=varStep)

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"插入空白行失败: {e}") from e

    def insertNullCol(self, varCol, varStep=1, varSheet=0):
        """
        2.1 在第N列前插入多列空白（优化版）

        insertNullCol(3)  在第3列前插入1列空白
        insertNullCol(3，5)  在第3列前插入5列空白

        :param varCol: 插入位置的列号或列字母，必须是大于0的整数或有效列字母
        :param varStep: 插入的列数，默认为1，必须是大于0的整数
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """

        # 参数校验
        if not isinstance(varCol, (int, str)):
            raise ValueError("varCol 必须是整数或字符串")
        if isinstance(varCol, int) and varCol <= 0:
            raise ValueError("varCol 必须是大于0的整数")
        if isinstance(varCol, str) and not varCol.isalpha():
            raise ValueError("varCol 必须是有效的列字母")
        if not isinstance(varStep, int) or varStep <= 0:
            raise ValueError("moreCol 必须是大于0的整数")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 转换列标识为索引（如果是字符串）
            if isinstance(varCol, str):
                varCol = column_index_from_string(varCol)

            # 执行插入操作
            sh.insert_cols(idx=varCol, amount=varStep)

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"插入空白列失败: {e}") from e

    def setCell(self, varRow, varCol, varContent, varSheet=0):
        """
        2.2 设置单元格（优化版）

        :param varRow: 行号，必须是大于0的整数
        :param varCol: 列号（整数）或列字母（字符串），必须合法
        :param varContent: 单元格内容，可为任意类型
        :param varSheet: 工作表索引（int）或名称（str），默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varRow, int) or varRow <= 0:
            raise ValueError("varRow 必须是大于0的整数")
        if not isinstance(varCol, (int, str)):
            raise ValueError("varCol 必须是整数或字符串")
        if isinstance(varCol, str) and not varCol.isalpha():
            raise ValueError("varCol 必须是有效的列字母")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 转换列标识为索引（如果是字符串）
            if isinstance(varCol, str):
                varCol = column_index_from_string(varCol)

            # 边界检查
            max_row = sh.max_row
            max_col = sh.max_column
            if varRow > max_row:
                raise ValueError(f"行号 {varRow} 超出范围，最大行数为 {max_row}")
            if varCol > max_col:
                raise ValueError(f"列号 {varCol} 超出范围，最大列数为 {max_col}")

            # 设置单元格值
            sh.cell(row=varRow, column=varCol, value=varContent)

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置单元格值失败: {e}") from e

    def insertRow(self, d_rowNumber_l_value, varSheet=0):
        """
        2.3 插入行数据（优化版）

        :param d_var: 行数据字典，键为行号（整数），值为行内容（列表）
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """

        # 参数校验
        if not isinstance(d_rowNumber_l_value, dict):
            raise ValueError("d_var 必须是一个字典")
        if not all(isinstance(k, int) and isinstance(v, list) for k, v in d_rowNumber_l_value.items()):
            raise ValueError("d_var 的键必须是整数，值必须是列表")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 插入空白行
            for k in d_rowNumber_l_value.keys():
                self.insertNullRow(k, varSheet=varSheet)

            # 设置行数据
            self.setRow(d_rowNumber_l_value, varSheet=varSheet)

            # 保存文件
            self.save()

        except Exception as e:
            raise IOError(f"插入行数据失败: {e}") from e

    def setRow(self, d_rowNumber_l_value, varSheet=0):
        """
        2.4 更新行数据（优化版）

        :param d_var: 行数据字典，键为行号（整数），值为行内容（列表）
        :param varSheet: 工作表索引（int）或名称（str），默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(d_rowNumber_l_value, dict):
            raise ValueError("d_var 必须是一个字典")
        if not all(isinstance(k, int) and isinstance(v, list) for k, v in d_rowNumber_l_value.items()):
            raise ValueError("d_var 的键必须是整数，值必须是列表")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 更新行数据
            for k, v in d_rowNumber_l_value.items():
                for i in range(len(v)):
                    if v[i] is not None:
                        sh.cell(row=k, column=i + 1, value=str(v[i]))

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"更新行数据失败: {e}") from e

    def appendRow(self, l_l_value, varSheet=0):
        """
        2.5 追加行数据（优化版）

        :param l_l_rows: 行数据列表，每个元素为一个列表，表示一行数据
        :param varSheet: 工作表索引（int）或名称（str），默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(l_l_value, list):
            raise ValueError("l_l_rows 必须是一个列表")
        if not all(isinstance(row, list) for row in l_l_value):
            raise ValueError("l_l_rows 中的每个元素必须是一个列表")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 追加行数据
            for row_data in l_l_value:
                sh.append(row_data)

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"追加行数据失败: {e}") from e

    def insertCol(self, d_col_l_value, varSheet=0):
        """
        2.6 插入列数据（优化版）

        insertCol({"a": ["姓名", "张三", "李四"], "c": ["年龄", "55", "34"]})
        insertCol({1: ["姓名", "张三", "李四"], 3: ["年龄", "55", "34"]})

        :param d_var: 列数据字典，键为列号（整数）或列字母（字符串），值为列内容（列表）
        :param varSheet: 工作表索引（int）或名称（str），默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(d_col_l_value, dict):
            raise ValueError("d_var 必须是一个字典")
        if not all(isinstance(k, (int, str)) and isinstance(v, list) for k, v in d_col_l_value.items()):
            raise ValueError("d_var 的键必须是整数或字符串，值必须是列表")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 转换列标识为索引并排序（去重）
            col_indices = []
            for col in d_col_l_value.keys():
                if isinstance(col, str):
                    col = column_index_from_string(col)
                col_indices.append(col)
            col_indices = sorted(set(col_indices))

            # 插入空白列
            for col in col_indices:
                sh.insert_cols(idx=col, amount=1)

            # 设置列数据
            for k, v in d_col_l_value.items():
                col_idx = column_index_from_string(k) if isinstance(k, str) else k
                for i in range(len(v)):
                    if v[i] is not None:
                        sh.cell(row=i + 1, column=col_idx, value=v[i])

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"插入列数据失败: {e}") from e

    def setCol(self, d_col_l_value, varSheet=0):
        """
        更新列数据（优化版）

        :param d_col_l_value: 列数据字典，键为列号（整数）或列字母（字符串），值为列内容（列表）
        :param varSheet: 工作表索引（int）或名称（str），默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(d_col_l_value, dict):
            raise ValueError("d_col_l_value 必须是一个字典")
        if not all(isinstance(k, (int, str)) and isinstance(v, list) for k, v in d_col_l_value.items()):
            raise ValueError("d_col_l_value 的键必须是整数或字符串，值必须是列表")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 遍历列数据字典
            for col_key, col_values in d_col_l_value.items():
                # 处理列标识：如果是字符串且为数字，则转换为列字母；否则直接使用
                if isinstance(col_key, str) and col_key.isdigit():
                    col_index = int(col_key)  # 直接使用数字索引
                elif isinstance(col_key, str):
                    col_index = column_index_from_string(col_key)  # 转换列字母为索引
                elif isinstance(col_key, int):
                    col_index = col_key  # 直接使用数字索引
                else:
                    raise ValueError(f"列标识 '{col_key}' 类型不合法，应为整数或字符串")

                # 写入列数据
                for row_idx, value in enumerate(col_values, start=1):
                    if value is not None:  # 跳过 None 值
                        sh.cell(row=row_idx, column=col_index, value=value)

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"设置列数据失败: {e}") from e

    def appendCol(self, l_l_value, varSheet=0):
        """
        2.8 追加列数据（优化版）

        :param l_l_cols: 列数据列表，每个元素为一个列表，表示一列数据
        :param varSheet: 工作表索引（int）或名称（str），默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(l_l_value, list):
            raise ValueError("l_l_cols 必须是一个列表")
        if not all(isinstance(col, list) for col in l_l_value):
            raise ValueError("l_l_cols 中的每个元素必须是一个列表")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 获取当前总列数
            total_cols = sh.max_column

            # 构建列字母与数据的映射
            col_mapping = {
                get_column_letter(total_cols + i + 1): col_data
                for i, col_data in enumerate(l_l_value)
            }

            # 设置列数据
            self.setCol(col_mapping, varSheet)

        except Exception as e:
            raise IOError(f"追加列数据失败: {e}") from e

    def clsRow(self, varNums, varSheet=0):
        """
        2.9 清空行（优化版）

        :param varNums: 行号，必须是大于0的整数
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varNums, int) or varNums <= 0:
            raise ValueError("varNums 必须是大于0的整数")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 检查行号是否超出范围
            if varNums > sh.max_row:
                raise ValueError(f"行号 {varNums} 超出范围，最大行数为 {sh.max_row}")

            # # 备份原始文件（可选）
            # backup_path = self.file.replace(".xlsx", "_backup.xlsx")
            # self.wb.save(backup_path)

            # 清空指定行的数据
            for col in range(1, sh.max_column + 1):
                cell = sh.cell(row=varNums, column=col)
                cell.value = None  # 清除值
                cell.font = None  # 清除字体样式
                cell.fill = PatternFill()  # 正确清除背景色
                cell.border = None  # 清除边框
                cell.alignment = None  # 清除对齐方式
                cell.number_format = ''  # 设置为空字符串，避免 None 错误
                cell.protection = None  # 清除保护设置

            # 保存更改
            self.save()

        except ValueError as ve:
            raise ValueError(f"参数错误: {ve}") from ve
        except IOError as ioe:
            raise IOError(f"文件操作失败: {ioe}") from ioe
        except Exception as e:
            raise IOError(f"清空行失败: {e}") from e

    def clsCol(self, varCol, varSheet=0, clear_header=True):
        """
        2.10 清空列（优化版）

        :param varCol: 列号，可以是整数（如2）或字符串（如'B'）
        :param varSheet: 工作表索引或名称，默认为0
        :param clear_header: 是否清除列标签（第一行），默认为True
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varCol, (int, str)):
            raise ValueError("varCol 必须是整数或字符串")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(clear_header, bool):
            raise ValueError("clear_header 必须是布尔值")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 转换列标识为索引
            if isinstance(varCol, str):
                varCol = column_index_from_string(varCol)

            # 确定起始行
            start_row = 1 if clear_header else 2

            # 清空列数据
            for i in range(start_row, sh.max_row + 1):
                sh.cell(row=i, column=varCol).value = None

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"清空列失败: {e}") from e

    def delRow(self, varFrom, varSeries=1, varSheet=0):
        """
        2.11 删除连续行（优化版）

        :param varFrom: 起始行号，必须是大于0的整数
        :param varSeries: 删除的行数，默认为1
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varFrom, int) or varFrom <= 0:
            raise ValueError("varFrom 必须是大于0的整数")
        if not isinstance(varSeries, int) or varSeries <= 0:
            raise ValueError("varSeries 必须是大于0的整数")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 执行删除操作
            sh.delete_rows(idx=varFrom, amount=varSeries)

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"删除行失败: {e}") from e

    def delCol(self, varFrom, varSeries=1, varSheet=0):
        """
        2.12 删除连续列（优化版）

        :param varFrom: 起始列，可以是整数（如2）或字符串（如'U'）
        :param varSeries: 删除的列数，默认为1
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varFrom, (int, str)):
            raise ValueError("varFrom 必须是整数或字符串")
        if not isinstance(varSeries, int) or varSeries <= 0:
            raise ValueError("varSeries 必须是大于0的整数")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 转换列标识为索引
            if isinstance(varFrom, str):
                varFrom = column_index_from_string(varFrom)

            # 执行删除操作
            sh.delete_cols(idx=varFrom, amount=varSeries)

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"删除列失败: {e}") from e

    def moveBlock(self, varBlock, varRow, varCol, varSheet=0):
        """
        2.13 移动区域数据（优化版）

        # Openpyxl_PO.moveBlock('C1:D2', 3, -2)  # 把'C1:D2'区域移动到 下三行左二列
        # Openpyxl_PO.moveBlock('A1:C14', 0, 3)  # 把'A1:C14'区域向右移动3列

        :param varFrom: 起始区域，如 'C1:D2'
        :param varRows: 移动的行数，正数向下移动，负数向上移动
        :param varCols: 移动的列数，正数向右移动，负数向左移动
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varBlock, str) or not varBlock:
            raise ValueError("varFrom 必须是一个非空字符串，表示起始区域")
        if not isinstance(varRow, int) or not isinstance(varCol, int):
            raise ValueError("varRows 和 varCols 必须是整数")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 执行移动操作
            sh.move_range(varBlock, rows=varRow, cols=varCol)

            # 保存更改
            self.save()
        except Exception as e:
            raise IOError(f"移动区域数据失败: {e}") from e

    def sortColHeader(self, varSheetName):
        """
        2.14 将表格第一行（标题）排序（从小打大）（优化版）

        :param varSheet1: 工作表名称
        :raises ValueError: 如果输入参数不合法
        :raises IOError: 如果文件读取或写入失败
        return：生成 {表名}_标题升序, 如表名字：sheet1， 结果：sheet1_标题升序
        """

        if varSheetName not in self.wb.sheetnames:
            raise ValueError(f"工作表 '{varSheetName}' 不存在于当前文件中。可用的工作表包括：{self.wb.sheetnames}")

        # 参数校验
        if not isinstance(varSheetName, str):
            raise ValueError("varSheet1 必须是一个字符串")

        try:
            # 获取第一行标题
            l_title = self.getL_row(1, varSheetName)
            if not l_title:
                raise ValueError("工作表中没有找到标题行")

            # 排序标题
            l_sortAsc = sorted(l_title)

            # 创建新工作表并写入排序后的标题
            varNewSheet = varSheetName + "_" + "标题升序"
            self.addSheet(varNewSheet)
            self.setRow({1: l_sortAsc}, varNewSheet)

            # 获取原始列数据
            col = self.getLL_col(varSheetName)

            # 根据排序后的标题重新排列列数据
            for i in range(len(col)):
                for j in range(len(l_sortAsc)):
                    if col[i][0] == l_sortAsc[j]:
                        self.setCol({str(j + 1): col[i]}, varNewSheet)

            self.save()

        except FileNotFoundError as e:
            raise IOError(f"文件未找到: {varSheetName}") from e
        except Exception as e:
            raise IOError(f"处理工作表失败: {varSheetName}") from e


    # todo [样式]

    def setColWidth(self, col, colQty, varSheet=0):
        """
        3.0 设置列宽（合并版）

        :param col: 列标识，可以是整数（如1）或字符串（如'f'）
        :param colQty: 列宽，必须是数字
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(col, (int, str)):
            raise ValueError("col 必须是整数或字符串")
        if not isinstance(colQty, (int, float)):
            raise ValueError("colQty 必须是数字")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 处理列标识：如果是整数，则转换为列字母；如果是字符串，则直接使用
            if isinstance(col, int):
                col_letter = get_column_letter(col)
            elif isinstance(col, str):
                col_letter = col
            else:
                raise ValueError("col 类型不合法，应为整数或字符串")

            # 设置列宽
            sh.column_dimensions[col_letter].width = colQty

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"设置列宽失败: {e}") from e

    def setRowColSize(self, row_qty_pairs=None, col_qty_pairs=None, varSheet=0):
        """
        3.1 设置行高与列宽（优化版）

        :param row_qty_pairs: 行高设置的元组列表，如 [(3, 30), (5, 40)] 表示第3行高30，第5行高40
        :param col_qty_pairs: 列宽设置的元组列表，如 [('f', 50), ('h', 60)] 表示第f列宽50，第h列宽60
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if row_qty_pairs is not None and not all(isinstance(pair, tuple) and len(pair) == 2 for pair in row_qty_pairs):
            raise ValueError("row_qty_pairs 必须是包含两个元素的元组列表")
        if col_qty_pairs is not None and not all(isinstance(pair, tuple) and len(pair) == 2 for pair in col_qty_pairs):
            raise ValueError("col_qty_pairs 必须是包含两个元素的元组列表")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 设置行高
            if row_qty_pairs:
                for row, qty in row_qty_pairs:
                    if not isinstance(row, int) or row <= 0:
                        raise ValueError(f"行号 {row} 必须是大于0的整数")
                    if not isinstance(qty, (int, float)) or qty < 0:
                        raise ValueError(f"行高 {qty} 必须是非负数")
                    sh.row_dimensions[row].height = qty

            # 设置列宽
            if col_qty_pairs:
                for col, qty in col_qty_pairs:
                    if not isinstance(col, (int, str)):
                        raise ValueError(f"列标识 {col} 必须是整数或字符串")
                    if not isinstance(qty, (int, float)) or qty < 0:
                        raise ValueError(f"列宽 {qty} 必须是非负数")
                    col_letter = get_column_letter(col) if isinstance(col, int) else col
                    sh.column_dimensions[col_letter].width = qty

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"设置单元格行高与列宽失败: {e}") from e

    def setAllSize(self, rowQty, colQty, varSheet=0):
        """
        3.2 设置所有单元格的行高与列宽（优化版）

        :param rowQty: 行高，必须是非负数
        :param colQty: 列宽，必须是非负数
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(rowQty, (int, float)) or rowQty < 0:
            raise ValueError("rowQty 必须是非负数")
        if not isinstance(colQty, (int, float)) or colQty < 0:
            raise ValueError("colQty 必须是非负数")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 获取最大行数和列数
            max_row = sh.max_row
            max_col = sh.max_column

            # 容错处理：如果工作表为空，则直接返回
            if max_row is None or max_col is None:
                print("[WARNING] 工作表为空，无需设置行高和列宽。")
                return

            # 批量设置行高
            for i in range(1, max_row + 1):
                sh.row_dimensions[i].height = rowQty

            # 批量设置列宽
            for i in range(1, max_col + 1):
                sh.column_dimensions[get_column_letter(i)].width = colQty

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"设置所有单元格行高与列宽失败: {e}") from e

    def setWrapText(self, varSheet=0):
        """
        3.3 设置自动换行（优化版）
        自动换行后，内容将完全显示出来，并动态调整行高以适配内容。

        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 获取最大行数和列数
            max_row = sh.max_row
            max_col = sh.max_column

            # 容错处理：如果工作表为空，则直接返回
            if max_row is None or max_col is None:
                print("[WARNING] 工作表为空，无需设置自动换行。")
                return

            # 遍历所有单元格并设置自动换行及动态行高
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sh.cell(row=row, column=col)
                    # 设置自动换行
                    cell.alignment = Alignment(wrapText=True)

                    # 动态调整行高以适配内容
                    if cell.value:
                        # 获取列宽（单位：字符宽度）
                        col_width = sh.column_dimensions[get_column_letter(col)].width or 10
                        # 获取字体大小（默认为 11）
                        font_size = cell.font.size or 11
                        # 估算文本行数
                        estimated_lines = self._estimate_text_lines(str(cell.value), col_width, font_size)
                        # 设置行高（每行约 1.2 倍字体高度）
                        sh.row_dimensions[row].height = max(sh.row_dimensions[row].height or 15,
                                                            estimated_lines * font_size * 1.2)

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"设置自动换行失败: {e}") from e
    def _estimate_text_lines(self, text, col_width, font_size):
        """
        估算文本所需行数（私有辅助函数）

        :param text: 单元格内容
        :param col_width: 列宽（字符数）
        :param font_size: 字体大小
        :return: 估算的行数
        """
        # 经验公式：每字符宽度约为字体大小的 0.6 倍
        char_width = font_size * 0.6
        # 每行最多容纳的字符数
        chars_per_line = col_width / char_width
        # 估算行数
        lines = len(text) / chars_per_line
        return int(lines) + 1  # 向上取整

    def setFreezePanes(self, coordinate, varSheet=0):
        """
        3.4 冻结窗口（优化版）
        支持冻结行、列或区域。

        :param coordinate: 冻结的起始单元格坐标，如 'A2' 表示冻结第1行，'B1' 表示冻结第A列，
                           'B2' 表示冻结第1行和第A列。
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法或坐标无效
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(coordinate, str) or not coordinate:
            raise ValueError("coordinate 必须是一个非空字符串，表示冻结的起始单元格坐标")

        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 校验坐标有效性
            if not self._is_valid_coordinate(coordinate):
                raise ValueError(f"无效的坐标: {coordinate}")

            # 设置冻结窗格
            sh.freeze_panes = coordinate

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"冻结窗口失败: {e}") from e
    def _is_valid_coordinate(self, coordinate):
        """
        校验坐标是否有效（私有辅助函数）

        :param coordinate: 单元格坐标，如 'A1', 'B2' 等
        :return: 坐标是否有效
        """
        try:
            from openpyxl.utils import coordinate_to_tuple
            row, col = coordinate_to_tuple(coordinate)
            return row >= 1 and col >= 1
        except Exception:
            return False

    def setAlignment(self, row, col, horizontal="center", vertical="center", text_rotation=0, wrap_text=False, varSheet=0):
        """
        3.5 设置对齐样式（优化版）
        支持设置水平对齐、垂直对齐、文本旋转和自动换行。

        :param row: 行号，必须是大于0的整数
        :param col: 列号（整数）或列字母（字符串），必须合法
        :param horizontal: 水平对齐方式，默认为 "center"
        :param vertical: 垂直对齐方式，默认为 "center"
        :param text_rotation: 文本旋转角度，默认为 0
        :param wrap_text: 是否自动换行，默认为 False
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(row, int) or row <= 0:
            raise ValueError("row 必须是大于0的整数")
        if not isinstance(col, (int, str)):
            raise ValueError("col 必须是整数或字符串")
        if isinstance(col, str) and not col.isalpha():
            raise ValueError("col 必须是有效的列字母")
        if not isinstance(horizontal, str) or horizontal not in (
                "general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed"
        ):
            raise ValueError("horizontal 必须是合法的对齐方式")
        if not isinstance(vertical, str) or vertical not in (
                "top", "center", "bottom", "justify", "distributed"
        ):
            raise ValueError("vertical 必须是合法的对齐方式")
        if not isinstance(text_rotation, int) or not (0 <= text_rotation <= 180):
            raise ValueError("text_rotation 必须是 0 到 180 之间的整数")
        if not isinstance(wrap_text, bool):
            raise ValueError("wrap_text 必须是布尔值")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象（缓存以减少重复调用）
            sh = self.sh(varSheet)

            # 转换列标识为索引（如果是字符串）
            if isinstance(col, str):
                col = column_index_from_string(col)

            # 边界检查
            max_row = sh.max_row
            max_col = sh.max_column
            if row > max_row:
                raise ValueError(f"行号 {row} 超出范围，最大行数为 {max_row}")
            if col > max_col:
                raise ValueError(f"列号 {col} 超出范围，最大列数为 {max_col}")

            # 设置单元格对齐样式
            sh.cell(row=row, column=col).alignment = Alignment(
                horizontal=horizontal,
                vertical=vertical,
                text_rotation=text_rotation,
                wrap_text=wrap_text,
            )

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置单元格对齐样式失败: {e}") from e

    def setRowColAlignment(self, row, l_col, horizontal="center", vertical="center", text_rotation=0, wrap_text=False, varSheet=0):
        """
        3.6 设置单行多列对齐样式（优化版）
        支持设置指定行中多列的对齐方式。

        :param row: 行号，必须是大于0的整数
        :param l_col: 列标识列表（如 [4, 6]）或字符串 "all"
        :param horizontal: 水平对齐方式，默认为 "center"
        :param vertical: 垂直对齐方式，默认为 "center"
        :param text_rotation: 文本旋转角度，默认为 0
        :param wrap_text: 是否自动换行，默认为 False
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(row, int) or row <= 0:
            raise ValueError("row 必须是大于0的整数")
        if not isinstance(l_col, (list, str)) or (
                isinstance(l_col, list) and not all(isinstance(c, (int, str)) for c in l_col)):
            raise ValueError("l_col 必须是列表（元素为整数或字符串）或字符串 'all'")
        if not isinstance(horizontal, str) or horizontal not in (
                "general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed"
        ):
            raise ValueError("horizontal 必须是合法的对齐方式")
        if not isinstance(vertical, str) or vertical not in (
                "top", "center", "bottom", "justify", "distributed"
        ):
            raise ValueError("vertical 必须是合法的对齐方式")
        if not isinstance(text_rotation, int) or not (0 <= text_rotation <= 180):
            raise ValueError("text_rotation 必须是 0 到 180 之间的整数")
        if not isinstance(wrap_text, bool):
            raise ValueError("wrap_text 必须是布尔值")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象（缓存以减少重复调用）
            sh = self.sh(varSheet)

            # 边界检查
            max_row = sh.max_row
            max_col = sh.max_column
            if row > max_row:
                raise ValueError(f"行号 {row} 超出范围，最大行数为 {max_row}")

            # 设置对齐样式
            alignment = Alignment(
                horizontal=horizontal,
                vertical=vertical,
                text_rotation=text_rotation,
                wrap_text=wrap_text,
            )

            if l_col == "all":
                # 全部列应用对齐样式
                for col in range(1, max_col + 1):
                    sh.cell(row=row, column=col).alignment = alignment
            else:
                # 指定列应用对齐样式
                col_indices = set()
                for col in l_col:
                    if isinstance(col, str):
                        col_indices.add(column_index_from_string(col))
                    elif isinstance(col, int):
                        col_indices.add(col)
                    else:
                        raise ValueError(f"列标识 '{col}' 类型不合法，应为整数或字符串")

                for col in sorted(col_indices):
                    if col > max_col:
                        raise ValueError(f"列号 {col} 超出范围，最大列数为 {max_col}")
                    sh.cell(row=row, column=col).alignment = alignment

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置单行多列对齐样式失败: {e}") from e

    def setAllAlignment(self, horizontal="center", vertical="center", text_rotation=0, wrap_text=False, varSheet=0):
        """
        3.7 设置所有单元格对齐样式（优化版）
        支持设置所有单元格的水平对齐、垂直对齐、文本旋转和自动换行。

        :param horizontal: 水平对齐方式，默认为 "center"
        :param vertical: 垂直对齐方式，默认为 "center"
        :param text_rotation: 文本旋转角度，默认为 0
        :param wrap_text: 是否自动换行，默认为 False
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(horizontal, str) or horizontal not in (
                "general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed"
        ):
            raise ValueError("horizontal 必须是合法的对齐方式")
        if not isinstance(vertical, str) or vertical not in (
                "top", "center", "bottom", "justify", "distributed"
        ):
            raise ValueError("vertical 必须是合法的对齐方式")
        if not isinstance(text_rotation, int) or not (0 <= text_rotation <= 180):
            raise ValueError("text_rotation 必须是 0 到 180 之间的整数")
        if not isinstance(wrap_text, bool):
            raise ValueError("wrap_text 必须是布尔值")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象（缓存以减少重复调用）
            sh = self.sh(varSheet)

            # 获取最大行数和列数（缓存以减少重复调用）
            max_row = sh.max_row
            max_col = sh.max_column

            # 容错处理：如果工作表为空，则直接返回
            if max_row is None or max_col is None:
                return

            # 创建对齐样式对象（复用以减少重复创建）
            alignment = Alignment(
                horizontal=horizontal,
                vertical=vertical,
                text_rotation=text_rotation,
                wrap_text=wrap_text,
            )

            # 批量设置所有单元格的对齐样式
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    sh.cell(row=r, column=c).alignment = alignment

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置所有单元格对齐样式失败: {e}") from e

    def setFilterCol(self, varCell="all", varSheet=0):
        """
        3.8 设置筛选列（优化版）
        支持设置全部筛选、取消筛选或指定区域筛选。

        :param varCell: 筛选范围，可以是 "all"（全部筛选）、""（取消筛选）或具体单元格范围（如 "A2"）
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varCell, str):
            raise ValueError("varCell 必须是字符串类型")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象（缓存以减少重复调用）
            sh = self.sh(varSheet)

            # 处理筛选逻辑
            if varCell == "all":
                # 全部筛选：使用整个工作表的范围
                sh.auto_filter.ref = sh.dimensions
            elif varCell == "":
                # 取消筛选：清空筛选范围
                sh.auto_filter.ref = None
            else:
                # 指定区域筛选：校验单元格范围有效性
                if not self._is_valid_cell_range(varCell):
                    raise ValueError(f"无效的单元格范围: {varCell}")
                sh.auto_filter.ref = varCell

            # 保存更改
            self.save()

        except Exception as e:
            raise IOError(f"设置筛选列失败: {e}") from e
    def _is_valid_cell_range(self, cell_range):
        """
        校验单元格范围是否有效（私有辅助函数）

        :param cell_range: 单元格范围字符串，如 "A1:B2"
        :return: 单元格范围是否有效
        """
        try:
            from openpyxl.utils import range_boundaries
            range_boundaries(cell_range)  # 尝试解析范围，失败会抛出异常
            return True
        except Exception:
            return False

    def setCellFont(self, row, col, name="微软雅黑", size=10, bold=False, italic=False, color=None, varSheet=0):
        """
        3.9 设置单元格字体（优化版）
        支持设置字体、字号、粗体、斜体、下划线和颜色。

        :param row: 行号，必须是大于0的整数
        :param col: 列号（整数）或列字母（字符串），必须合法
        :param name: 字体名称，默认为 "微软雅黑"
        :param size: 字号，默认为 10
        :param bold: 是否加粗，默认为 False
        :param italic: 是否斜体，默认为 False
        :param color: 字体颜色（十六进制），默认为 None
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(row, int) or row <= 0:
            raise ValueError("row 必须是大于0的整数")
        if not isinstance(col, (int, str)):
            raise ValueError("col 必须是整数或字符串")
        if isinstance(col, str) and not col.isalpha():
            raise ValueError("col 必须是有效的列字母")
        if not isinstance(name, str):
            raise ValueError("name 必须是字符串")
        if not isinstance(size, (int, float)) or size <= 0:
            raise ValueError("size 必须是大于0的数字")
        if not isinstance(bold, bool):
            raise ValueError("bold 必须是布尔值")
        if not isinstance(italic, bool):
            raise ValueError("italic 必须是布尔值")
        if color is not None and not isinstance(color, str):
            raise ValueError("color 必须是字符串或 None")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象（缓存以减少重复调用）
            sh = self.sh(varSheet)

            # 转换列标识为索引（如果是字符串）
            if isinstance(col, str):
                col = column_index_from_string(col)

            # 边界检查
            max_row = sh.max_row
            max_col = sh.max_column
            if row > max_row:
                raise ValueError(f"行号 {row} 超出范围，最大行数为 {max_row}")
            if col > max_col:
                raise ValueError(f"列号 {col} 超出范围，最大列数为 {max_col}")

            # 设置单元格字体
            sh.cell(row=row, column=col).font = Font(
                name=name, size=size, bold=bold, italic=italic, color=color
            )

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置单元格字体失败: {e}") from e

    def setRowColFont(self, row, l_col, name="微软雅黑", size=16, bold=False, italic=False, color="000000", varSheet=0):

        # 3.10 设置单行多列字体（字体、字号、粗斜体、下划线、颜色）
        # setRowColFont(1, ["e", "h"])
        # setRowColFont(1, "all", color="000000")
        sh = self.sh(varSheet)
        cols = sh.max_column
        if l_col == "all":
            for i in range((cols)):
                sh.cell(row=row, column=i + 1).font = Font(
                    name=name, size=size, bold=bold, italic=italic, color=color
                )
        else:
            for i in range(
                column_index_from_string(l_col[0]),
                int(column_index_from_string(l_col[1])) + 1,
            ):
                sh.cell(row=row, column=i).font = Font(
                    name=name, size=size, bold=bold, italic=italic, color=color
                )
        self.save()

    def setAllFont(self, name="微软雅黑", size=16, bold=False, italic=False, color="000000", varSheet=0):
        """
        3.11 设置所有单元格字体（优化版）
        支持设置所有单元格的字体、字号、粗体、斜体、下划线和颜色。

        :param name: 字体名称，默认为 "微软雅黑"
        :param size: 字号，默认为 16
        :param bold: 是否加粗，默认为 False
        :param italic: 是否斜体，默认为 False
        :param color: 字体颜色（十六进制），默认为 "000000"
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(name, str):
            raise ValueError("name 必须是字符串")
        if not isinstance(size, (int, float)) or size <= 0:
            raise ValueError("size 必须是大于0的数字")
        if not isinstance(bold, bool):
            raise ValueError("bold 必须是布尔值")
        if not isinstance(italic, bool):
            raise ValueError("italic 必须是布尔值")
        if not isinstance(color, str):
            raise ValueError("color 必须是字符串")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象（缓存以减少重复调用）
            sh = self.sh(varSheet)

            # 获取最大行数和列数（缓存以减少重复调用）
            max_row = sh.max_row
            max_col = sh.max_column

            # 容错处理：如果工作表为空，则直接返回
            if max_row is None or max_col is None:
                print("[WARNING] 工作表为空，无需设置字体。")
                return

            # 创建字体对象（复用以减少重复创建）
            font = Font(name=name, size=size, bold=bold, italic=italic, color=color)

            # 批量设置所有单元格的字体
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    sh.cell(row=r, column=c).font = font

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置所有单元格字体失败: {e}") from e

    def setBorder(self, row, col, left=["thin", "ff0000"], right=["thick", "ff0000"], top=["thin", "ff0000"], bottom=["thick", "ff0000"], varSheet=0):
        """
        3.12 设置单元格边框（优化版）
        支持设置上下左右边框样式和颜色。

        :param row: 行号，必须是大于0的整数
        :param col: 列号（整数）或列字母（字符串），必须合法
        :param left: 左边框样式和颜色，如 ['thin', 'ff0000']
        :param right: 右边框样式和颜色，如 ['thick', 'ff0000']
        :param top: 上边框样式和颜色，如 ['thin', 'ff0000']
        :param bottom: 下边框样式和颜色，如 ['thick', 'ff0000']
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(row, int) or row <= 0:
            raise ValueError("row 必须是大于0的整数")
        if not isinstance(col, (int, str)):
            raise ValueError("col 必须是整数或字符串")
        if isinstance(col, str) and not col.isalpha():
            raise ValueError("col 必须是有效的列字母")
        if not isinstance(left, list) or len(left) != 2:
            raise ValueError("left 必须是一个包含两个元素的列表，如 ['thin', 'ff0000']")
        if not isinstance(right, list) or len(right) != 2:
            raise ValueError("right 必须是一个包含两个元素的列表，如 ['thick', 'ff0000']")
        if not isinstance(top, list) or len(top) != 2:
            raise ValueError("top 必须是一个包含两个元素的列表，如 ['thin', 'ff0000']")
        if not isinstance(bottom, list) or len(bottom) != 2:
            raise ValueError("bottom 必须是一个包含两个元素的列表，如 ['thick', 'ff0000']")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象（缓存以减少重复调用）
            sh = self.sh(varSheet)

            # 转换列标识为索引（如果是字符串）
            if isinstance(col, str):
                col = column_index_from_string(col)

            # 边界检查
            max_row = sh.max_row
            max_col = sh.max_column
            if row > max_row:
                raise ValueError(f"行号 {row} 超出范围，最大行数为 {max_row}")
            if col > max_col:
                raise ValueError(f"列号 {col} 超出范围，最大列数为 {max_col}")

            # 校验边框样式合法性
            valid_styles = [
                "double", "mediumDashDotDot", "slantDashDot", "dashDotDot", "dotted",
                "hair", "mediumDashed", "dashed", "dashDot", "thin", "mediumDashDot",
                "medium", "thick"
            ]
            for side in [left, right, top, bottom]:
                if side[0] not in valid_styles:
                    raise ValueError(f"无效的边框样式: {side[0]}")

            # 创建边框对象
            border = Border(
                left=Side(style=left[0], color=left[1]),
                right=Side(style=right[0], color=right[1]),
                top=Side(style=top[0], color=top[1]),
                bottom=Side(style=bottom[0], color=bottom[1]),
            )

            # 设置单元格边框
            sh.cell(row=row, column=col).border = border

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置单元格边框失败: {e}") from e


    def setGradientFill(self, row, col, stop=["FFFFFF", "99ccff", "000000"], varSheet=0):
        """
        3.13 设置单元格填充渐变色

        :param row: 行号，必须是大于0的整数
        :param col: 列号（整数）或列字母（字符串），必须合法
        :param stop: 渐变颜色列表，每个元素为6位十六进制字符串，如 ["FFFFFF", "99ccff", "000000"]
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法或超出范围
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(row, int) or row <= 0:
            raise ValueError("row 必须是大于0的整数")
        if not isinstance(col, (int, str)):
            raise ValueError("col 必须是整数或字符串")
        if isinstance(col, str) and not col.isalpha():
            raise ValueError("col 必须是有效的列字母")
        if not isinstance(stop, list) or len(stop) < 2:
            raise ValueError("stop 必须是一个包含至少两个元素的列表")
        if not all(isinstance(color, str) and len(color) == 6 and all(c in '0123456789ABCDEFabcdef' for c in color) for
                   color in stop):
            raise ValueError("stop 中的每个元素必须是6位十六进制字符串")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 转换列标识为索引（如果是字符串）
            if isinstance(col, str):
                col = column_index_from_string(col)

            # 边界检查
            max_row = sh.max_row
            max_col = sh.max_column
            if row > max_row:
                raise ValueError(f"行号 {row} 超出范围，最大行数为 {max_row}")
            if col > max_col:
                raise ValueError(f"列号 {col} 超出范围，最大列数为 {max_col}")

            # 设置单元格渐变色
            gradient_fill = GradientFill(stop=tuple(stop))
            sh.cell(row=row, column=col).fill = gradient_fill

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置单元格渐变色失败: {e}") from e

    def setBackgroundColor(self, row, col, varColor=None, clear_color=False, varSheet=0):
        """
        3.14 设置单元格背景色（优化版）

        :param row: 行号，可以是整数（如5）或 None（清除所有行）
        :param col: 列号（整数）或列字母（字符串），可以是 None（清除所有列）
        :param varColor: 背景色（十六进制字符串）或 None（清除背景色）
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法或超出范围
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if row is not None and (not isinstance(row, int) or row <= 0):
            raise ValueError("row 必须是大于0的整数或 None")
        if col is not None and not isinstance(col, (int, str)):
            raise ValueError("col 必须是整数、字符串或 None")
        if isinstance(col, str) and not col.isalpha():
            raise ValueError("col 必须是有效的列字母")
        if varColor is not None and (
                not isinstance(varColor, str) or
                len(varColor) != 6 or
                not all(c in '0123456789ABCDEFabcdef' for c in varColor)
        ):
            raise ValueError("varColor 必须是6位十六进制字符串或 None")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 获取最大行数和列数
            max_row = sh.max_row
            max_col = sh.max_column

            # 容错处理：如果工作表为空，则直接返回
            if max_row is None or max_col is None:
                print("[WARNING] 工作表为空，无需设置背景色。")
                return

            # 转换列标识为索引（如果是字符串）
            if isinstance(col, str):
                col = column_index_from_string(col)

            # 清除所有单元格背景色
            if row is None and col is None:
                style = PatternFill(fill_type=None)
                for i in range(1, max_row + 1):
                    for j in range(1, max_col + 1):
                        sh.cell(i, j).fill = style
            else:
                # 边界检查
                if row is not None and row > max_row:
                    raise ValueError(f"行号 {row} 超出范围，最大行数为 {max_row}")
                if col is not None and col > max_col:
                    raise ValueError(f"列号 {col} 超出范围，最大列数为 {max_col}")

                # 设置单元格背景色
                if varColor is not None:
                    style = PatternFill("solid", fgColor=varColor)

                if clear_color == True:
                    style = PatternFill(fill_type=None)  # 清除背景色

                if row is not None and col is not None:
                    sh.cell(row, col).fill = style
                elif row is not None:
                    for j in range(1, max_col + 1):
                        sh.cell(row, j).fill = style
                elif col is not None:
                    for i in range(1, max_row + 1):
                        sh.cell(i, col).fill = style

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置单元格背景色失败: {e}") from e

    def setRowColBackgroundColor(self, row, l_col, varColor=None, varSheet=0, clear_color=False):
        """
        3.15 设置单行多列背景色（优化版）

        :param row: 行号，必须是大于0的整数
        :param l_col: 列标识列表（如 [4, 6]）或字符串 "all"
        :param varColor: 背景色（十六进制字符串），不能为 None（除非 clear_color 为 True）
        :param varSheet: 工作表索引或名称，默认为0
        :param clear_color: 是否清除背景色，默认为 False
        :raises ValueError: 如果参数不合法或超出范围
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(row, int) or row <= 0:
            raise ValueError("row 必须是大于0的整数")
        if not isinstance(l_col, (list, str)) or (
                isinstance(l_col, list) and not all(isinstance(c, (int, str)) for c in l_col)):
            raise ValueError("l_col 必须是列表（元素为整数或字符串）或字符串 'all'")
        if not clear_color and (not isinstance(varColor, str) or len(varColor) != 6 or not all(
                c in '0123456789ABCDEFabcdef' for c in varColor)):
            raise ValueError("varColor 必须是6位十六进制字符串（除非 clear_color 为 True）")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(clear_color, bool):
            raise ValueError("clear_color 必须是布尔值")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 边界检查
            max_row = sh.max_row
            max_col = sh.max_column
            if row > max_row:
                raise ValueError(f"行号 {row} 超出范围，最大行数为 {max_row}")

            # 设置背景色样式
            if clear_color:
                style = PatternFill(fill_type=None)  # 清除背景色
            else:
                style = PatternFill("solid", fgColor=varColor)

            if l_col == "all":
                # 全部列应用背景色或清除背景色
                for col in range(1, max_col + 1):
                    sh.cell(row=row, column=col).fill = style
            else:
                # 指定列应用背景色或清除背景色
                col_indices = set()
                for col in l_col:
                    if isinstance(col, str):
                        col_indices.add(column_index_from_string(col))
                    elif isinstance(col, int):
                        col_indices.add(col)
                    else:
                        raise ValueError(f"列标识 '{col}' 类型不合法，应为整数或字符串")

                for col in sorted(col_indices):
                    if col > max_col:
                        raise ValueError(f"列号 {col} 超出范围，最大列数为 {max_col}")
                    sh.cell(row=row, column=col).fill = style

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置单行多列背景色失败: {e}") from e

    # def setRowColBackgroundColor(self, row, l_col, varColor, varSheet=0):
    #     """
    #     3.15 设置单行多列背景色（优化版）
    #
    #     :param row: 行号，必须是大于0的整数
    #     :param l_col: 列标识列表（如 [4, 6]）或字符串 "all"
    #     :param varColor: 背景色（十六进制字符串），不能为 None
    #     :param varSheet: 工作表索引或名称，默认为0
    #     :raises ValueError: 如果参数不合法或超出范围
    #     :raises IOError: 如果操作失败
    #     """
    #     # 参数校验
    #     if not isinstance(row, int) or row <= 0:
    #         raise ValueError("row 必须是大于0的整数")
    #     if not isinstance(l_col, (list, str)) or (
    #             isinstance(l_col, list) and not all(isinstance(c, (int, str)) for c in l_col)):
    #         raise ValueError("l_col 必须是列表（元素为整数或字符串）或字符串 'all'")
    #     if not isinstance(varColor, str) or len(varColor) != 6 or not all(
    #             c in '0123456789ABCDEFabcdef' for c in varColor):
    #         raise ValueError("varColor 必须是6位十六进制字符串")
    #     if not isinstance(varSheet, (int, str)):
    #         raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
    #
    #     try:
    #         # 获取工作表对象
    #         sh = self.sh(varSheet)
    #
    #         # 边界检查
    #         max_row = sh.max_row
    #         max_col = sh.max_column
    #         if row > max_row:
    #             raise ValueError(f"行号 {row} 超出范围，最大行数为 {max_row}")
    #
    #         # 设置背景色样式
    #         style = PatternFill("solid", fgColor=varColor)
    #
    #         if l_col == "all":
    #             # 全部列应用背景色
    #             for col in range(1, max_col + 1):
    #                 sh.cell(row=row, column=col).fill = style
    #         else:
    #             # 指定列应用背景色
    #             col_indices = set()
    #             for col in l_col:
    #                 if isinstance(col, str):
    #                     col_indices.add(column_index_from_string(col))
    #                 elif isinstance(col, int):
    #                     col_indices.add(col)
    #                 else:
    #                     raise ValueError(f"列标识 '{col}' 类型不合法，应为整数或字符串")
    #
    #             for col in sorted(col_indices):
    #                 if col > max_col:
    #                     raise ValueError(f"列号 {col} 超出范围，最大列数为 {max_col}")
    #                 sh.cell(row=row, column=col).fill = style
    #
    #         # 保存文件（仅在必要时调用）
    #         self.save()
    #
    #     except Exception as e:
    #         raise IOError(f"设置单行多列背景色失败: {e}") from e

    def setAllBackgroundColor(self, varColor=None, clear_color=False, varSheet=0):
        """
        3.16 设置所有单元格背景色（优化版）

        :param varColor: 背景色（十六进制字符串）或 None（清除背景色）
        :param varSheet: 工作表索引或名称，默认为0
        :raises ValueError: 如果参数不合法或超出范围
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if varColor is not None and (
                not isinstance(varColor, str) or
                len(varColor) != 6 or
                not all(c in '0123456789ABCDEFabcdef' for c in varColor)
        ):
            raise ValueError("varColor 必须是6位十六进制字符串或 None")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 获取最大行数和列数
            max_row = sh.max_row
            max_col = sh.max_column

            # 容错处理：如果工作表为空，则直接返回
            if max_row is None or max_col is None:
                print("[WARNING] 工作表为空，无需设置背景色。")
                return

            # 设置背景色样式
            if varColor is not None:
                style = PatternFill("solid", fgColor=varColor)

            if clear_color == True:
                style = PatternFill(fill_type=None)  # 清除背景色

            # 批量设置所有单元格背景色
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    sh.cell(row=r, column=c).fill = style

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置所有单元格背景色失败: {e}") from e

    def setBandRowsColor(self, row, varSkip, varColor=None, varSheet=0, clear_color=False):
        """
        3.17 设置整行(可间隔)背景色（优化版）

        :param row: 起始行号，必须是大于0的整数
        :param varSkip: 间隔行数，必须是非负整数
        :param varColor: 背景色（十六进制字符串），可以为 None（清除背景色）
        :param varSheet: 工作表索引或名称，默认为0
        :param clear_color: 是否清除背景色，默认为 False
        :raises ValueError: 如果参数不合法或超出范围
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(row, int) or row <= 0:
            raise ValueError("row 必须是大于0的整数")
        if not isinstance(varSkip, int) or varSkip < 0:
            raise ValueError("varSkip 必须是非负整数")
        if not clear_color and (varColor is None or not isinstance(varColor, str) or len(varColor) != 6 or not all(
                c in '0123456789ABCDEFabcdef' for c in varColor)):
            raise ValueError("varColor 必须是6位十六进制字符串（除非 clear_color 为 True）")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(clear_color, bool):
            raise ValueError("clear_color 必须是布尔值")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 边界检查
            max_row = sh.max_row
            max_col = sh.max_column
            if row > max_row:
                raise ValueError(f"行号 {row} 超出范围，最大行数为 {max_row}")

            # 设置背景色样式
            if clear_color:
                style = PatternFill(fill_type=None)  # 清除背景色
            else:
                style = PatternFill("solid", fgColor=varColor)

            # 批量设置指定行的背景色
            for i in range(row, max_row + 1, varSkip + 1):
                for j in range(1, max_col + 1):
                    sh.cell(i, j).fill = style

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置整行背景色失败: {e}") from e

    def setBandColsColor(self, col, varSkip, varColor=None, varSheet=0, clear_color=False):
        """
        3.18 设置整列(可间隔)背景色（优化版）

        :param col: 起始列号，必须是大于0的整数
        :param varSkip: 间隔列数，必须是非负整数
        :param varColor: 背景色（十六进制字符串），可以为 None（清除背景色）
        :param varSheet: 工作表索引或名称，默认为0
        :param clear_color: 是否清除背景色，默认为 False
        :raises ValueError: 如果参数不合法或超出范围
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(col, int) or col <= 0:
            raise ValueError("col 必须是大于0的整数")
        if not isinstance(varSkip, int) or varSkip < 0:
            raise ValueError("varSkip 必须是非负整数")
        if not clear_color and (varColor is None or not isinstance(varColor, str) or len(varColor) != 6 or not all(
                c in '0123456789ABCDEFabcdef' for c in varColor)):
            raise ValueError("varColor 必须是6位十六进制字符串（除非 clear_color 为 True）")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(clear_color, bool):
            raise ValueError("clear_color 必须是布尔值")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 边界检查
            max_row = sh.max_row
            max_col = sh.max_column
            if col > max_col:
                raise ValueError(f"列号 {col} 超出范围，最大列数为 {max_col}")

            # 设置背景色样式
            if clear_color:
                style = PatternFill(fill_type=None)  # 清除背景色
            else:
                style = PatternFill("solid", fgColor=varColor)

            # 批量设置指定列的背景色
            for i in range(1, max_row + 1):
                for j in range(col, max_col + 1, varSkip + 1):
                    sh.cell(i, j).fill = style

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置整列背景色失败: {e}") from e

    def setSheetColor(self, varColor=None, varSheet=0, clear_color=False):
        """
        3.19 设置工作表背景颜色（优化版）

        :param varColor: 背景色（十六进制字符串），可以为 None（清除背景色）
        :param varSheet: 工作表索引或名称，默认为0
        :param clear_color: 是否清除背景色，默认为 False
        :raises ValueError: 如果参数不合法或超出范围
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not clear_color and (varColor is None or not isinstance(varColor, str) or len(varColor) != 6 or not all(
                c in '0123456789ABCDEFabcdef' for c in varColor)):
            raise ValueError("varColor 必须是6位十六进制字符串（除非 clear_color 为 True）")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(clear_color, bool):
            raise ValueError("clear_color 必须是布尔值")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 设置工作表标签颜色
            if clear_color:
                sh.sheet_properties.tabColor = None  # 清除背景色
            else:
                sh.sheet_properties.tabColor = varColor

            # 保存文件（仅在必要时调用）
            self.save()

        except Exception as e:
            raise IOError(f"设置工作表背景颜色失败: {e}") from e


    # todo [获取]


    def getL_shape(self, varSheet=0):
        """
        3.1 获取总行列数（优化版）

        :param varSheet: 工作表索引或名称，默认为0
        :return: 总行数和总列数组成的列表，如 [4, 3]
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 获取总行数和总列数
            total_rows = sh.max_row
            total_cols = sh.max_column

            # 容错处理：如果工作表为空，返回默认值
            if total_rows is None or total_cols is None:
                return [0, 0]

            return [total_rows, total_cols]

        except Exception as e:
            raise IOError(f"获取总行列数失败: {e}") from e

    def getCell(self, varRow, varCol, varSheet=0):
        """
        3.2 获取单元格的值（优化版）

        :param varRow: 行号，必须是大于0的整数
        :param varCol: 列号，必须是大于0的整数
        :param varSheet: 工作表索引或名称，默认为0
        :return: 单元格的值，如 "hello"
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varRow, int) or varRow <= 0:
            raise ValueError("varRow 必须是大于0的整数")
        if not isinstance(varCol, int) or varCol <= 0:
            raise ValueError("varCol 必须是大于0的整数")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 边界检查
            max_row = sh.max_row
            max_col = sh.max_column
            if varRow > max_row:
                raise ValueError(f"行号 {varRow} 超出范围，最大行数为 {max_row}")
            if varCol > max_col:
                raise ValueError(f"列号 {varCol} 超出范围，最大列数为 {max_col}")

            # 获取单元格值
            return sh.cell(row=varRow, column=varCol).value

        except Exception as e:
            raise IOError(f"获取单元格值失败: {e}") from e

    def getL_row(self, varSeq, varSheet=0):
        """
        3.3 获取一行数据（优化版）

        :param varSeq: 行号，必须是大于0的整数
        :param varSheet: 工作表索引或名称，默认为0
        :return: 指定行的数据列表，如 [10, 5, 10]
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSeq, int) or varSeq <= 0:
            raise ValueError("varSeq 必须是大于0的整数")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 检查行号是否超出范围
            total_rows = sh.max_row
            if varSeq > total_rows:
                raise ValueError(f"行号 {varSeq} 超出范围，最大行数为 {total_rows}")

            # 获取指定行数据
            target_row = list(sh.rows)[varSeq - 1]
            l_row = [cell.value for cell in target_row]

            return l_row

        except Exception as e:
            raise IOError(f"获取行数据失败: {e}") from e

    def getL_col(self, varCol, varSheet=0, include_header=True):
        """
        3.4 获取一列的值（优化版）

        :param varCol: 列标识，可以是整数（如2）或字符串（如'B'）
        :param varSheet: 工作表索引或名称，默认为0
        :param include_header: 是否包含标题行，默认为True
        :return: 指定列的数据列表，如 ['山丘', 30, 25, 30, 10, 5, 10]
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varCol, (int, str)):
            raise ValueError("varCol 必须是整数或字符串")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(include_header, bool):
            raise ValueError("include_header 必须是布尔值")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 转换列标识为索引（如果是字符串）
            if isinstance(varCol, str):
                varCol = column_index_from_string(varCol)

            # 获取列数据
            col_data = [cell.value for cell in list(sh.columns)[varCol - 1]]

            # 根据 include_header 参数决定是否包含标题行
            if not include_header:
                col_data = col_data[1:]  # 跳过第一行（标题行）

            return col_data

        except Exception as e:
            raise IOError(f"获取列数据失败: {e}") from e

    def getLL_row(self, varSheet=0, include_header=True):
        """
        3.5.1 获取每行数据（优化版）

        :param varSheet: 工作表索引或名称，默认为0
        :param include_header: 是否包含标题行，默认为True
        :return: 每行数据组成的二维列表，如 [['Number具体数', '高地', 'jinhaoyoyo'], [2, 40, 30], ...]
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(include_header, bool):
            raise ValueError("include_header 必须是布尔值")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 根据 include_header 参数决定是否包含标题行
            start_index = 0 if include_header else 1

            # 构建每行数据
            l_l_row = [
                [cell.value for cell in row] for row in list(sh.rows)[start_index:]
            ]

            return l_l_row

        except Exception as e:
            raise IOError(f"获取行数据失败: {e}") from e

    def getD_rowNumber_row(self, varSheet=0, include_header=True):
        """
        3.5.2 获取带行号的每行数据（优化版）

        :param varSheet: 工作表索引或名称，默认为0
        :param include_header: 是否包含标题行，默认为True
        :return: 行号为键、行数据为值的字典，如 {1: ['Number具体数', '高地', 'jinhaoyoyo'], 2: [2, 40, 30], ...}
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(include_header, bool):
            raise ValueError("include_header 必须是布尔值")

        try:
            # 获取所有行数据
            l_l_row = self.getLL_row(varSheet)
            if not l_l_row:
                raise ValueError("工作表中没有数据")

            # 根据 include_header 参数决定是否包含标题行
            start_index = 0 if include_header else 1

            # 构建行号到行数据的映射
            d_seq_row = {
                i + 1: l_l_row[i] for i in range(start_index, len(l_l_row))
            }

            return d_seq_row

        except Exception as e:
            raise IOError(f"获取行数据失败: {e}") from e

    def getLL_rowOfPartialCol(self, l_col, varSheet=0):
        """
        3.5.3 获取部分列的行数据（优化版）

        :param l_col: 需要获取的列标识列表，如 [1, 3] 或 ['A', 'C']
        :param varSheet: 工作表索引或名称，默认为0
        :return: 每行指定列的数据组成的二维列表，如 [['Number具体数', 'jinhaoyoyo'], [2, 30], ...]
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(l_col, list) or not l_col:
            raise ValueError("l_col 必须是一个非空列表")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 转换列标识为索引（如果是字符串）
            col_indices = []
            for col in l_col:
                if isinstance(col, str):
                    col_indices.append(column_index_from_string(col))
                elif isinstance(col, int):
                    col_indices.append(col)
                else:
                    raise ValueError(f"列标识 '{col}' 类型不合法，应为整数或字符串")

            # 去重（保留首次出现的值）
            col_indices = sorted(set(col_indices), key=col_indices.index)

            # 构建每行指定列的数据
            l_l_row = [
                [sh.cell(row, col).value for col in col_indices]
                for row in range(1, sh.max_row + 1)
            ]

            return l_l_row

        except Exception as e:
            raise IOError(f"获取行数据失败: {e}") from e

    def getD_rowNumber_rowOfpartialCol(self, l_col, varSheet=0):
        """
        3.5.4 获取带行号的部分列的行数据（优化版）

        :param l_col: 需要获取的列标识列表，如 [1, 3] 或 ['A', 'C']
        :param varSheet: 工作表索引或名称，默认为0
        :return: 行号为键、行数据为值的字典，如 {1: ['Number具体数', 'jinhaoyoyo'], 2: [2, 30], ...}
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(l_col, list) or not l_col:
            raise ValueError("l_col 必须是一个非空列表")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取部分列的行数据
            l_l_col = self.getLL_rowOfPartialCol(l_col, varSheet)
            if not l_l_col:
                raise ValueError("工作表中没有数据")

            # 构建行号到行数据的映射
            d_seq_row = {i + 1: l_l_col[i] for i in range(len(l_l_col))}

            return d_seq_row

        except Exception as e:
            raise IOError(f"获取行数据失败: {e}") from e

    def getLL_col(self, varSheet=0, include_header=True):
        """
        3.6.1 获取每列数据（优化版）

        :param varSheet: 工作表索引或名称，默认为0
        :param include_header: 是否包含标题行，默认为True
        :return: 每列数据组成的二维列表，如 [['age', 2, 12], ['city', 'shanghai', 'beijin']]
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(include_header, bool):
            raise ValueError("include_header 必须是布尔值")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 根据 include_header 参数决定是否包含标题行
            start_index = 0 if include_header else 1

            # 构建每列数据
            l_l_col = [
                [cell.value for cell in col][start_index:] for col in sh.columns
            ]

            return l_l_col

        except Exception as e:
            raise IOError(f"获取列数据失败: {e}") from e

    def getD_colNumber_col(self, varSheet=0, include_header=True):
        """
        3.6.2 获取带列序号的每列数据（优化版）

        :param varSheet: 工作表索引或名称，默认为0
        :param include_header: 是否包含标题行，默认为True
        :return: 列序号为键、列数据为值的字典，如 {1: ['age', 2, 12], 2: ['city', 'shanghai', 'beijin']}
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(include_header, bool):
            raise ValueError("include_header 必须是布尔值")

        try:
            # 获取所有列数据
            l_l_col = self.getLL_col(varSheet)
            if not l_l_col:
                raise ValueError("工作表中没有数据")

            # 根据 include_header 参数决定是否包含标题行
            start_index = 0 if include_header else 1

            # 构建列序号到列数据的映射
            d_seq_col = {
                i + 1: l_l_col[i][start_index:] for i in range(len(l_l_col))
            }

            return d_seq_col

        except Exception as e:
            raise IOError(f"获取列数据失败: {e}") from e

    def getD_colLetter_col(self, varSheet=0, include_header=True):
        """
        3.6.3 获取带列字母的每列数据（优化版）

        :param varSheet: 工作表索引或名称，默认为0
        :param include_header: 是否包含标题行，默认为True
        :return: 列字母为键、列数据为值的字典，如 {'A': [1, 2, 3], 'B': ['a', 'b', 'c']}
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")
        if not isinstance(include_header, bool):
            raise ValueError("include_header 必须是布尔值")

        try:
            # 获取所有列数据
            l_l_col = self.getLL_col(varSheet)
            if not l_l_col:
                raise ValueError("工作表中没有数据")

            # 根据 include_header 参数决定是否包含标题行
            start_index = 0 if include_header else 1

            # 构建列字母到列数据的映射
            d_seq_col = {
                get_column_letter(i + 1): l_l_col[i][start_index:] for i in range(len(l_l_col))
            }

            return d_seq_col

        except Exception as e:
            raise IOError(f"获取列数据失败: {e}") from e


    def getL_columnHeaderNumber(self, l_partialTitle, varSheet=0):
        """
        3.8.1 获取列标签的列标number（优化版）

        :param l_partialTitle: 部分标题列表，如 ['姓名', '性别']
        :param varSheet: 工作表索引或名称，默认为0
        :return: 标题对应的列序号列表，如 [1, 3]
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(l_partialTitle, list):
            raise ValueError("l_partialTitle 必须是一个列表")
        if not all(isinstance(item, str) for item in l_partialTitle):
            raise ValueError("l_partialTitle 中的所有元素必须是字符串")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取第一行标题
            l_title = self.getL_row(1, varSheet)
            if not l_title:
                raise ValueError("工作表中没有找到标题行")

            # 构建标题到列索引的映射（使用字典加速查找）
            title_to_index = {title: idx for idx, title in enumerate(l_title)}

            # 查找匹配的标题并转换为列序号
            result = []
            for title in l_partialTitle:
                if title in title_to_index:
                    col_index = title_to_index[title] + 1  # 列索引从1开始
                    result.append(col_index)
                else:
                    # 如果标题未找到，可以选择跳过或抛出警告
                    print(f"[Warning] 标题 '{title}' 未在表格中找到")

            return result

        except Exception as e:
            raise IOError(f"获取标题列序号失败: {e}") from e

    def getL_columnHeaderLetter(self, l_partialTitle, varSheet=0):
        """
        3.8.2 获取列标签的列标letter（优化版）

        :param l_partialTitle: 部分标题列表，如 ['姓名', '性别']
        :param varSheet: 工作表索引或名称，默认为0
        :return: 标题对应的列字母列表，如 ['A', 'C']
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(l_partialTitle, list):
            raise ValueError("l_partialTitle 必须是一个列表")
        if not all(isinstance(item, str) for item in l_partialTitle):
            raise ValueError("l_partialTitle 中的所有元素必须是字符串")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取第一行标题
            l_title = self.getL_row(1, varSheet)
            if not l_title:
                raise ValueError("工作表中没有找到标题行")

            # 构建标题到列索引的映射（使用字典加速查找）
            title_to_index = {title: idx for idx, title in enumerate(l_title)}

            # 查找匹配的标题并转换为列字母
            result = []
            for title in l_partialTitle:
                if title in title_to_index:
                    col_index = title_to_index[title] + 1  # 列索引从1开始
                    result.append(get_column_letter(col_index))
                else:
                    # 如果标题未找到，可以选择跳过或抛出警告
                    print(f"[Warning] 标题 '{title}' 未在表格中找到")

            return result

        except Exception as e:
            raise IOError(f"获取标题列字母失败: {e}") from e

    def getD_colNumber_columnTitle(self, l_partialTitle, varSheet=0):
        """
        3.8.3 将标题转列字典序列（优化版）

        :param l_partialTitle: 部分标题列表，如 ['姓名', '性别']
        :param varSheet: 工作表索引或名称，默认为0
        :return: 标题与列序号的字典，如 {1: '姓名', 2: '性别'}
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(l_partialTitle, list):
            raise ValueError("l_partialTitle 必须是一个列表")
        if not all(isinstance(item, str) for item in l_partialTitle):
            raise ValueError("l_partialTitle 中的所有元素必须是字符串")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取第一行标题
            l_title = self.getL_row(1, varSheet)
            if not l_title:
                raise ValueError("工作表中没有找到标题行")

            # 构建标题到列序号的映射
            d_title_to_seq = {}
            title_set = set(l_title)  # 使用集合加速查找
            for title in l_partialTitle:
                if title in title_set:
                    index = l_title.index(title) + 1
                    d_title_to_seq[index] = title

            return d_title_to_seq

        except Exception as e:
            raise IOError(f"处理标题转列字典序列失败: {e}") from e

    def getD_colLetter_columnTitle(self, l_partialTitle, varSheet=0):
        """
        3.8.4 将标题转列字典字母（优化版）

        :param l_partialTitle: 部分标题列表，如 ['姓名', '性别']
        :param varSheet: 工作表索引或名称，默认为0
        :return: 标题与列字母的字典，如 {'B': '姓名', 'E': '性别'}
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(l_partialTitle, list):
            raise ValueError("l_partialTitle 必须是一个列表")
        if not all(isinstance(item, str) for item in l_partialTitle):
            raise ValueError("l_partialTitle 中的所有元素必须是字符串")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取第一行标题
            l_title = self.getL_row(1, varSheet)
            if not l_title:
                raise ValueError("工作表中没有找到标题行")

            # 构建标题到列字母的映射
            d_title_to_letter = {}
            for title in l_partialTitle:
                if title in l_title:
                    index = l_title.index(title) + 1
                    letter = get_column_letter(index)
                    d_title_to_letter[letter] = title

            return d_title_to_letter

        except Exception as e:
            raise IOError(f"处理标题转列字典字母失败: {e}") from e

    def getLL_partialColOfPartialCol(self, l_getCol, l_ignoreRow, varSheet=0):
        """
        3.9 获取部分列的列值（可忽略多行）（优化版）

        :param l_getCol: 需要获取的列标识列表，如 [1, 3] 或 ['A', 'C']
        :param l_ignoreRow: 需要忽略的行号列表，如 [1, 2]
        :param varSheet: 工作表索引或名称，默认为0
        :return: 每列的值组成的二维列表，如 [[10, 20], [30, 40]]
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(l_getCol, list):
            raise ValueError("l_getCol 必须是一个列表")
        if not isinstance(l_ignoreRow, list):
            raise ValueError("l_ignoreRow 必须是一个列表")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 转换列标识为索引（如果是字符串）
            col_indices = []
            for col in l_getCol:
                if isinstance(col, str):
                    col_indices.append(column_index_from_string(col))
                elif isinstance(col, int):
                    col_indices.append(col)
                else:
                    raise ValueError(f"列标识 '{col}' 类型不合法，应为整数或字符串")

            # 转换忽略行号为集合（提升查找效率）
            ignore_rows = set(l_ignoreRow)

            # 初始化结果容器
            result = []

            # 遍历每一列
            for col_idx in col_indices:
                col_values = []
                # 遍历每一行
                for row in range(1, sh.max_row + 1):
                    if row not in ignore_rows:
                        col_values.append(sh.cell(row, col_idx).value)
                result.append(col_values)

            return result

        except Exception as e:
            raise IOError(f"获取列值失败: {e}") from e


    def getCoordinate(self, varRow, varCol, varSheet=0):
        """
        3.10 获取单元格的坐标（优化版）

        :param varRow: 行号，必须是大于0的整数
        :param varCol: 列号，必须是大于0的整数
        :param varSheet: 工作表索引或名称，默认为0
        :return: 单元格的坐标字符串，如 'A1'
        :raises ValueError: 如果参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varRow, int) or varRow <= 0:
            raise ValueError("varRow 必须是大于0的整数")
        if not isinstance(varCol, int) or varCol <= 0:
            raise ValueError("varCol 必须是大于0的整数")
        if not isinstance(varSheet, (int, str)):
            raise ValueError("varSheet 必须是整数（索引）或字符串（名称）")

        try:
            # 获取工作表对象
            sh = self.sh(varSheet)

            # 获取单元格坐标
            return sh.cell(row=varRow, column=varCol).coordinate

        except Exception as e:
            raise IOError(f"获取单元格坐标失败: {e}") from e

    def getDimensions(self, varSheet=0):

        # 3.11 获取所有数据的坐标
        sh = self.sh(varSheet)
        return sh.dimensions


    # todo [两表比较]

    def getD_excel_cell_By_Diff(self, l_file1row, l_file2row):
        """
        getD_excel_cell_By_Diff
        5.2 两个excel的sheet进行比较，输出有差异值。（两表标题与行数必须一致） （优化版）
        
        # print(Openpyxl_PO.getD_excel_cell_By_Diff(Openpyxl_PO.getLL_row("Sheet2"), Openpyxl_PO2.getLL_row("Sheet2")))

        :param l_file1row: 第一个文件的行数据列表
        :param l_file2row: 第二个文件的行数据列表
        :return: 包含差异内容的字典，格式为 {"left": {...}, "right": {...}}；若无差异则返回 None
        :raises ValueError: 如果输入参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(l_file1row, list) or not isinstance(l_file2row, list):
            raise ValueError("输入参数必须是列表类型")

        try:
            # 初始化结果容器
            d_left = {}
            d_right = {}
            d_all = {}

            # 检查行数是否一致
            if len(l_file1row) != len(l_file2row):
                raise ValueError("两个文件的行数不一致，无法比较")

            # 遍历每一行，查找差异
            for i in range(len(l_file1row)):
                if l_file1row[i] != l_file2row[i]:
                    d_left_sub = {}
                    d_right_sub = {}
                    for j in range(len(l_file1row[i])):
                        if l_file1row[i][j] != l_file2row[i][j]:
                            # 记录差异字段及其值
                            d_left_sub[l_file1row[0][j]] = l_file1row[i][j]
                            d_right_sub[l_file2row[0][j]] = l_file2row[i][j]
                    if d_left_sub or d_right_sub:
                        d_left[i + 1] = d_left_sub
                        d_right[i + 1] = d_right_sub

            # 构造最终结果
            if d_left or d_right:
                d_all["第一个excel"] = d_left
                d_all["第二个excel"] = d_right
                return d_all
            else:
                return None  # 无差异时返回 None

        except Exception as e:
            raise IOError(f"比较文件时发生错误: {e}") from e

    def setColorByDiff(self, varSheet1, varSheet2, color1="FF0000", color2="ffeb9c", skip_empty=True):
        """
        5.2 两工作表比较，对差异内容标注颜色（优化版）

        :param varSheet1: 第一张工作表名称或索引
        :param varSheet2: 第二张工作表名称或索引
        :param color1: Sheet1 中差异单元格的颜色（默认红色 FF0000）
        :param color2: Sheet2 中差异单元格的颜色（默认橙色 ffeb9c）
        :param skip_empty: 是否跳过空值比较（默认 True）
        :raises ValueError: 如果参数不合法或工作表不存在
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if not isinstance(varSheet1, (int, str)):
            raise ValueError("varSheet1 必须是整数（索引）或字符串（名称）")
        if not isinstance(varSheet2, (int, str)):
            raise ValueError("varSheet2 必须是整数（索引）或字符串（名称）")
        if not isinstance(color1, str) or len(color1) != 6 or not all(c in '0123456789ABCDEFabcdef' for c in color1):
            raise ValueError("color1 必须是6位十六进制字符串")
        if not isinstance(color2, str) or len(color2) != 6 or not all(c in '0123456789ABCDEFabcdef' for c in color2):
            raise ValueError("color2 必须是6位十六进制字符串")
        if not isinstance(skip_empty, bool):
            raise ValueError("skip_empty 必须是布尔值")

        try:
            # 获取两张表的所有行数据
            l_sheetOneRow = self.getLL_row(varSheet1)
            l_sheetTwoRow = self.getLL_row(varSheet2)

            # 校验工作表是否存在
            if l_sheetOneRow is None or l_sheetTwoRow is None:
                raise ValueError(f"工作表 '{varSheet1}' 或 '{varSheet2}' 不存在！")

            # 校验行列数是否一致
            if len(l_sheetOneRow) != len(l_sheetTwoRow):
                raise ValueError("两张表的行数不一致，无法比较！")
            if any(len(row1) != len(row2) for row1, row2 in zip(l_sheetOneRow, l_sheetTwoRow)):
                raise ValueError("两张表的列数不一致，无法比较！")

            # 遍历每一行和每一列，标记差异
            for i in range(len(l_sheetOneRow)):
                row1 = l_sheetOneRow[i]
                row2 = l_sheetTwoRow[i]
                for j in range(len(row1)):
                    val1 = row1[j]
                    val2 = row2[j]

                    # 跳过空值（如果启用 skip_empty）
                    if skip_empty and (val1 is None or val1 == "" or val2 is None or val2 == ""):
                        continue

                    # 统一类型后比较值
                    if str(val1) != str(val2):  # 强制转换为字符串比较
                        # 设置 Sheet1 的背景色
                        self.switchSheet(varSheet1)  # 显式切换到 Sheet1
                        self.setBackgroundColor(i + 1, j + 1, color1, varSheet1)
                        print(f"[DEBUG] Sheet1 差异单元格 ({i + 1}, {j + 1}): '{val1}'")

                        # 设置 Sheet2 的背景色
                        self.switchSheet(varSheet2)  # 显式切换到 Sheet2
                        self.setBackgroundColor(i + 1, j + 1, color2, varSheet2)
                        print(f"[DEBUG] Sheet2 差异单元格 ({i + 1}, {j + 1}): '{val2}'")

                print(f"[INFO] 已完成第 {i + 1} 行的比较")

            # 保存文件
            self.save()

        except Exception as e:
            raise IOError(f"比较工作表失败: {e}") from e

    def genSheetByDiff(self, varSheet1, varSheet2):
        """
        5.3 比较两工作表，对差异内容标注颜色，生成新表Sheet1_Sheet2（优化版）
        支持灵活应对标题位置不一致的问题，并保留所有数据（包括无差异的值）

        :param varSheet1: 第一个工作表名称
        :param varSheet2: 第二个工作表名称
        :return: 生成的新工作表名称
        :raises ValueError: 如果输入参数不合法
        :raises IOError: 如果操作失败
        """
        # 参数校验
        if varSheet1 not in self.wb.sheetnames:
            raise ValueError(f"工作表 '{varSheet1}' 不存在于当前文件中。可用的工作表包括：{self.wb.sheetnames}")
        if varSheet2 not in self.wb.sheetnames:
            raise ValueError(f"工作表 '{varSheet2}' 不存在于当前文件中。可用的工作表包括：{self.wb.sheetnames}")

        try:
            # 生成安全的工作表名称（替换非法字符）
            diff_sheet_name = f"{varSheet1}_{varSheet2}_diff"

            # 如果该工作表已存在，则先删除
            if diff_sheet_name in self.wb.sheetnames:
                self.delSheet(diff_sheet_name)

            # 创建新工作表用于存储比对结果
            new_sheet = self.wb.create_sheet(title=diff_sheet_name)

            # 获取两个工作表的数据
            data1 = self.getLL_row(varSheet1)
            data2 = self.getLL_row(varSheet2)

            # 获取最大行数和列数（以较大的为准）
            max_rows = max(len(data1), len(data2))
            max_cols = max(
                max(len(row) for row in data1) if data1 else 0,
                max(len(row) for row in data2) if data2 else 0
            )

            # 执行比对逻辑并填充新表
            for i in range(max_rows):
                for j in range(max_cols):
                    # 获取当前单元格的值（如果超出范围则设为 None）
                    val1 = data1[i][j] if i < len(data1) and j < len(data1[i]) else None
                    val2 = data2[i][j] if i < len(data2) and j < len(data2[i]) else None

                    # 填充新表
                    if val1 == val2:
                        # 无差异：直接复制值
                        new_sheet.cell(row=i + 1, column=j + 1).value = val1
                    else:
                        # 有差异：标注差异并设置背景色
                        new_sheet.cell(row=i + 1, column=j + 1).value = f"{val1} vs {val2}"
                        new_sheet.cell(row=i + 1, column=j + 1).fill = PatternFill(
                            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                        )

            # 保存文件
            self.save()

            # 返回新工作表名称
            return diff_sheet_name

        except Exception as e:
            raise IOError(f"处理工作表失败: {e}") from e

