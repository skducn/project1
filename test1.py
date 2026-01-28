import os
import re
import chardet
import openpyxl
import docx
import PyPDF2
import pandas as pd

import warnings
warnings.filterwarnings('ignore', message='Workbook contains no default style, apply openpyxl\'s default')


# 全局配置
DEFAULT_CONTEXT_ROWS = 3  # 上下文默认展示行数
IGNORE_CASE = True  # 默认忽略大小写


class DocSearchAgent:
    def __init__(self):
        self.file_parsed_data = {}  # 存储解析后的文件数据：{文件路径: [(位置元数据, 内容), ...]}

    def _detect_encoding(self, file_path):
        """检测文件编码，解决TXT/MD乱码问题"""
        with open(file_path, 'rb') as f:
            result = chardet.detect(f.read(1024))
        return result['encoding'] or 'utf-8'

    def _parse_txt_md(self, file_path):
        """解析TXT/MD文件"""
        encoding = self._detect_encoding(file_path)
        parsed_data = []
        with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
            lines = f.readlines()
            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if line:  # 忽略空行
                    parsed_data.append((f"行号：{line_num}", line))
        self.file_parsed_data[file_path] = parsed_data

    def _parse_excel(self, file_path):
        """解析Excel文件（xlsx/xls）"""
        parsed_data = []
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
                    for col_num, cell_value in enumerate(row, 1):
                        cell_value = str(cell_value).strip() if cell_value is not None else ""
                        if cell_value:  # 忽略空单元格
                            pos = f"工作表：{sheet_name} | 行号：{row_num} | 单元格：{chr(64 + col_num)}{row_num}"
                            parsed_data.append((pos, cell_value))
            wb.close()
        except:
            # 兼容xls格式
            df = pd.read_excel(file_path, sheet_name=None)
            for sheet_name, data in df.items():
                for row_num, row in data.iterrows():
                    row_num += 2  # 对应Excel实际行号（跳过表头）
                    for col_num, cell_value in enumerate(row, 1):
                        cell_value = str(cell_value).strip() if cell_value is not None else ""
                        if cell_value:
                            pos = f"工作表：{sheet_name} | 行号：{row_num} | 单元格：{chr(64 + col_num)}{row_num}"
                            parsed_data.append((pos, cell_value))
        self.file_parsed_data[file_path] = parsed_data

    def _parse_csv(self, file_path):
        """解析Csv文件"""
        parsed_data = []
        encoding = self._detect_encoding(file_path)
        df = pd.read_csv(file_path, encoding=encoding, errors='ignore')
        for row_num, row in df.iterrows():
            row_num += 2  # 对应实际行号
            for col_num, cell_value in enumerate(row, 1):
                cell_value = str(cell_value).strip() if cell_value is not None else ""
                if cell_value:
                    pos = f"行号：{row_num} | 单元格：{chr(64 + col_num)}{row_num}"
                    parsed_data.append((pos, cell_value))
        self.file_parsed_data[file_path] = parsed_data

    def _parse_pdf(self, file_path):
        """解析PDF文件"""
        parsed_data = []
        with open(file_path, 'rb') as f:
            pdf_reader = PyPDF2.PdfReader(f)
            for page_num, page in enumerate(pdf_reader.pages, 1):
                text = page.extract_text()
                if not text:
                    continue
                lines = text.split('\n')
                for line_num, line in enumerate(lines, 1):
                    line = line.strip()
                    if line:
                        pos = f"页码：{page_num} | 行号：{line_num}"
                        parsed_data.append((pos, line))
        self.file_parsed_data[file_path] = parsed_data

    def _parse_docx(self, file_path):
        """解析Word文件（docx）"""
        parsed_data = []
        doc = docx.Document(file_path)
        for para_num, para in enumerate(doc.paragraphs, 1):
            text = para.text.strip()
            if not text:
                continue
            lines = text.split('\n')
            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if line:
                    pos = f"段落号：{para_num} | 行号：{line_num}"
                    parsed_data.append((pos, line))
        self.file_parsed_data[file_path] = parsed_data

    def parse_file(self, file_path):
        """统一解析入口，自动识别格式"""
        file_suffix = os.path.splitext(file_path)[1].lower()
        if file_suffix in ['.txt', '.md']:
            self._parse_txt_md(file_path)
        elif file_suffix in ['.xlsx', '.xls']:
            self._parse_excel(file_path)
        elif file_suffix == '.csv':
            self._parse_csv(file_path)
        elif file_suffix == '.pdf':
            self._parse_pdf(file_path)
        elif file_suffix == '.docx':
            self._parse_docx(file_path)
        else:
            raise ValueError(f"不支持的文件格式：{file_suffix}")
        print(f"✅ 解析完成：{file_path}")

    def parse_folder(self, folder_path, filter_suffix=None):
        """解析文件夹下所有文件，可指定格式过滤"""
        if filter_suffix is None:
            filter_suffix = ['.txt', '.md', '.xlsx', '.xls', '.csv', '.pdf', '.docx']
        for root, _, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                file_suffix = os.path.splitext(file_path)[1].lower()
                if file_suffix in filter_suffix:
                    try:
                        self.parse_file(file_path)
                    except Exception as e:
                        print(f"❌ 解析失败：{file_path}，错误：{str(e)[:50]}")

    def _get_search_pattern(self, keywords, search_mode):
        """生成检索正则表达式"""
        if IGNORE_CASE:
            flags = re.IGNORECASE
        else:
            flags = 0
        # 关键字预处理
        keywords = [k.strip() for k in keywords if k.strip()]
        if not keywords:
            raise ValueError("请输入有效关键字")

        if search_mode == '精准':
            # 精准匹配：整词匹配
            pattern = r'\b(' + '|'.join(re.escape(k) for k in keywords) + r')\b'
        elif search_mode == '模糊':
            # 模糊匹配：包含关键字即可
            pattern = '|'.join(re.escape(k) for k in keywords)
        elif search_mode == '组合':
            # 组合匹配：&为与，|为或，先处理&再处理|
            keyword_str = ' '.join(keywords)
            # 替换&为正则的正向预查
            keyword_str = keyword_str.replace('&', r')(?=.*')
            pattern = r'(?=.*' + keyword_str + r').*'
        else:
            raise ValueError("检索模式仅支持：精准/模糊/组合")
        return re.compile(pattern, flags=flags)

    def _highlight_keyword(self, text, keywords):
        """高亮关键字（CLI版：红色）"""
        if IGNORE_CASE:
            keywords = [k.lower() for k in keywords]
            text_lower = text.lower()
        else:
            text_lower = text
        # 遍历关键字，高亮
        for k in keywords:
            start = 0
            while True:
                pos = text_lower.find(k, start)
                if pos == -1:
                    break
                # ANSI红色高亮：\033[31m关键字\033[0m
                text = text[:pos] + '\033[31m' + text[pos:pos + len(k)] + '\033[0m'
                text_lower = text_lower[:pos] + '\033[31m' + text_lower[pos:pos + len(k)] + '\033[0m'
                start = pos + len(k) + 9  # 加上高亮字符长度
        return text

    def search(self, keywords_input, search_mode='模糊'):
        """核心检索入口"""
        # 解析输入的关键字
        if '&' in keywords_input or '|' in keywords_input:
            keywords = [keywords_input]
        else:
            keywords = re.split(r'[,，\s]+', keywords_input)
        # 生成检索模式
        pattern = self._get_search_pattern(keywords, search_mode)
        # 开始检索
        result_count = 0
        print(f"\n========== 检索结果（模式：{search_mode}，关键字：{keywords_input}）==========")
        for file_path, parsed_data in self.file_parsed_data.items():
            file_matched = False
            for pos, content in parsed_data:
                if pattern.search(content):
                    if not file_matched:
                        print(f"\n【匹配文件】：{file_path}")
                        file_matched = True
                    result_count += 1
                    # 高亮关键字
                    content_highlight = self._highlight_keyword(content, [k for k in keywords if k.strip()])
                    print(f"【定位信息】：{pos}")
                    print(f"【匹配内容】：{content_highlight}\n")
        if result_count == 0:
            print("❌ 未找到匹配内容")
        else:
            print(f"✅ 检索完成，共找到{result_count}条匹配结果")
        return result_count


# 命令行交互入口
def main():
    agent = DocSearchAgent()
    print("===== 文档关键字定位智能体 =====")
    print("支持格式：TXT/MD/Excel(xlsx/xls)/Csv/PDF/Word(docx)")
    print("检索模式：精准/模糊/组合（组合使用&与|或，例：黄金&5000美元）\n")
    # 选择检索范围
    while True:
        scope_choice = input("请选择检索范围（1-单个文件，2-整个文件夹）：")
        if scope_choice in ['1', '2']:
            break
        print("输入错误，请选择1或2")
    # 输入路径
    if scope_choice == '1':
        file_path = input("请输入文件绝对路径：").strip('"').strip("'")
        if not os.path.isfile(file_path):
            print("文件路径不存在！")
            return
        agent.parse_file(file_path)
    else:
        folder_path = input("请输入文件夹绝对路径：").strip('"').strip("'")
        if not os.path.isdir(folder_path):
            print("文件夹路径不存在！")
            return
        agent.parse_folder(folder_path)
    # 选择检索模式
    while True:
        search_mode = input("请选择检索模式（精准/模糊/组合）：").strip()
        if search_mode in ['精准', '模糊', '组合']:
            break
        print("输入错误，请选择精准/模糊/组合")
    # 输入关键字并检索
    keywords_input = input("请输入检索关键字：").strip()
    agent.search(keywords_input, search_mode)


if __name__ == "__main__":
    main()